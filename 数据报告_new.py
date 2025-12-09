import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import pandas as pd
import chardet
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import numpy as np
from datetime import datetime
import traceback
import sys
import linecache
from pypinyin import pinyin, Style

# 土壤属性配置（参考土壤属性多文件分析.py）
SOIL_ATTR_CONFIG = {
    'OM': {
        'name': '有机质',
        'unit': 'g/kg',
        'reverse_display': True,
        'levels': [
            (10, '5级', '低'),
            (20, '4级', '较低'),
            (30, '3级', '中'),
            (40, '2级', '较高'),
            (float('inf'), '1级', '高')
        ]
    },
    'TN': {
        'name': '全氮',
        'unit': 'g/kg',
        'reverse_display': True,
        'levels': [
            (0.5, '5级', '极缺'),
            (1.0, '4级', '缺乏'),
            (1.5, '3级', '中等'),
            (2.0, '2级', '较丰富'),
            (float('inf'), '1级', '丰富')
        ]
    },
    'TP': {
        'name': '全磷',
        'unit': 'g/kg',
        'reverse_display': True,
        'levels': [
            (0.4, '5级', '极缺'),
            (0.6, '4级', '缺乏'),
            (0.8, '3级', '中等'),
            (1.0, '2级', '较丰富'),
            (float('inf'), '1级', '丰富')
        ]
    },
    'TK': {
        'name': '全钾',
        'unit': 'g/kg',
        'reverse_display': True,
        'levels': [
            (10, '5级', '极缺'),
            (15, '4级', '缺乏'),
            (20, '3级', '中等'),
            (25, '2级', '较丰富'),
            (float('inf'), '1级', '丰富')
        ]
    },
    'AP': {
        'name': '有效磷',
        'unit': 'mg/kg',
        'reverse_display': True,
        'levels': [
            (5, '5级', '极缺'),
            (10, '4级', '缺乏'),
            (20, '3级', '中等'),
            (40, '2级', '较丰富'),
            (float('inf'), '1级', '丰富')
        ]
    },
    'AK': {
        'name': '速效钾',
        'unit': 'mg/kg',
        'reverse_display': True,
        'levels': [
            (50, '5级', '极缺'),
            (100, '4级', '缺乏'),
            (150, '3级', '中等'),
            (200, '2级', '较丰富'),
            (float('inf'), '1级', '丰富')
        ]
    },
    'CEC': {
        'name': '阳离子交换量',
        'unit': 'cmol(+)/kg',
        'reverse_display': True,
        'levels': [
            (5, '5级', '低'),
            (10, '4级', '较低'),
            (15, '3级', '中'),
            (20, '2级', '较高'),
            (float('inf'), '1级', '高')
        ]
    },
    'ph': {
        'name': 'pH',
        'unit': '',
        'reverse_display': False,
        'levels': [
            (4.5, '1级', '强酸性'),
            (5.5, '2级', '酸性'),
            (6.5, '3级', '弱酸性'),
            (7.5, '4级', '中性'),
            (8.5, '5级', '弱碱性'),
            (9.0, '6级', '碱性'),
            (14.0, '7级', '强碱性')
        ]
    },
    # 特殊用地限制属性（只统计耕园地）
    'AFE': {
        'name': '有效铁',
        'unit': 'mg/kg',
        'reverse_display': True,
        'land_filter': 'cultivated_garden',
        'levels': [
            (2.5, '5级', '极缺'),
            (4.5, '4级', '缺乏'),
            (10.0, '3级', '中等'),
            (20.0, '2级', '丰富'),
            (float('inf'), '1级', '偏高')
        ]
    },
    'AMN': {
        'name': '有效锰',
        'unit': 'mg/kg',
        'reverse_display': True,
        'land_filter': 'cultivated_garden',
        'levels': [
            (1.0, '5级', '极缺'),
            (5.0, '4级', '缺乏'),
            (15.0, '3级', '中等'),
            (30.0, '2级', '丰富'),
            (float('inf'), '1级', '偏高')
        ]
    },
    'ACU': {
        'name': '有效铜',
        'unit': 'mg/kg',
        'reverse_display': True,
        'land_filter': 'cultivated_garden',
        'levels': [
            (0.1, '5级', '极缺'),
            (0.2, '4级', '缺乏'),
            (1.0, '3级', '中等'),
            (1.8, '2级', '丰富'),
            (float('inf'), '1级', '偏高')
        ]
    },
    'AZN': {
        'name': '有效锌',
        'unit': 'mg/kg',
        'reverse_display': True,
        'land_filter': 'cultivated_garden',
        'levels': [
            (0.3, '5级', '极缺'),
            (0.5, '4级', '缺乏'),
            (1.0, '3级', '中等'),
            (3.0, '2级', '丰富'),
            (float('inf'), '1级', '偏高')
        ]
    },
    'AB': {
        'name': '有效硼',
        'unit': 'mg/kg',
        'reverse_display': True,
        'land_filter': 'cultivated_garden',
        'levels': [
            (0.2, '5级', '极缺'),
            (0.5, '4级', '缺乏'),
            (1.0, '3级', '中等'),
            (2.0, '2级', '丰富'),
            (float('inf'), '1级', '偏高')
        ]
    },
    'AMO': {
        'name': '有效钼',
        'unit': 'mg/kg',
        'reverse_display': True,
        'land_filter': 'cultivated_garden',
        'levels': [
            (0.1, '5级', '极缺'),
            (0.15, '4级', '缺乏'),
            (0.2, '3级', '中等'),
            (0.3, '2级', '丰富'),
            (float('inf'), '1级', '偏高')
        ]
    },
    'ECA': {
        'name': '交换性钙',
        'unit': 'cmol(1/2Ca²⁺)/kg',
        'reverse_display': True,
        'land_filter': 'cultivated_garden',
        'levels': [
            (1.0, '5级', '极缺'),
            (4.0, '4级', '缺乏'),
            (10.0, '3级', '中等'),
            (15.0, '2级', '丰富'),
            (float('inf'), '1级', '偏高')
        ]
    },
    'EMG': {
        'name': '交换性镁',
        'unit': 'cmol(1/2Mg²⁺)/kg',
        'reverse_display': True,
        'land_filter': 'cultivated_garden',
        'levels': [
            (0.5, '5级', '极缺'),
            (1.0, '4级', '缺乏'),
            (1.5, '3级', '中等'),
            (2.0, '2级', '丰富'),
            (float('inf'), '1级', '偏高')
        ]
    },
    'SK': {
        'name': '缓效钾',
        'unit': 'mg/kg',
        'reverse_display': True,
        'land_filter': 'cultivated_garden',
        'levels': [
            (100, '5级', '极缺'),
            (300, '4级', '缺乏'),
            (500, '3级', '中等'),
            (700, '2级', '较丰富'),
            (float('inf'), '1级', '丰富')
        ]
    },
    'AS1': {
        'name': '有效硫',
        'unit': 'mg/kg',
        'reverse_display': True,
        'land_filter': 'cultivated_garden',
        'levels': [
            (10.0, '5级', '极缺'),
            (20.0, '4级', '缺乏'),
            (30.0, '3级', '中等'),
            (40.0, '2级', '丰富'),
            (float('inf'), '1级', '偏高')
        ]
    },
    # 只统计水田的属性
    'ASI': {
        'name': '有效硅',
        'unit': 'mg/kg',
        'reverse_display': True,
        'land_filter': 'paddy_only',
        'levels': [
            (50, '5级', '极缺'),
            (100, '4级', '缺乏'),
            (150, '3级', '中等'),
            (250, '2级', '丰富'),
            (float('inf'), '1级', '偏高')
        ]
    },
    # 只统计耕地的属性
    'GZCHD': {
        'name': '耕作层厚度',
        'unit': 'cm',
        'reverse_display': True,
        'land_filter': 'cultivated_only',
        'levels': [
            (10, '5级', '薄'),
            (15, '4级', '较薄'),
            (20, '3级', '中'),
            (25, '2级', '较厚'),
            (float('inf'), '1级', '厚')
        ]
    },
}

# 罗马数字映射
ROMAN_MAP = {
    '1级': 'Ⅰ级',
    '2级': 'Ⅱ级',
    '3级': 'Ⅲ级',
    '4级': 'Ⅳ级',
    '5级': 'Ⅴ级',
    '6级': 'Ⅵ级',
    '7级': 'Ⅶ级'
}

# 土地利用类型配置
LAND_USE_CONFIG = {
    "耕地": ["水田", "水浇地", "旱地"],
    "园地": ["果园", "茶园", "其他园地"],
    "林地": ["林地"],
    "草地": ["草地"],
    "其他": ["其他"]
}

# 用户提供的土壤类型顺序
SOIL_TYPE_ORDER = [
    # 红壤相关
    ('红壤', '棕红壤', '红泥质棕红壤'),
    ('红壤', '红壤性土', '砂泥质红壤性土'),
    ('红壤', '红壤性土', '麻砂质红壤性土'),
    
    # 黄棕壤相关
    ('黄棕壤', '典型黄棕壤', '暗泥质黄棕壤'),
    ('黄棕壤', '典型黄棕壤', '麻砂质黄棕壤'),
    ('黄棕壤', '典型黄棕壤', '红砂质黄棕壤'),
    ('黄棕壤', '典型黄棕壤', '黄土质黄棕壤'),
    ('黄棕壤', '典型黄棕壤', '砂泥质黄棕壤'),
    ('黄棕壤', '黄棕壤性土', '砂泥质黄棕壤性土'),
    
    # 棕壤相关
    ('棕壤', '典型棕壤', '麻砂质典型棕壤'),
    ('棕壤', '白浆化棕壤', '麻砂质白浆化棕壤'),
    ('棕壤', '白浆化棕壤', '泥砂质白浆化棕壤'),
    ('棕壤', '潮棕壤', '泥砂质潮棕壤'),
    
    # 褐土相关
    ('褐土', '淋溶褐土', '黄土质淋溶褐土'),
    ('褐土', '淋溶褐土', '灰泥质淋溶褐土'),
    ('褐土', '淋溶褐土', '暗泥质淋溶褐土'),
    ('褐土', '潮褐土', '泥砂质潮褐土'),
    
    # 其他土壤类型
    ('红黏土', '红黏土', '红黏土'),
    
    # 石灰岩土相关
    ('石灰岩土', '黑色石灰土', '黑色石灰土'),
    ('石灰岩土', '棕色石灰土', '棕色石灰土'),
    
    # 火山灰土相关
    ('火山灰土', '暗火山灰土', '暗火山灰土'),
    
    # 紫色土相关
    ('紫色土', '酸性紫色土', '壤质酸性紫色土'),
    ('紫色土', '酸性紫色土', '黏质酸性紫色土'),
    ('紫色土', '中性紫色土', '砂质中性紫色土'),
    ('紫色土', '中性紫色土', '壤质中性紫色土'),
    ('紫色土', '中性紫色土', '黏质中性紫色土'),
    ('紫色土', '石灰性紫色土', '壤质石灰性紫色土'),
    
    # 粗骨土相关
    ('粗骨土', '酸性粗骨土', '麻砂质酸性粗骨土'),
    ('粗骨土', '酸性粗骨土', '硅质酸性粗骨土'),
    ('粗骨土', '中性粗骨土', '麻砂质中性粗骨土'),
    ('粗骨土', '钙质粗骨土', '灰泥质钙质粗骨土'),
    
    # 潮土相关
    ('潮土', '典型潮土', '砂质潮土'),
    ('潮土', '典型潮土', '壤质潮土'),
    ('潮土', '典型潮土', '黏质潮土'),
    ('潮土', '灰潮土', '灰潮土'),
    ('潮土', '灰潮土', '石灰性灰潮土'),
    ('潮土', '盐化潮土（含碱化潮土）', '氯化物盐化潮土'),
    ('潮土', '盐化潮土（含碱化潮土）', '硫酸盐盐化潮土'),
    ('潮土', '盐化潮土（含碱化潮土）', '苏打盐化潮土'),
    
    # 砂姜黑土相关
    ('砂姜黑土', '典型砂姜黑土', '黑腐砂姜黑土（黑姜土）'),
    ('砂姜黑土', '典型砂姜黑土', '覆泥砂姜黑土（覆泥黑姜土）'),
    ('砂姜黑土', '盐化砂姜黑土', '氯化物盐化砂姜黑土'),
    
    # 沼泽土相关
    ('沼泽土', '腐泥沼泽土', '腐泥沼泽土'),
    ('沼泽土', '草甸沼泽土', '草甸沼泽土'),
    ('沼泽土', '草甸沼泽土', '石灰性草甸沼泽土'),
    
    # 滨海盐土相关
    ('滨海盐土', '典型滨海盐土', '氯化物滨海盐土'),
    ('滨海盐土', '滨海沼泽盐土', '氯化物沼泽滨海盐土'),
    ('滨海盐土', '滨海潮滩盐土', '氯化物潮滩滨海盐土'),
    
    # 水稻土相关
    ('水稻土', '淹育水稻土', '浅马肝泥田'),
    ('水稻土', '渗育水稻土', '渗灰泥田'),
    ('水稻土', '渗育水稻土', '渗潮泥砂田'),
    ('水稻土', '渗育水稻土', '渗潮泥田'),
    ('水稻土', '渗育水稻土', '渗湖泥田'),
    ('水稻土', '渗育水稻土', '渗涂泥田'),
    ('水稻土', '渗育水稻土', '渗淡涂泥田'),
    ('水稻土', '渗育水稻土', '渗麻砂泥田'),
    ('水稻土', '渗育水稻土', '渗潮白土田'),
    ('水稻土', '渗育水稻土', '渗马肝泥田'),
    ('水稻土', '潴育水稻土', '潮泥田'),
    ('水稻土', '潴育水稻土', '湖泥田'),
    ('水稻土', '潴育水稻土', '马肝泥田'),
    ('水稻土', '潜育水稻土', '青湖泥田'),
    ('水稻土', '潜育水稻土', '青马肝泥田'),
    ('水稻土', '脱潜水稻土', '黄斑黏田'),
    ('水稻土', '脱潜水稻土', '黄斑泥田'),
    ('水稻土', '漂洗水稻土', '漂潮白土田'),
    ('水稻土', '漂洗水稻土', '漂马肝泥田'),
    ('水稻土', '盐渍水稻土', '氯化物潮泥田'),
    ('水稻土', '盐渍水稻土', '氯化物涂泥田'),
    ('水稻土', '盐渍水稻土', '氯化物湖泥田'),
    ('水稻土', '盐渍水稻土', '硫酸盐潮泥田'),
    ('水稻土', '盐渍水稻土', '硫酸盐涂泥田'),
    ('水稻土', '盐渍水稻土', '苏打潮泥田'),
    ('水稻土', '盐渍水稻土', '苏打涂泥田'),
    ('水稻土', '盐渍水稻土', '苏打湖泥田'),
    
    # 人工土相关
    ('人工土', '填充土', '工矿填充土'),
    ('人工土', '填充土', '城镇填充土'),
    ('人工土', '扰动土', '运移扰动土')
]

# 从SOIL_TYPE_ORDER中提取土类、亚类、土属的层级结构
SOIL_TYPE_HIERARCHY = {}
for soil_type in SOIL_TYPE_ORDER:
    major, sub, genus = soil_type
    if major not in SOIL_TYPE_HIERARCHY:
        SOIL_TYPE_HIERARCHY[major] = {}
    if sub not in SOIL_TYPE_HIERARCHY[major]:
        SOIL_TYPE_HIERARCHY[major][sub] = []
    if genus not in SOIL_TYPE_HIERARCHY[major][sub]:
        SOIL_TYPE_HIERARCHY[major][sub].append(genus)


def format_small_value(value):
    """
    格式化小数值:
    - 大于等于 0.001: 显示3位小数
    - 小于 0.001: 显示到第一位非零数字
    例如: 0.000041 显示为 0.00004
    """
    if pd.isna(value) or value == 0:
        return round(value, 3) if not pd.isna(value) else value
    
    abs_val = abs(value)
    
    # 大于等于 0.001，正常显示3位小数
    if abs_val >= 0.001:
        return round(value, 3)
    
    # 小于 0.001，显示到第一位非零数字
    import math
    if abs_val == 0:
        return 0
    
    # 计算需要保留的小数位数（到第一位非零数字）
    decimal_places = -int(math.floor(math.log10(abs_val)))
    return round(value, decimal_places)


def format_percentage(value):
    """
    格式化百分比数值:
    - 如果超过100，显示100
    - 否则按照 format_small_value 逻辑处理
    """
    if pd.isna(value):
        return value
    
    # 如果超过100，显示100
    if value > 100:
        return 100
    
    return format_small_value(value)


def normalize_attr_column_name(col_name: str) -> str:
    """将原始列名映射为 SOIL_ATTR_CONFIG 中的标准键"""
    if pd.isna(col_name):
        return ""
    col_str = str(col_name).strip()

    # 显式别名映射（根据常见情况扩展）
    alias_map = {
        'pH': 'ph',
        'PH': 'ph',
        '酸碱度': 'ph',
        '有机质含量': 'OM',
        '有机质(g/kg)': 'OM',
        # 可继续添加...
    }

    # 先查别名
    if col_str in alias_map:
        return alias_map[col_str]

    # 再直接匹配（不区分大小写）
    for key in SOIL_ATTR_CONFIG.keys():
        if col_str.lower() == key.lower():
            return key

    # 否则返回原字符串（大概率不在配置中）
    return col_str


def classify_by_config(value, attr_key):
    """
    根据配置字典对属性值进行分级
    返回罗马数字级别
    区间规则：
    - 第一级：value ≤ threshold1
    - 中间级：prev_threshold < value ≤ threshold
    - 最后一级：value > last_threshold
    """
    if pd.isna(value):
        return None
    # 只统计大于0的值，0值和负值不参与分级
    if value <= 0:
        return None
    config = SOIL_ATTR_CONFIG.get(attr_key)
    if not config:
        return None

    levels = config['levels']
    
    # 遍历所有级别
    for i, (threshold, level, _) in enumerate(levels):
        if i == 0:
            # 第一个级别：value ≤ threshold
            if value <= threshold:
                return ROMAN_MAP.get(level, level)
        else:
            # 中间和最后级别：prev_threshold < value ≤ threshold
            prev_threshold = levels[i-1][0]
            if prev_threshold < value <= threshold:
                return ROMAN_MAP.get(level, level)
    
    return None


def ensure_land_class_column(df):
    """
    统一二级地类列名为 '二级地类'，兼容 'DLMC'（不区分大小写）
    参考自土壤属性多文件分析.py
    """
    if '二级地类' in df.columns:
        return df
    for col in df.columns:
        if str(col).strip().upper() == 'DLMC':
            return df.rename(columns={col: '二级地类'})
    # 尝试寻找地类名称列
    if '地类名称' in df.columns:
        return df.rename(columns={'地类名称': '二级地类'})
    if 'dlmc' in df.columns:
        return df.rename(columns={'dlmc': '二级地类'})
    raise ValueError("未找到二级地类列（期望列名：'二级地类' 或 'DLMC' 或 '地类名称'）")


def get_land_class(dlmc):
    """
    将二级地类映射为 (一级地类, 二级地类)
    参考自土壤属性多文件分析.py
    """
    if pd.isna(dlmc):
        return None
    s = str(dlmc).strip()
    if s in ["水田", "水浇地", "旱地"]:
        return ("耕地", s)
    elif s in ["果园", "茶园"]:
        return ("园地", s)
    elif "园地" in s:
        return ("园地", "其他园地")
    elif "林地" in s:
        return ("林地", "林地")
    elif "草地" in s:
        return ("草地", "草地")
    else:
        return ("其他", "其他")


def apply_land_filter(df, attr_key):
    """
    根据属性的用地限制过滤数据
    参考自土壤属性多文件分析.py的process_files函数
    """
    config = SOIL_ATTR_CONFIG.get(attr_key, {})
    land_filter = config.get('land_filter')
    
    if not land_filter:
        return df
    
    try:
        df = ensure_land_class_column(df)
        land_info = df['二级地类'].apply(get_land_class)
        df[['一级地类', '二级地类名']] = pd.DataFrame(land_info.tolist(), index=df.index)
        
        if land_filter == 'cultivated_garden':
            # 只统计耕地和园地
            df = df[df['一级地类'].isin(['耕地', '园地'])]
        elif land_filter == 'paddy_only':
            # 只统计水田
            df = df[(df['一级地类'] == '耕地') & (df['二级地类名'] == '水田')]
        elif land_filter == 'cultivated_only':
            # 只统计耕地
            df = df[df['一级地类'] == '耕地']
    except Exception as e:
        print(f"  [警告] 用地过滤失败: {e}")
    
    return df


def get_grade_ranges(attr_key):
    """
    根据属性配置生成分级范围文本
    返回一个字典，key为罗马数字级别，value为范围文本
    """
    config = SOIL_ATTR_CONFIG.get(attr_key)
    if not config:
        return {}
    
    levels = config['levels']
    ranges = {}
    
    # 处理每个级别
    for i, (threshold, level, _) in enumerate(levels):
        roman_level = ROMAN_MAP.get(level, level)
        
        if i == 0:
            # 第一个级别：≤threshold
            ranges[roman_level] = f'≤{threshold}'
        elif threshold == float('inf'):
            # 最后一个级别：>previous_threshold
            prev_threshold = levels[i-1][0]
            ranges[roman_level] = f'>{prev_threshold}'
        else:
            # 中间级别：>previous_threshold～threshold（使用中文波浪号连接）
            prev_threshold = levels[i-1][0]
            ranges[roman_level] = f'{prev_threshold}～{threshold}'
    
    return ranges


def get_grade_order(attr_key):
    """
    根据属性获取等级顺序列表
    """
    config = SOIL_ATTR_CONFIG.get(attr_key)
    if not config:
        return ['Ⅰ级', 'Ⅱ级', 'Ⅲ级', 'Ⅳ级', 'Ⅴ级']
    
    # 从配置中提取所有等级
    levels = config.get('levels', [])
    grade_set = set()
    for _, level, _ in levels:
        roman_level = ROMAN_MAP.get(level, level)
        grade_set.add(roman_level)
    
    # 按罗马数字排序（从Ⅰ到Ⅶ）
    all_grades = ['Ⅰ级', 'Ⅱ级', 'Ⅲ级', 'Ⅳ级', 'Ⅴ级', 'Ⅵ级', 'Ⅶ级']
    return [g for g in all_grades if g in grade_set]


def calculate_average_grade(df, attr_key, area_col='面积'):
    """
    计算平均等级（加权平均，根据面积加权）
    """
    grade_map = {
        'Ⅰ级': 1,
        'Ⅱ级': 2,
        'Ⅲ级': 3,
        'Ⅳ级': 4,
        'Ⅴ级': 5,
        'Ⅵ级': 6,
        'Ⅶ级': 7
    }
    
    # 创建副本避免SettingWithCopyWarning
    df = df.copy()
    # 确保属性列是数值类型
    df[attr_key] = pd.to_numeric(df[attr_key], errors='coerce')
    # 过滤有效数据
    valid_df = df[(df[attr_key] > 0) & (df[area_col].notna()) & (df[area_col] > 0)].copy()
    
    if valid_df.empty:
        return None
    
    # 分级
    valid_df['等级'] = valid_df[attr_key].apply(lambda x: classify_by_config(x, attr_key))
    # 转换等级为数值
    valid_df['等级数值'] = valid_df['等级'].map(grade_map)
    # 过滤无效等级
    valid_df = valid_df[valid_df['等级数值'].notna()]
    
    if valid_df.empty:
        return None
    
    # 计算加权平均
    total_area = valid_df[area_col].sum()
    weighted_sum = (valid_df['等级数值'] * valid_df[area_col]).sum()
    average_grade = weighted_sum / total_area
    
    return round(average_grade, 2)


def generate_town_summary(ws, df, attr_key):
    """
    生成分乡镇统计表格
    """
    from pypinyin import pinyin, Style
    
    config = SOIL_ATTR_CONFIG[attr_key]
    attr_name = config['name']
    unit = config['unit']
    ws.title = f"{attr_name}分乡镇统计"
    
    # 准备数据
    df_clean = df.copy()
    df_clean[attr_key] = pd.to_numeric(df_clean[attr_key], errors='coerce')
    df_clean = df_clean[df_clean[attr_key] > 0].copy()
    
    # 确保行政区名称列存在
    town_col = None
    for col in ['行政区名称', '乡镇', '镇/街道', '街道']:
        if col in df_clean.columns:
            town_col = col
            break
    if town_col is None:
        df_clean['行政区名称'] = ''
        town_col = '行政区名称'
    
    # 分级
    df_clean['等级'] = df_clean[attr_key].apply(lambda x: classify_by_config(x, attr_key))
    
    # 获取所有乡镇并按拼音排序
    towns = df_clean[town_col].dropna().unique().tolist()
    towns = [t for t in towns if t and str(t).strip()]
    
    def get_pinyin_key(name):
        try:
            return ''.join([p[0] for p in pinyin(str(name), style=Style.TONE3)])
        except:
            return str(name)
    
    towns = sorted(towns, key=get_pinyin_key)
    
    # 获取等级列表（按顺序）
    grade_order = get_grade_order(attr_key)
    
    # 获取分级范围
    grade_ranges = get_grade_ranges(attr_key)
    
    # 写入标题
    ws.merge_cells('A1:H1')
    ws['A1'] = f'{attr_name}分乡镇统计'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
   # 写入表头（第2-4行）
    # 第2行：行政县（合并A2:B4）、分级（合并5列）、指标平均等级
    ws.merge_cells('A2:B4')
    ws['A2'] = '行政县'
    ws['A2'].font = Font(bold=True)
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws.merge_cells('C2:G2')
    ws['C2'] = '分级'
    ws['C2'].font = Font(bold=True)
    ws['C2'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws.merge_cells('H2:H4')
    ws['H2'] = '指标平均等级'
    ws['H2'].font = Font(bold=True)
    ws['H2'].alignment = Alignment(horizontal='center', vertical='center')
    
    # 第3行：分级名称
    ws['C3'] = 'Ⅰ'
    ws['D3'] = 'Ⅱ'
    ws['E3'] = 'Ⅲ'
    ws['F3'] = 'Ⅳ'
    ws['G3'] = 'Ⅴ'
    
    for col in ['C', 'D', 'E', 'F', 'G']:
        ws[f'{col}3'].font = Font(bold=True)
        ws[f'{col}3'].alignment = Alignment(horizontal='center', vertical='center')
    
    # 第4行：分级范围（从配置自动生成）
    ws['C4'] = grade_ranges.get('Ⅰ级', '')
    ws['D4'] = grade_ranges.get('Ⅱ级', '')
    ws['E4'] = grade_ranges.get('Ⅲ级', '')
    ws['F4'] = grade_ranges.get('Ⅳ级', '')
    ws['G4'] = grade_ranges.get('Ⅴ级', '')
    
    for col in ['C', 'D', 'E', 'F', 'G']:
        ws[f'{col}4'].font = Font(bold=True)
        ws[f'{col}4'].alignment = Alignment(horizontal='center', vertical='center')
    
    # 写入数据
    current_row = 5
    
    for town in towns:
        town_df = df_clean[df_clean[town_col] == town]
        
        # 统计各级别面积
        grade_counts = town_df.groupby('等级')['面积'].sum().reindex(grade_order, fill_value=0)
        total_area = grade_counts.sum()
        
        # 计算平均等级
        avg_grade = calculate_average_grade(town_df, attr_key)
        
        # 写入面积行：A=镇名, B=面积, C-G=分级数据, H=平均等级
        ws.cell(current_row, 1, town)
        ws.cell(current_row, 2, '面积')
        ws.cell(current_row, 3, format_small_value(grade_counts['Ⅰ级']))
        ws.cell(current_row, 4, format_small_value(grade_counts['Ⅱ级']))
        ws.cell(current_row, 5, format_small_value(grade_counts['Ⅲ级']))
        ws.cell(current_row, 6, format_small_value(grade_counts['Ⅳ级']))
        ws.cell(current_row, 7, format_small_value(grade_counts['Ⅴ级']))
        ws.cell(current_row, 8, avg_grade if avg_grade is not None else '')
        
        # 合并镇/街道名称单元格（2行）
        ws.merge_cells(f'A{current_row}:A{current_row+1}')
        ws.cell(current_row, 1).alignment = Alignment(horizontal='center', vertical='center')
        
        # 合并平均等级单元格（2行）
        ws.merge_cells(f'H{current_row}:H{current_row+1}')
        ws.cell(current_row, 8).alignment = Alignment(horizontal='center', vertical='center')
        
        # 写入占比行
        ws.cell(current_row+1, 2, '占比')
        ws.cell(current_row+1, 3, format_percentage((grade_counts['Ⅰ级'] / total_area * 100) if total_area > 0 else 0))
        ws.cell(current_row+1, 4, format_percentage((grade_counts['Ⅱ级'] / total_area * 100) if total_area > 0 else 0))
        ws.cell(current_row+1, 5, format_percentage((grade_counts['Ⅲ级'] / total_area * 100) if total_area > 0 else 0))
        ws.cell(current_row+1, 6, format_percentage((grade_counts['Ⅳ级'] / total_area * 100) if total_area > 0 else 0))
        ws.cell(current_row+1, 7, format_percentage((grade_counts['Ⅴ级'] / total_area * 100) if total_area > 0 else 0))
        
        current_row += 2
    
    # 写入全域统计
    total_df = df_clean
    total_grade_counts = total_df.groupby('等级')['面积'].sum().reindex(grade_order, fill_value=0)
    total_total_area = total_grade_counts.sum()
    total_avg_grade = calculate_average_grade(total_df, attr_key)
    
    ws.cell(current_row, 1, '全域')
    ws.cell(current_row, 2, '面积')
    ws.cell(current_row, 3, format_small_value(total_grade_counts['Ⅰ级']))
    ws.cell(current_row, 4, format_small_value(total_grade_counts['Ⅱ级']))
    ws.cell(current_row, 5, format_small_value(total_grade_counts['Ⅲ级']))
    ws.cell(current_row, 6, format_small_value(total_grade_counts['Ⅳ级']))
    ws.cell(current_row, 7, format_small_value(total_grade_counts['Ⅴ级']))
    ws.cell(current_row, 8, total_avg_grade if total_avg_grade is not None else '')
    
    # 合并全域名称单元格
    ws.merge_cells(f'A{current_row}:A{current_row+1}')
    ws.cell(current_row, 1).alignment = Alignment(horizontal='center', vertical='center')
    
    # 合并平均等级单元格
    ws.merge_cells(f'H{current_row}:H{current_row+1}')
    ws.cell(current_row, 8).alignment = Alignment(horizontal='center', vertical='center')
    
    ws.cell(current_row+1, 2, '占比')
    ws.cell(current_row+1, 3, format_percentage((total_grade_counts['Ⅰ级'] / total_total_area * 100) if total_total_area > 0 else 0))
    ws.cell(current_row+1, 4, format_percentage((total_grade_counts['Ⅱ级'] / total_total_area * 100) if total_total_area > 0 else 0))
    ws.cell(current_row+1, 5, format_percentage((total_grade_counts['Ⅲ级'] / total_total_area * 100) if total_total_area > 0 else 0))
    ws.cell(current_row+1, 6, format_percentage((total_grade_counts['Ⅳ级'] / total_total_area * 100) if total_total_area > 0 else 0))
    ws.cell(current_row+1, 7, format_percentage((total_grade_counts['Ⅴ级'] / total_total_area * 100) if total_total_area > 0 else 0))

    # 设置边框和对齐
    thin = Side(border_style="thin")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    
    for row in ws.iter_rows(min_row=2, max_row=current_row+1, min_col=1, max_col=8):
        for cell in row:
            cell.border = border
            if cell.alignment.horizontal is None:
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
    # 调整列宽
    ws.column_dimensions['A'].width = 14
    for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws.column_dimensions[col].width = 12


def generate_land_use_summary(ws, df, attr_key):
    """
    生成土地利用类型统计表格
    """
    config = SOIL_ATTR_CONFIG[attr_key]
    attr_name = config['name']
    unit = config['unit']
    ws.title = f"{attr_name}土地利用类型统计"
    
    # 准备数据
    df_clean = df.copy()
    df_clean[attr_key] = pd.to_numeric(df_clean[attr_key], errors='coerce')
    df_clean = df_clean[df_clean[attr_key] > 0].copy()
    
    # 确保土地利用类型列存在（支持多种列名）
    if '二级地类' in df_clean.columns:
        df_clean['土地利用类型'] = df_clean['二级地类']
    elif 'DLMC' in df_clean.columns:
        df_clean['土地利用类型'] = df_clean['DLMC']
    elif 'dlmc' in df_clean.columns:
        df_clean['土地利用类型'] = df_clean['dlmc']
    elif '地类名称' in df_clean.columns:
        df_clean['土地利用类型'] = df_clean['地类名称']
    else:
        df_clean['土地利用类型'] = ''
    
    # 映射到一级和二级地类
    def map_land_use(land_type):
        if pd.isna(land_type):
            return ('其他', '其他')
        land_type = str(land_type).strip()
        if land_type in ['水田', '水浇地', '旱地']:
            return ('耕地', land_type)
        elif land_type in ['果园', '茶园']:
            return ('园地', land_type)
        elif '园地' in land_type:
            return ('园地', '其他园地')
        elif '林地' in land_type:
            return ('林地', '林地')
        elif '草地' in land_type:
            return ('草地', '草地')
        else:
            return ('其他', '其他')
    
    df_clean[['一级地类', '二级地类']] = df_clean['土地利用类型'].apply(map_land_use).apply(pd.Series)
    
    # 检查面积列是否存在
    if '面积' not in df_clean.columns:
        # 尝试使用其他可能的列名
        if 'AREA' in df_clean.columns:
            df_clean['面积'] = df_clean['AREA']
        elif 'area' in df_clean.columns:
            df_clean['面积'] = df_clean['area']
        else:
            return
    
    # 分级
    df_clean['等级'] = df_clean[attr_key].apply(lambda x: classify_by_config(x, attr_key))
    
    # 获取等级列表
    grade_order = get_grade_order(attr_key)
    
    # 获取分级范围
    grade_ranges = get_grade_ranges(attr_key)
    
    # 写入标题
    ws.merge_cells('A1:I1')
    ws['A1'] = f'{attr_name}土地利用类型统计'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # 写入表头（第2-4行）
    # 第2行：土地利用类型（合并A2:C4） | 分级（5列） | 平均等级
    ws.merge_cells('A2:C4')
    ws['A2'] = '土地利用类型'
    ws['A2'].font = Font(bold=True)
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws.merge_cells('D2:H2')
    ws['D2'] = '分级'
    ws['D2'].font = Font(bold=True)
    ws['D2'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws.merge_cells('I2:I4')
    ws['I2'] = '指标平均等级'
    ws['I2'].font = Font(bold=True)
    ws['I2'].alignment = Alignment(horizontal='center', vertical='center')
    
    # 第3行：等级（Ⅰ、Ⅱ、Ⅲ、Ⅳ、Ⅴ）
    ws['D3'] = 'Ⅰ'
    ws['E3'] = 'Ⅱ'
    ws['F3'] = 'Ⅲ'
    ws['G3'] = 'Ⅳ'
    ws['H3'] = 'Ⅴ'
    
    # 第4行：区间（从配置自动生成）
    ws['D4'] = grade_ranges.get("Ⅰ级", "")
    ws['E4'] = grade_ranges.get("Ⅱ级", "")
    ws['F4'] = grade_ranges.get("Ⅲ级", "")
    ws['G4'] = grade_ranges.get("Ⅳ级", "")
    ws['H4'] = grade_ranges.get("Ⅴ级", "")
    
    for col in ['D', 'E', 'F', 'G', 'H']:
        ws[f'{col}3'].font = Font(bold=True)
        ws[f'{col}3'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'{col}4'].font = Font(bold=True)
        ws[f'{col}4'].alignment = Alignment(horizontal='center', vertical='center')
    
    # 固定的土地利用类型结构
    land_structure = [
        ('耕地', ['水田', '水浇地', '旱地']),
        ('园地', ['果园', '茶园', '其他园地']),
        ('林地', []),
        ('草地', []),
        ('其他', [])
    ]
    
    # 写入数据（从第5行开始）
    current_row = 5
    
    for primary_type, secondaries in land_structure:
        primary_df = df_clean[df_clean['一级地类'] == primary_type]
        
        # 一级地类的起始行
        primary_start_row = current_row
        
        # 如果有二级地类，先处理二级地类，然后汇总到一级地类
        if secondaries:
            # 用于汇总二级地类数据
            primary_grade_counts_sum = pd.Series(0, index=grade_order)
            primary_total_area_sum = 0
            
            # 先写入一级地类行（后面会填充数据）
            primary_row = current_row
            ws.cell(current_row, 1, '')
            ws.cell(current_row, 2, '')
            ws.cell(current_row, 3, '面积')
            current_row += 2  # 预留两行给一级地类
            
            # 记录二级地类开始行
            secondary_start_row = current_row
            
            # 处理每个二级地类
            for secondary_type in secondaries:
                secondary_df = primary_df[primary_df['二级地类'] == secondary_type]
                
                if not secondary_df.empty:
                    secondary_grade_counts = secondary_df.groupby('等级')['面积'].sum().reindex(grade_order, fill_value=0)
                    secondary_total_area = secondary_grade_counts.sum()
                    secondary_avg_grade = calculate_average_grade(secondary_df, attr_key)
                    
                    # 累加到一级地类总和
                    primary_grade_counts_sum += secondary_grade_counts
                    primary_total_area_sum += secondary_total_area
                    
                    # 写入二级地类面积行
                    ws.cell(current_row, 1, '')
                    ws.cell(current_row, 2, secondary_type)
                    ws.cell(current_row, 3, '面积')
                    ws.cell(current_row, 4, format_small_value(secondary_grade_counts['Ⅰ级']))
                    ws.cell(current_row, 5, format_small_value(secondary_grade_counts['Ⅱ级']))
                    ws.cell(current_row, 6, format_small_value(secondary_grade_counts['Ⅲ级']))
                    ws.cell(current_row, 7, format_small_value(secondary_grade_counts['Ⅳ级']))
                    ws.cell(current_row, 8, format_small_value(secondary_grade_counts['Ⅴ级']))
                    ws.cell(current_row, 9, secondary_avg_grade if secondary_avg_grade is not None else '')
                    
                    # 写入二级地类占比行
                    ws.cell(current_row+1, 1, '')
                    ws.cell(current_row+1, 3, '占比')
                    ws.cell(current_row+1, 4, format_percentage((secondary_grade_counts['Ⅰ级'] / secondary_total_area * 100) if secondary_total_area > 0 else 0))
                    ws.cell(current_row+1, 5, format_percentage((secondary_grade_counts['Ⅱ级'] / secondary_total_area * 100) if secondary_total_area > 0 else 0))
                    ws.cell(current_row+1, 6, format_percentage((secondary_grade_counts['Ⅲ级'] / secondary_total_area * 100) if secondary_total_area > 0 else 0))
                    ws.cell(current_row+1, 7, format_percentage((secondary_grade_counts['Ⅳ级'] / secondary_total_area * 100) if secondary_total_area > 0 else 0))
                    ws.cell(current_row+1, 8, format_percentage((secondary_grade_counts['Ⅴ级'] / secondary_total_area * 100) if secondary_total_area > 0 else 0))
                    
                    # 合并B列（二级地类名称）
                    ws.merge_cells(f'B{current_row}:B{current_row+1}')
                    ws.cell(current_row, 2).alignment = Alignment(horizontal='center', vertical='center')
                    
                    # 合并平均等级单元格（2行）
                    ws.merge_cells(f'I{current_row}:I{current_row+1}')
                    ws.cell(current_row, 9).alignment = Alignment(horizontal='center', vertical='center')
                    
                    current_row += 2
                else:
                    # 即使没有数据也要写出结构
                    ws.cell(current_row, 1, '')
                    ws.cell(current_row, 2, secondary_type)
                    ws.cell(current_row, 3, '面积')
                    ws.cell(current_row, 4, 0)
                    ws.cell(current_row, 5, 0)
                    ws.cell(current_row, 6, 0)
                    ws.cell(current_row, 7, 0)
                    ws.cell(current_row, 8, 0)
                    ws.cell(current_row, 9, '')
                    
                    ws.cell(current_row+1, 1, '')
                    ws.cell(current_row+1, 3, '占比')
                    ws.cell(current_row+1, 4, '0.00%')
                    ws.cell(current_row+1, 5, '0.00%')
                    ws.cell(current_row+1, 6, '0.00%')
                    ws.cell(current_row+1, 7, '0.00%')
                    ws.cell(current_row+1, 8, '0.00%')
                    
                    ws.merge_cells(f'B{current_row}:B{current_row+1}')
                    ws.cell(current_row, 2).alignment = Alignment(horizontal='center', vertical='center')
                    
                    ws.merge_cells(f'I{current_row}:I{current_row+1}')
                    ws.cell(current_row, 9).alignment = Alignment(horizontal='center', vertical='center')
                    
                    current_row += 2
            
            # 现在回填一级地类的数据（二级地类的总和）
            if primary_total_area_sum > 0:
                primary_avg_grade = calculate_average_grade(primary_df, attr_key)
                
                ws.cell(primary_row, 4, format_small_value(primary_grade_counts_sum['Ⅰ级']))
                ws.cell(primary_row, 5, format_small_value(primary_grade_counts_sum['Ⅱ级']))
                ws.cell(primary_row, 6, format_small_value(primary_grade_counts_sum['Ⅲ级']))
                ws.cell(primary_row, 7, format_small_value(primary_grade_counts_sum['Ⅳ级']))
                ws.cell(primary_row, 8, format_small_value(primary_grade_counts_sum['Ⅴ级']))
                ws.cell(primary_row, 9, primary_avg_grade if primary_avg_grade is not None else '')
                
                ws.cell(primary_row+1, 3, '占比')
                ws.cell(primary_row+1, 4, format_percentage((primary_grade_counts_sum['Ⅰ级'] / primary_total_area_sum * 100) if primary_total_area_sum > 0 else 0))
                ws.cell(primary_row+1, 5, format_percentage((primary_grade_counts_sum['Ⅱ级'] / primary_total_area_sum * 100) if primary_total_area_sum > 0 else 0))
                ws.cell(primary_row+1, 6, format_percentage((primary_grade_counts_sum['Ⅲ级'] / primary_total_area_sum * 100) if primary_total_area_sum > 0 else 0))
                ws.cell(primary_row+1, 7, format_percentage((primary_grade_counts_sum['Ⅳ级'] / primary_total_area_sum * 100) if primary_total_area_sum > 0 else 0))
                ws.cell(primary_row+1, 8, format_percentage((primary_grade_counts_sum['Ⅴ级'] / primary_total_area_sum * 100) if primary_total_area_sum > 0 else 0))
            else:
                ws.cell(primary_row, 4, 0)
                ws.cell(primary_row, 5, 0)
                ws.cell(primary_row, 6, 0)
                ws.cell(primary_row, 7, 0)
                ws.cell(primary_row, 8, 0)
                ws.cell(primary_row, 9, '')
                
                ws.cell(primary_row+1, 3, '占比')
                ws.cell(primary_row+1, 4, '0.00%')
                ws.cell(primary_row+1, 5, '0.00%')
                ws.cell(primary_row+1, 6, '0.00%')
                ws.cell(primary_row+1, 7, '0.00%')
                ws.cell(primary_row+1, 8, '0.00%')
            
            # 合并A-B列（一级地类名称，2行）
            ws.merge_cells(f'A{primary_row}:B{primary_row+1}')
            ws.cell(primary_row, 1, primary_type)
            ws.cell(primary_row, 1).alignment = Alignment(horizontal='center', vertical='center')
            
            # 合并平均等级单元格（2行）
            ws.merge_cells(f'I{primary_row}:I{primary_row+1}')
            ws.cell(primary_row, 9).alignment = Alignment(horizontal='center', vertical='center')
            
            # 合并A列（一级地类跨越所有二级地类）
            if current_row > secondary_start_row:
                ws.merge_cells(f'A{secondary_start_row}:A{current_row-1}')
                ws.cell(secondary_start_row, 1, primary_type)
                ws.cell(secondary_start_row, 1).alignment = Alignment(horizontal='center', vertical='center')
        else:
            # 没有二级地类，直接统计一级地类
            if not primary_df.empty:
                primary_grade_counts = primary_df.groupby('等级')['面积'].sum().reindex(grade_order, fill_value=0)
                primary_total_area = primary_grade_counts.sum()
                primary_avg_grade = calculate_average_grade(primary_df, attr_key)
                
                ws.cell(current_row, 1, '')
                ws.cell(current_row, 2, '')
                ws.cell(current_row, 3, '面积')
                ws.cell(current_row, 4, format_small_value(primary_grade_counts['Ⅰ级']))
                ws.cell(current_row, 5, format_small_value(primary_grade_counts['Ⅱ级']))
                ws.cell(current_row, 6, format_small_value(primary_grade_counts['Ⅲ级']))
                ws.cell(current_row, 7, format_small_value(primary_grade_counts['Ⅳ级']))
                ws.cell(current_row, 8, format_small_value(primary_grade_counts['Ⅴ级']))
                ws.cell(current_row, 9, primary_avg_grade if primary_avg_grade is not None else '')
                
                ws.cell(current_row+1, 3, '占比')
                ws.cell(current_row+1, 4, format_percentage((primary_grade_counts['Ⅰ级'] / primary_total_area * 100) if primary_total_area > 0 else 0))
                ws.cell(current_row+1, 5, format_percentage((primary_grade_counts['Ⅱ级'] / primary_total_area * 100) if primary_total_area > 0 else 0))
                ws.cell(current_row+1, 6, format_percentage((primary_grade_counts['Ⅲ级'] / primary_total_area * 100) if primary_total_area > 0 else 0))
                ws.cell(current_row+1, 7, format_percentage((primary_grade_counts['Ⅳ级'] / primary_total_area * 100) if primary_total_area > 0 else 0))
                ws.cell(current_row+1, 8, format_percentage((primary_grade_counts['Ⅴ级'] / primary_total_area * 100) if primary_total_area > 0 else 0))
            else:
                ws.cell(current_row, 1, '')
                ws.cell(current_row, 2, '')
                ws.cell(current_row, 3, '面积')
                ws.cell(current_row, 4, 0)
                ws.cell(current_row, 5, 0)
                ws.cell(current_row, 6, 0)
                ws.cell(current_row, 7, 0)
                ws.cell(current_row, 8, 0)
                ws.cell(current_row, 9, '')
                
                ws.cell(current_row+1, 3, '占比')
                ws.cell(current_row+1, 4, '0.00%')
                ws.cell(current_row+1, 5, '0.00%')
                ws.cell(current_row+1, 6, '0.00%')
                ws.cell(current_row+1, 7, '0.00%')
                ws.cell(current_row+1, 8, '0.00%')
            
            # 合并A-B列（一级地类名称，2行）
            ws.merge_cells(f'A{current_row}:B{current_row+1}')
            ws.cell(current_row, 1, primary_type)
            ws.cell(current_row, 1).alignment = Alignment(horizontal='center', vertical='center')
            
            # 合并平均等级单元格（2行）
            ws.merge_cells(f'I{current_row}:I{current_row+1}')
            ws.cell(current_row, 9).alignment = Alignment(horizontal='center', vertical='center')
            
            current_row += 2
    
    # 写入全域统计
    total_df = df_clean
    total_grade_counts = total_df.groupby('等级')['面积'].sum().reindex(grade_order, fill_value=0)
    total_total_area = total_grade_counts.sum()
    total_avg_grade = calculate_average_grade(total_df, attr_key)
    
    ws.cell(current_row, 1, '')
    ws.cell(current_row, 2, '')
    ws.cell(current_row, 3, '面积')
    ws.cell(current_row, 4, format_small_value(total_grade_counts['Ⅰ级']))
    ws.cell(current_row, 5, format_small_value(total_grade_counts['Ⅱ级']))
    ws.cell(current_row, 6, format_small_value(total_grade_counts['Ⅲ级']))
    ws.cell(current_row, 7, format_small_value(total_grade_counts['Ⅳ级']))
    ws.cell(current_row, 8, format_small_value(total_grade_counts['Ⅴ级']))
    ws.cell(current_row, 9, total_avg_grade if total_avg_grade is not None else '')
    
    # 写入占比行
    ws.cell(current_row+1, 3, '占比')
    ws.cell(current_row+1, 4, format_percentage((total_grade_counts['Ⅰ级'] / total_total_area * 100) if total_total_area > 0 else 0))
    ws.cell(current_row+1, 5, format_percentage((total_grade_counts['Ⅱ级'] / total_total_area * 100) if total_total_area > 0 else 0))
    ws.cell(current_row+1, 6, format_percentage((total_grade_counts['Ⅲ级'] / total_total_area * 100) if total_total_area > 0 else 0))
    ws.cell(current_row+1, 7, format_percentage((total_grade_counts['Ⅳ级'] / total_total_area * 100) if total_total_area > 0 else 0))
    ws.cell(current_row+1, 8, format_percentage((total_grade_counts['Ⅴ级'] / total_total_area * 100) if total_total_area > 0 else 0))
    
    # 合并A-B列（全域名称）
    ws.merge_cells(f'A{current_row}:B{current_row+1}')
    ws.cell(current_row, 1, '全域')
    ws.cell(current_row, 1).alignment = Alignment(horizontal='center', vertical='center')
    
    # 合并平均等级单元格（2行）
    ws.merge_cells(f'I{current_row}:I{current_row+1}')
    ws.cell(current_row, 9).alignment = Alignment(horizontal='center', vertical='center')
    
    # 设置边框和对齐
    thin = Side(border_style="thin")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    
    for row in ws.iter_rows(min_row=1, max_row=current_row+1, min_col=1, max_col=9):
        for cell in row:
            cell.border = border
            if cell.row > 3:  # 数据行居中
                cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 调整列宽
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    for col in ['D', 'E', 'F', 'G', 'H', 'I']:
        ws.column_dimensions[col].width = 12


def generate_sample_point_summary(ws, df, attr_key):
    """
    生成样点数据统计表格（第4张表）
    """
    config = SOIL_ATTR_CONFIG[attr_key]
    attr_name = config['name']
    unit = config['unit']
    ws.title = f"{attr_name}样点统计"
    
    # 准备数据
    df_clean = df.copy()
    df_clean[attr_key] = pd.to_numeric(df_clean[attr_key], errors='coerce')
    df_clean = df_clean[df_clean[attr_key] > 0].copy()
    
    if df_clean.empty:
        return
    
    # 分级
    df_clean['等级'] = df_clean[attr_key].apply(lambda x: classify_by_config(x, attr_key))
    
   # 获取等级列表（动态支持5级或7级）
    grade_order = get_grade_order(attr_key)
    num_grades = len(grade_order)
    
    # 获取分级范围
    grade_ranges = get_grade_ranges(attr_key)
    
    # 动态计算列数（A列 + 等级列数）
    total_cols = 1 + num_grades
    last_col_letter = get_column_letter(total_cols)
    
    # 写入标题
    ws.merge_cells(f'A1:{last_col_letter}1')
    ws['A1'] = f'{attr_name}样点统计'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # 写入表头（第2-3行）
    ws['A2'] = '等级'
    for i, grade in enumerate(grade_order, start=2):
        col_letter = get_column_letter(i)
        # 写入罗马数字（去掉"级"字）
        ws[f'{col_letter}2'] = grade.replace('级', '')
        ws[f'{col_letter}2'].font = Font(bold=True)
        ws[f'{col_letter}2'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws['A3'] = '级别范围'
    for i, grade in enumerate(grade_order, start=2):
        col_letter = get_column_letter(i)
        ws[f'{col_letter}3'] = grade_ranges.get(grade, '')
        ws[f'{col_letter}3'].font = Font(bold=True)
        ws[f'{col_letter}3'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws['A2'].font = Font(bold=True)
    ws['A3'].font = Font(bold=True)
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A3'].alignment = Alignment(horizontal='center', vertical='center')
    
    # 统计各级别样点数
    grade_counts = df_clean['等级'].value_counts().reindex(grade_order, fill_value=0)
    total_samples = len(df_clean)
    
    # 计算统计值
    mean_value = df_clean[attr_key].mean()
    median_value = df_clean[attr_key].median()
    min_value = df_clean[attr_key].min()
    max_value = df_clean[attr_key].max()
    
    # 第4行：样点数
    ws['A4'] = '样点数'
    for i, grade in enumerate(grade_order, start=2):
        col_letter = get_column_letter(i)
        ws[f'{col_letter}4'] = int(grade_counts[grade])
    
    # 第5行：样点数占比
    ws['A5'] = '样点数占比'
    for i, grade in enumerate(grade_order, start=2):
        col_letter = get_column_letter(i)
        ws[f'{col_letter}5'] = format_percentage((grade_counts[grade] / total_samples * 100) if total_samples > 0 else 0)
    
    # 第6行：全域均值 | 全域最小值（横向排列）
    ws['A6'] = '全域均值'
    ws.merge_cells('B6:C6')  # 合并B6C6
    ws['B6'] = format_small_value(mean_value)
    ws['B6'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws['D6'] = '全域最小值'
    ws.merge_cells('E6:F6')  # 合并E6F6
    ws['E6'] = format_small_value(min_value)
    ws['E6'].alignment = Alignment(horizontal='center', vertical='center')
    
    # 第7行：全域中位值 | 全域最大值（横向排列）
    ws['A7'] = '全域中位值'
    ws.merge_cells('B7:C7')  # 合并B7C7
    ws['B7'] = format_small_value(median_value)
    ws['B7'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws['D7'] = '全域最大值'
    ws.merge_cells('E7:F7')  # 合并E7F7
    ws['E7'] = format_small_value(max_value)
    ws['E7'].alignment = Alignment(horizontal='center', vertical='center')
    
    # 设置边框
    thin = Side(border_style="thin")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    
    for row in ws.iter_rows(min_row=2, max_row=7, min_col=1, max_col=total_cols):
        for cell in row:
            cell.border = border
            if cell.alignment.horizontal is None:
                cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 调整列宽
    for i in range(1, total_cols + 1):
        ws.column_dimensions[get_column_letter(i)].width = 15


def generate_town_sample_summary(ws, df, attr_key):
    """
    生成分行政区样点统计表格（第5张表）
    从样点数据读取，统计每个行政区的均值、最小值、最大值、样点数、样点占比
    """
    from pypinyin import pinyin, Style
    
    config = SOIL_ATTR_CONFIG[attr_key]
    attr_name = config['name']
    unit = config['unit']
    ws.title = f"{attr_name}分行政区样点统计"
    
    # 准备数据
    df_clean = df.copy()
    df_clean[attr_key] = pd.to_numeric(df_clean[attr_key], errors='coerce')
    df_clean = df_clean[df_clean[attr_key] > 0].copy()
    
    if df_clean.empty:
        return
    
    # 确保行政区名称列存在
    town_col = None
    for col in ['行政区名称', '乡镇', '镇/街道', '街道', '行政区']:
        if col in df_clean.columns:
            town_col = col
            break
    if town_col is None:
        df_clean['行政区名称'] = ''
        town_col = '行政区名称'
    
    # 获取所有行政区并按拼音排序
    towns = df_clean[town_col].dropna().unique().tolist()
    towns = [t for t in towns if t and str(t).strip()]
    
    def get_pinyin_key(name):
        try:
            return ''.join([p[0] for p in pinyin(str(name), style=Style.TONE3)])
        except:
            return str(name)
    
    towns = sorted(towns, key=get_pinyin_key)
    
    # 写入标题
    ws.merge_cells('A1:F1')
    ws['A1'] = f'{attr_name}分行政区样点统计'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # 写入表头（第2-3行）
    ws.merge_cells('A2:A3')
    ws['A2'] = '行政区'
    ws['A2'].font = Font(bold=True)
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws.merge_cells('B2:F2')
    ws['B2'] = '样点统计'
    ws['B2'].font = Font(bold=True)
    ws['B2'].alignment = Alignment(horizontal='center', vertical='center')
    
    # 第3行：子表头
    ws['B3'] = '均值'
    ws['C3'] = '最小值'
    ws['D3'] = '最大值'
    ws['E3'] = '样点数'
    ws['F3'] = '样点占比'
    
    for col in ['B', 'C', 'D', 'E', 'F']:
        ws[f'{col}3'].font = Font(bold=True)
        ws[f'{col}3'].alignment = Alignment(horizontal='center', vertical='center')
    
    # 写入数据
    current_row = 4
    total_samples = len(df_clean)
    
    for town in towns:
        town_df = df_clean[df_clean[town_col] == town]
        
        if not town_df.empty:
            mean_value = town_df[attr_key].mean()
            min_value = town_df[attr_key].min()
            max_value = town_df[attr_key].max()
            sample_count = len(town_df)
            sample_ratio = (sample_count / total_samples * 100) if total_samples > 0 else 0
            
            ws.cell(current_row, 1, town)
            ws.cell(current_row, 2, format_small_value(mean_value))
            ws.cell(current_row, 3, format_small_value(min_value))
            ws.cell(current_row, 4, format_small_value(max_value))
            ws.cell(current_row, 5, sample_count)
            ws.cell(current_row, 6, format_percentage(sample_ratio))
            
            current_row += 1
    
    # 写入全域统计
    total_mean = df_clean[attr_key].mean()
    total_min = df_clean[attr_key].min()
    total_max = df_clean[attr_key].max()
    
    ws.cell(current_row, 1, '全域')
    ws.cell(current_row, 2, format_small_value(total_mean))
    ws.cell(current_row, 3, format_small_value(total_min))
    ws.cell(current_row, 4, format_small_value(total_max))
    ws.cell(current_row, 5, total_samples)
    ws.cell(current_row, 6, '100')
    
    # 设置边框和对齐
    thin = Side(border_style="thin")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    
    for row in ws.iter_rows(min_row=2, max_row=current_row, min_col=1, max_col=6):
        for cell in row:
            cell.border = border
            if cell.alignment.horizontal is None:
                cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 调整列宽
    ws.column_dimensions['A'].width = 14
    for col in ['B', 'C', 'D', 'E', 'F']:
        ws.column_dimensions[col].width = 12


def generate_land_use_sample_summary(ws, df, attr_key):
    """
    生成土地利用类型样点统计表格（第6张表）
    从样点数据读取
    """
    config = SOIL_ATTR_CONFIG[attr_key]
    attr_name = config['name']
    unit = config['unit']
    ws.title = f"{attr_name}土地利用类型样点统计"
    
    # 准备数据
    df_clean = df.copy()
    df_clean[attr_key] = pd.to_numeric(df_clean[attr_key], errors='coerce')
    df_clean = df_clean[df_clean[attr_key] > 0].copy()
    
    if df_clean.empty:
        return
    
    # 确保土地利用类型列存在（支持多种列名）
    if '二级地类' in df_clean.columns:
        df_clean['土地利用类型'] = df_clean['二级地类']
    elif 'DLMC' in df_clean.columns:
        df_clean['土地利用类型'] = df_clean['DLMC']
    elif 'dlmc' in df_clean.columns:
        df_clean['土地利用类型'] = df_clean['dlmc']
    elif '地类名称' in df_clean.columns:
        df_clean['土地利用类型'] = df_clean['地类名称']
    else:
        df_clean['土地利用类型'] = ''
    
    # 映射到一级和二级地类
    def map_land_use(land_type):
        if pd.isna(land_type):
            return ('其他', '其他')
        land_type = str(land_type).strip()
        if land_type in ['水田', '水浇地', '旱地']:
            return ('耕地', land_type)
        elif land_type in ['果园', '茶园']:
            return ('园地', land_type)
        elif '园地' in land_type:
            return ('园地', '其他园地')
        elif '林地' in land_type:
            return ('林地', '林地')
        elif '草地' in land_type:
            return ('草地', '草地')
        else:
            return ('其他', '其他')
    
    df_clean[['一级地类', '二级地类']] = df_clean['土地利用类型'].apply(map_land_use).apply(pd.Series)
    
    # 写入标题
    ws.merge_cells('A1:G1')
    ws['A1'] = f'{attr_name}土地利用类型样点统计'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # 写入表头（第2-3行）
    ws.merge_cells('A2:B3')
    ws['A2'] = '土地利用类型'
    ws['A2'].font = Font(bold=True)
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws.merge_cells('C2:G2')
    ws['C2'] = '样点统计'
    ws['C2'].font = Font(bold=True)
    ws['C2'].alignment = Alignment(horizontal='center', vertical='center')
    
    # 第3行：子表头
    ws['C3'] = '均值'
    ws['D3'] = '最小值'
    ws['E3'] = '最大值'
    ws['F3'] = '样点数'
    ws['G3'] = '样点占比'
    
    for col in ['C', 'D', 'E', 'F', 'G']:
        ws[f'{col}3'].font = Font(bold=True)
        ws[f'{col}3'].alignment = Alignment(horizontal='center', vertical='center')
    
    # 固定的土地利用类型结构
    land_structure = [
        ('耕地', ['水田', '水浇地', '旱地']),
        ('园地', ['果园', '茶园', '其他园地']),
        ('林地', []),
        ('草地', []),
        ('其他', [])
    ]
    
    # 写入数据（从第4行开始）
    current_row = 4
    total_samples = len(df_clean)
    
    for primary_type, secondaries in land_structure:
        primary_df = df_clean[df_clean['一级地类'] == primary_type]
        
        # 一级地类的起始行
        primary_start_row = current_row
        
        # 如果有二级地类，先处理二级地类，然后汇总到一级地类
        if secondaries:
            # 用于汇总二级地类数据
            primary_samples_sum = 0
            primary_values = []
            
            # 先写入一级地类行（后面会填充数据）
            primary_row = current_row
            current_row += 1  # 预留一行给一级地类
            
            # 记录二级地类开始行
            secondary_start_row = current_row
            
            # 处理每个二级地类
            for secondary_type in secondaries:
                secondary_df = primary_df[primary_df['二级地类'] == secondary_type]
                
                if not secondary_df.empty:
                    mean_value = secondary_df[attr_key].mean()
                    min_value = secondary_df[attr_key].min()
                    max_value = secondary_df[attr_key].max()
                    sample_count = len(secondary_df)
                    sample_ratio = (sample_count / total_samples * 100) if total_samples > 0 else 0
                    
                    # 累加到一级地类总和
                    primary_samples_sum += sample_count
                    primary_values.extend(secondary_df[attr_key].tolist())
                    
                    # 写入二级地类数据
                    ws.cell(current_row, 1, '')
                    ws.cell(current_row, 2, secondary_type)
                    ws.cell(current_row, 3, format_small_value(mean_value))
                    ws.cell(current_row, 4, format_small_value(min_value))
                    ws.cell(current_row, 5, format_small_value(max_value))
                    ws.cell(current_row, 6, sample_count)
                    ws.cell(current_row, 7, format_percentage(sample_ratio))
                    
                    current_row += 1
                else:
                    # 即使没有数据也要写出结构
                    ws.cell(current_row, 1, '')
                    ws.cell(current_row, 2, secondary_type)
                    ws.cell(current_row, 3, 0)
                    ws.cell(current_row, 4, 0)
                    ws.cell(current_row, 5, 0)
                    ws.cell(current_row, 6, 0)
                    ws.cell(current_row, 7, '0.00%')
                    
                    current_row += 1
            
            # 现在回填一级地类的数据（二级地类的总和）
            if primary_samples_sum > 0 and primary_values:
                primary_mean = sum(primary_values) / len(primary_values)
                primary_min = min(primary_values)
                primary_max = max(primary_values)
                primary_ratio = (primary_samples_sum / total_samples * 100) if total_samples > 0 else 0
                
                ws.cell(primary_row, 3, format_small_value(primary_mean))
                ws.cell(primary_row, 4, format_small_value(primary_min))
                ws.cell(primary_row, 5, format_small_value(primary_max))
                ws.cell(primary_row, 6, primary_samples_sum)
                ws.cell(primary_row, 7, format_percentage(primary_ratio))
            else:
                ws.cell(primary_row, 3, 0)
                ws.cell(primary_row, 4, 0)
                ws.cell(primary_row, 5, 0)
                ws.cell(primary_row, 6, 0)
                ws.cell(primary_row, 7, '0.00%')
            
            # 合并A-B列（一级地类名称）
            ws.merge_cells(f'A{primary_row}:B{primary_row}')
            ws.cell(primary_row, 1, primary_type)
            ws.cell(primary_row, 1).alignment = Alignment(horizontal='center', vertical='center')
            
            # 合并A列（一级地类跨越所有二级地类）
            if current_row > secondary_start_row:
                ws.merge_cells(f'A{secondary_start_row}:A{current_row-1}')
                ws.cell(secondary_start_row, 1, primary_type)
                ws.cell(secondary_start_row, 1).alignment = Alignment(horizontal='center', vertical='center')
        else:
            # 没有二级地类，直接统计一级地类
            if not primary_df.empty:
                mean_value = primary_df[attr_key].mean()
                min_value = primary_df[attr_key].min()
                max_value = primary_df[attr_key].max()
                sample_count = len(primary_df)
                sample_ratio = (sample_count / total_samples * 100) if total_samples > 0 else 0
                
                ws.cell(current_row, 3, format_small_value(mean_value))
                ws.cell(current_row, 4, format_small_value(min_value))
                ws.cell(current_row, 5, format_small_value(max_value))
                ws.cell(current_row, 6, sample_count)
                ws.cell(current_row, 7, format_percentage(sample_ratio))
            else:
                ws.cell(current_row, 3, 0)
                ws.cell(current_row, 4, 0)
                ws.cell(current_row, 5, 0)
                ws.cell(current_row, 6, 0)
                ws.cell(current_row, 7, '0.00%')
            
            # 合并A-B列（一级地类名称）
            ws.merge_cells(f'A{current_row}:B{current_row}')
            ws.cell(current_row, 1, primary_type)
            ws.cell(current_row, 1).alignment = Alignment(horizontal='center', vertical='center')
            
            current_row += 1
    
    # 写入全域统计
    total_mean = df_clean[attr_key].mean()
    total_min = df_clean[attr_key].min()
    total_max = df_clean[attr_key].max()
    
    ws.merge_cells(f'A{current_row}:B{current_row}')
    ws.cell(current_row, 1, '全域')
    ws.cell(current_row, 1).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(current_row, 3, format_small_value(total_mean))
    ws.cell(current_row, 4, format_small_value(total_min))
    ws.cell(current_row, 5, format_small_value(total_max))
    ws.cell(current_row, 6, total_samples)
    ws.cell(current_row, 7, '100')
    
    # 设置边框和对齐
    thin = Side(border_style="thin")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    
    for row in ws.iter_rows(min_row=2, max_row=current_row, min_col=1, max_col=7):
        for cell in row:
            cell.border = border
            if cell.alignment.horizontal is None:
                cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 调整列宽
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 12
    for col in ['C', 'D', 'E', 'F', 'G']:
        ws.column_dimensions[col].width = 12


def generate_soil_type_summary(ws, df, attr_key):
    """
    生成土壤类型统计表格
    """
    config = SOIL_ATTR_CONFIG[attr_key]
    attr_name = config['name']
    unit = config['unit']
    ws.title = f"{attr_name}土壤类型统计"
    
    # 准备数据
    df_clean = df.copy()
    df_clean[attr_key] = pd.to_numeric(df_clean[attr_key], errors='coerce')
    df_clean = df_clean[df_clean[attr_key] > 0].copy()
    
    # 确保土壤类型列存在（支持TL/土类、YL/亚类、TS/土属）
    # 统一列名为中文
    column_mapping = {}
    if 'TL' in df_clean.columns:
        column_mapping['TL'] = '土类'
    if 'YL' in df_clean.columns:
        column_mapping['YL'] = '亚类'
    if 'TS' in df_clean.columns:
        column_mapping['TS'] = '土属'
    
    if column_mapping:
        df_clean = df_clean.rename(columns=column_mapping)
    
    # 确保列存在
    for col in ['土类', '亚类', '土属']:
        if col not in df_clean.columns:
            df_clean[col] = ''
    
    # 分级
    df_clean['等级'] = df_clean[attr_key].apply(lambda x: classify_by_config(x, attr_key))
    
    # 获取等级列表
    grade_order = get_grade_order(attr_key)
    
    # 获取分级范围
    grade_ranges = get_grade_ranges(attr_key)
    
    # 写入标题
    ws.merge_cells('A1:J1')
    ws['A1'] = f'{attr_name}土壤类型统计'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # 写入表头（合并A2:A3, B2:B3, C2:C3, I2:I3）
    ws.merge_cells('A2:A3')
    ws['A2'] = '土类'
    ws['A2'].font = Font(bold=True)
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws.merge_cells('B2:B3')
    ws['B2'] = '亚类'
    ws['B2'].font = Font(bold=True)
    ws['B2'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws.merge_cells('C2:C3')
    ws['C2'] = '土属'
    ws['C2'].font = Font(bold=True)
    ws['C2'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws['D2'] = 'Ⅰ级'
    ws['E2'] = 'Ⅱ级'
    ws['F2'] = 'Ⅲ级'
    ws['G2'] = 'Ⅳ级'
    ws['H2'] = 'Ⅴ级'
    
    # 合并I2:I3显示"平均等级"
    ws.merge_cells('I2:I3')
    ws['I2'] = '平均等级'
    ws['I2'].font = Font(bold=True)
    ws['I2'].alignment = Alignment(horizontal='center', vertical='center')
    
    # 设置第3行的分级范围（从配置自动生成）
    ws['D3'] = grade_ranges.get('Ⅰ级', '')
    ws['E3'] = grade_ranges.get('Ⅱ级', '')
    ws['F3'] = grade_ranges.get('Ⅲ级', '')
    ws['G3'] = grade_ranges.get('Ⅳ级', '')
    ws['H3'] = grade_ranges.get('Ⅴ级', '')
    
    # 格式化表头
    for col in ['D', 'E', 'F', 'G', 'H']:
        ws[f'{col}2'].font = Font(bold=True)
        ws[f'{col}3'].font = Font(bold=True)
        ws[f'{col}2'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'{col}3'].alignment = Alignment(horizontal='center', vertical='center')
    
    # 写入数据
    current_row = 4
    
    # 使用字典收集每个土类-亚类-土属组合的数据
    soil_stats = {}
    
    # 逐行遍历数据
    for idx, row in df_clean.iterrows():
        major_type = row['土类'] if pd.notna(row['土类']) and row['土类'] != '' else '未分类'
        sub_type = row['亚类']
        genus = row['土属']
        
        # 只要亚类和土属有值就统计（土类可以为空或默认为'未分类'）
        if pd.isna(sub_type) or sub_type == '' or \
           pd.isna(genus) or genus == '':
            continue
        
        # 创建组合键
        key = (major_type, sub_type, genus)
        
        # 初始化该组合的统计数据
        if key not in soil_stats:
            soil_stats[key] = {
                '土类': major_type,
                '亚类': sub_type,
                '土属': genus,
                'data': []  # 存储该组合的所有数据行
            }
        
        # 添加数据到该组合
        soil_stats[key]['data'].append(row)
    
    # 对每个组合计算统计指标
    results = []
    for key, stats in soil_stats.items():
        major_type = stats['土类']
        sub_type = stats['亚类']
        genus = stats['土属']
        data_rows = stats['data']
        
        # 将数据行转换为DataFrame进行统计
        group_df = pd.DataFrame(data_rows)
        
        # 统计各级别面积
        grade_counts = group_df.groupby('等级')['面积'].sum().reindex(grade_order, fill_value=0)
        total_area = grade_counts.sum()
        avg_grade = calculate_average_grade(group_df, attr_key)
        
        results.append({
            '土类': major_type,
            '亚类': sub_type,
            '土属': genus,
            'grade_counts': grade_counts,
            'total_area': total_area,
            'avg_grade': avg_grade
        })
    
    # 按照土类→亚类→土属的顺序排序
    results_sorted = sorted(results, key=lambda x: (x['土类'], x['亚类'], x['土属']))
    
    # 先写入数据，后合并单元格
    merge_ranges = []  # 记录需要合并的范围
    
    # 记录每个土类和亚类的起始和结束行
    major_type_ranges = {}  # {'土类名': (start_row, end_row)}
    sub_type_ranges = {}    # {('土类名', '亚类名'): (start_row, end_row)}
    
    current_major = None
    current_sub = None
    major_start_row = None
    sub_start_row = None
    
    # 写入排序后的结果
    for idx, result in enumerate(results_sorted):
        major_type = result['土类']
        sub_type = result['亚类']
        genus = result['土属']
        grade_counts = result['grade_counts']
        total_area = result['total_area']
        avg_grade = result['avg_grade']
        
        # 跟踪土类变化
        if current_major != major_type:
            if current_major is not None and major_start_row is not None:
                # 记录上一个土类的范围
                major_type_ranges[current_major] = (major_start_row, current_row - 1)
            current_major = major_type
            major_start_row = current_row
        
        # 跟踪亚类变化
        if current_sub != (major_type, sub_type):
            if current_sub is not None and sub_start_row is not None:
                # 记录上一个亚类的范围
                sub_type_ranges[current_sub] = (sub_start_row, current_row - 1)
            current_sub = (major_type, sub_type)
            sub_start_row = current_row
        
        # 写入面积行
        ws.cell(current_row, 1, major_type)
        ws.cell(current_row, 2, sub_type)
        ws.cell(current_row, 3, genus)  # 土属名称
        for i, grade in enumerate(grade_order, start=4):
            ws.cell(current_row, i, format_small_value(grade_counts[grade]))
        ws.cell(current_row, 9, avg_grade if avg_grade is not None else '')  # 平均等级
        
        # 写入占比行
        ws.cell(current_row+1, 1, '')
        ws.cell(current_row+1, 2, '')
        ws.cell(current_row+1, 3, '')  # 土属占比行为空
        for i, grade in enumerate(grade_order, start=4):
            pct = (grade_counts[grade] / total_area * 100) if total_area > 0 else 0
            ws.cell(current_row+1, i, format_percentage(pct))
        ws.cell(current_row+1, 9, '')  # 平均等级占比行为空
        
        # 合并土属的2行（C列）
        ws.merge_cells(f'C{current_row}:C{current_row+1}')
        ws[f'C{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
        
        # 合并平均等级的2行（I列）
        ws.merge_cells(f'I{current_row}:I{current_row+1}')
        ws[f'I{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
        
        current_row += 2
    
    # 记录最后一个土类和亚类的范围
    if current_major is not None and major_start_row is not None:
        major_type_ranges[current_major] = (major_start_row, current_row - 1)
    if current_sub is not None and sub_start_row is not None:
        sub_type_ranges[current_sub] = (sub_start_row, current_row - 1)
    
    # 合并土类单元格（A列）
    for major_type, (start_row, end_row) in major_type_ranges.items():
        if end_row > start_row:  # 只有多行才合并
            ws.merge_cells(f'A{start_row}:A{end_row}')
            ws[f'A{start_row}'].alignment = Alignment(horizontal='center', vertical='center')
    
    # 合并亚类单元格（B列）
    for (major_type, sub_type), (start_row, end_row) in sub_type_ranges.items():
        if end_row > start_row:  # 只有多行才合并
            ws.merge_cells(f'B{start_row}:B{end_row}')
            ws[f'B{start_row}'].alignment = Alignment(horizontal='center', vertical='center')
    
    # 写入全域统计
    total_df = df_clean
    total_grade_counts = total_df.groupby('等级')['面积'].sum().reindex(grade_order, fill_value=0)
    total_total_area = total_grade_counts.sum()
    total_avg_grade = calculate_average_grade(total_df, attr_key)
    
    # 合并全域A-C列的3×2区域
    ws.merge_cells(f'A{current_row}:C{current_row+1}')
    ws[f'A{current_row}'] = '全域'
    ws[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
    
    # 写入全域面积数据
    for i, grade in enumerate(grade_order, start=4):
        ws.cell(current_row, i, format_small_value(total_grade_counts[grade]))
    ws.cell(current_row, 9, total_avg_grade if total_avg_grade is not None else '')  # 平均等级
    
    # 写入全域占比数据
    for i, grade in enumerate(grade_order, start=4):
        pct = (total_grade_counts[grade] / total_total_area * 100) if total_total_area > 0 else 0
        ws.cell(current_row+1, i, format_percentage(pct))
    ws.cell(current_row+1, 9, '')  # 平均等级占比行为空
    
    # 合并全域平均等级的2行（I列）
    ws.merge_cells(f'I{current_row}:I{current_row+1}')
    ws[f'I{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
    
    # 设置边框和对齐
    thin = Side(border_style="thin")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    
    for row in ws.iter_rows(min_row=1, max_row=current_row+1, min_col=1, max_col=9):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 调整列宽
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
        ws.column_dimensions[col].width = 12


def generate_soil_type_sample_summary(ws, df, attr_key):
    """
    生成土壤类型样点统计表格（第7张表）
    从样点数据读取，按亚类-土属层级展示
    """
    config = SOIL_ATTR_CONFIG[attr_key]
    attr_name = config['name']
    unit = config['unit']
    ws.title = f"{attr_name}土壤类型样点统计"
    
    # 准备数据
    df_clean = df.copy()
    df_clean[attr_key] = pd.to_numeric(df_clean[attr_key], errors='coerce')
    df_clean = df_clean[df_clean[attr_key] > 0].copy()
    
    if df_clean.empty:
        return
    
    # 确保土壤类型列存在（支持TL/土类、YL/亚类、TS/土属）
    column_mapping = {}
    if 'TL' in df_clean.columns:
        column_mapping['TL'] = '土类'
    if 'YL' in df_clean.columns:
        column_mapping['YL'] = '亚类'
    if 'TS' in df_clean.columns:
        column_mapping['TS'] = '土属'
    
    if column_mapping:
        df_clean = df_clean.rename(columns=column_mapping)
    
    # 确保列存在
    for col in ['土类', '亚类', '土属']:
        if col not in df_clean.columns:
            df_clean[col] = ''
    
    # 写入标题
    ws.merge_cells('A1:G1')
    ws['A1'] = f'{attr_name}土壤类型样点统计'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # 写入表头（第2-3行）
    ws.merge_cells('A2:B3')
    ws['A2'] = '土壤类型'
    ws['A2'].font = Font(bold=True)
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws.merge_cells('C2:G2')
    ws['C2'] = '样点统计'
    ws['C2'].font = Font(bold=True)
    ws['C2'].alignment = Alignment(horizontal='center', vertical='center')
    
    # 第3行：子表头
    ws['C3'] = '均值'
    ws['D3'] = '最小值'
    ws['E3'] = '最大值'
    ws['F3'] = '样点数'
    ws['G3'] = '样点占比'
    
    for col in ['C', 'D', 'E', 'F', 'G']:
        ws[f'{col}3'].font = Font(bold=True)
        ws[f'{col}3'].alignment = Alignment(horizontal='center', vertical='center')
    
    # 写入数据
    current_row = 4
    total_samples = len(df_clean)
    
    # 按亚类-土属组织数据
    soil_stats = {}
    
    # 逾行遍历数据
    for idx, row in df_clean.iterrows():
        sub_type = row['亚类']
        genus = row['土属']
        
        # 只要亚类和土属有值就统计
        if pd.isna(sub_type) or sub_type == '' or \
           pd.isna(genus) or genus == '':
            continue
        
        # 创建组合键
        key = (sub_type, genus)
        
        # 初始化该组合的统计数据
        if key not in soil_stats:
            soil_stats[key] = {
                '亚类': sub_type,
                '土属': genus,
                'data': []
            }
        
        # 添加数据到该组合
        soil_stats[key]['data'].append(row)
    
    # 对每个组合计算统计指标
    results = []
    for key, stats in soil_stats.items():
        sub_type = stats['亚类']
        genus = stats['土属']
        data_rows = stats['data']
        
        # 将数据行转换为DataFrame进行统计
        group_df = pd.DataFrame(data_rows)
        
        mean_value = group_df[attr_key].mean()
        min_value = group_df[attr_key].min()
        max_value = group_df[attr_key].max()
        sample_count = len(group_df)
        sample_ratio = (sample_count / total_samples * 100) if total_samples > 0 else 0
        
        results.append({
            '亚类': sub_type,
            '土属': genus,
            'mean_value': mean_value,
            'min_value': min_value,
            'max_value': max_value,
            'sample_count': sample_count,
            'sample_ratio': sample_ratio
        })
    
    # 按照亚类→土属的顺序排序
    results_sorted = sorted(results, key=lambda x: (x['亚类'], x['土属']))
    
    # 先写入数据，后合并单元格
    merge_ranges = []  # 记录需要合并的范围
    
    # 记录每个亚类的起始和结束行
    sub_type_ranges = {}  # {'亚类名': (start_row, end_row)}
    
    current_sub = None
    sub_start_row = None
    
    # 写入排序后的结果
    for idx, result in enumerate(results_sorted):
        sub_type = result['亚类']
        genus = result['土属']
        mean_value = result['mean_value']
        min_value = result['min_value']
        max_value = result['max_value']
        sample_count = result['sample_count']
        sample_ratio = result['sample_ratio']
        
        # 跟踪亚类变化
        if current_sub != sub_type:
            if current_sub is not None and sub_start_row is not None:
                # 记录上一个亚类的范围
                sub_type_ranges[current_sub] = (sub_start_row, current_row - 1)
            current_sub = sub_type
            sub_start_row = current_row
        
        # 写入数据行
        ws.cell(current_row, 1, sub_type)
        ws.cell(current_row, 2, genus)
        ws.cell(current_row, 3, format_small_value(mean_value))
        ws.cell(current_row, 4, format_small_value(min_value))
        ws.cell(current_row, 5, format_small_value(max_value))
        ws.cell(current_row, 6, sample_count)
        ws.cell(current_row, 7, format_percentage(sample_ratio))
        
        current_row += 1
    
    # 记录最后一个亚类的范围
    if current_sub is not None and sub_start_row is not None:
        sub_type_ranges[current_sub] = (sub_start_row, current_row - 1)
    
    # 合并亚类单元格（A列）
    for sub_type, (start_row, end_row) in sub_type_ranges.items():
        if end_row > start_row:  # 只有多行才合并
            ws.merge_cells(f'A{start_row}:A{end_row}')
            ws[f'A{start_row}'].alignment = Alignment(horizontal='center', vertical='center')
    
    # 写入全域统计
    total_mean = df_clean[attr_key].mean()
    total_min = df_clean[attr_key].min()
    total_max = df_clean[attr_key].max()
    
    # 合并全域A-B列
    ws.merge_cells(f'A{current_row}:B{current_row}')
    ws[f'A{current_row}'] = '全域'
    ws[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws.cell(current_row, 3, format_small_value(total_mean))
    ws.cell(current_row, 4, format_small_value(total_min))
    ws.cell(current_row, 5, format_small_value(total_max))
    ws.cell(current_row, 6, total_samples)
    ws.cell(current_row, 7, '100')
    
    # 设置边框和对齐
    thin = Side(border_style="thin")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    
    for row in ws.iter_rows(min_row=2, max_row=current_row, min_col=1, max_col=7):
        for cell in row:
            cell.border = border
            if cell.alignment.horizontal is None:
                cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 调整列宽
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        ws.column_dimensions[col].width = 12


def generate_overall_statistics_summary(wb, df_sample):
    """
    生成全域属性统计汇怺表（第8张表）
    将所有土壤属性的统计指标汇怺在一张表中，只生成一次
    """
    ws = wb.create_sheet()
    ws.title = "全域属性统计汇怺"
    
    # 写入标题
    ws.merge_cells('A1:I1')
    ws['A1'] = '全域土壤属性统计汇怺'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # 写入表头
    headers = ['土壤属性', '样本数', '最小值', '最大值', '极差', '中位数', '平均值', '标准差', '变异系数']
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(2, col_idx, header)
        ws.cell(2, col_idx).font = Font(bold=True)
        ws.cell(2, col_idx).alignment = Alignment(horizontal='center', vertical='center')
    
    # 写入数据
    current_row = 3
    
    # 遍历所有配置的属性
    for attr_key in SOIL_ATTR_CONFIG.keys():
        config = SOIL_ATTR_CONFIG[attr_key]
        attr_name = config['name']
        unit = config['unit']
        
        # 检查该属性是否在数据中
        if attr_key not in df_sample.columns:
            continue
        
        # 准备数据
        df_clean = df_sample.copy()
        df_clean[attr_key] = pd.to_numeric(df_clean[attr_key], errors='coerce')
        df_clean = df_clean[df_clean[attr_key] > 0].copy()
        
        if df_clean.empty:
            continue
        
        # 计算统计指标
        sample_count = len(df_clean)
        min_value = df_clean[attr_key].min()
        max_value = df_clean[attr_key].max()
        range_value = max_value - min_value  # 极差
        median_value = df_clean[attr_key].median()
        mean_value = df_clean[attr_key].mean()
        std_value = df_clean[attr_key].std()  # 标准差
        cv_value = (std_value / mean_value) if mean_value != 0 else 0  # 变异系数
        
        # 属性名称（包含单位）
        if unit:
            display_name = f'{attr_name}（{unit}）'
        else:
            display_name = attr_name
        
        # 写入数据行
        ws.cell(current_row, 1, display_name)
        ws.cell(current_row, 2, sample_count)
        ws.cell(current_row, 3, format_small_value(min_value))
        ws.cell(current_row, 4, format_small_value(max_value))
        ws.cell(current_row, 5, format_small_value(range_value))
        ws.cell(current_row, 6, format_small_value(median_value))
        ws.cell(current_row, 7, format_small_value(mean_value))
        ws.cell(current_row, 8, format_small_value(std_value))
        ws.cell(current_row, 9, format_small_value(cv_value))
        
        current_row += 1
    
    # 设置边框和对齐
    thin = Side(border_style="thin")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    
    for row in ws.iter_rows(min_row=2, max_row=current_row-1, min_col=1, max_col=9):
        for cell in row:
            cell.border = border
            if cell.column > 1:  # 数值列居中
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:  # 属性名称列左对齐
                cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # 调整列宽
    ws.column_dimensions['A'].width = 20
    for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
        ws.column_dimensions[col].width = 12


def generate_percentile_statistics_summary(wb, df_sample):
    """
    生成全域属性百分位数统计表（第9张表）
    统计各个属性在2%、5%、10%、20%、80%、90%、95%、98%百分位数的值，只生成一次
    """
    ws = wb.create_sheet()
    ws.title = "全域属性百分位数统计"
    
    # 写入标题
    ws.merge_cells('A1:I1')
    ws['A1'] = '全域土壤属性百分位数统计'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # 写入表头
    headers = ['土壤属性', '2%', '5%', '10%', '20%', '80%', '90%', '95%', '98%']
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(2, col_idx, header)
        ws.cell(2, col_idx).font = Font(bold=True)
        ws.cell(2, col_idx).alignment = Alignment(horizontal='center', vertical='center')
    
    # 写入数据
    current_row = 3
    
    # 百分位数列表
    percentiles = [0.02, 0.05, 0.10, 0.20, 0.80, 0.90, 0.95, 0.98]
    
    # 遍历所有配置的属性
    for attr_key in SOIL_ATTR_CONFIG.keys():
        config = SOIL_ATTR_CONFIG[attr_key]
        attr_name = config['name']
        unit = config['unit']
        
        # 检查该属性是否在数据中
        if attr_key not in df_sample.columns:
            continue
        
        # 准备数据
        df_clean = df_sample.copy()
        df_clean[attr_key] = pd.to_numeric(df_clean[attr_key], errors='coerce')
        df_clean = df_clean[df_clean[attr_key] > 0].copy()
        
        if df_clean.empty:
            continue
        
        # 属性名称（包含单位）
        if unit:
            display_name = f'{attr_name}（{unit}）'
        else:
            display_name = attr_name
        
        # 写入属性名称
        ws.cell(current_row, 1, display_name)
        
        # 计算并写入各个百分位数
        for col_idx, percentile in enumerate(percentiles, start=2):
            percentile_value = df_clean[attr_key].quantile(percentile)
            ws.cell(current_row, col_idx, format_small_value(percentile_value))
        
        current_row += 1
    
    # 设置边框和对齐
    thin = Side(border_style="thin")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    
    for row in ws.iter_rows(min_row=2, max_row=current_row-1, min_col=1, max_col=9):
        for cell in row:
            cell.border = border
            if cell.column > 1:  # 数值列居中
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:  # 属性名称列左对齐
                cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # 调整列宽
    ws.column_dimensions['A'].width = 20
    for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
        ws.column_dimensions[col].width = 12


def read_data_file(file):
    """
    读取数据文件（CSV或Excel）
    """
    if file.endswith('.csv'):
        try:
            # 检测编码
            with open(file, 'rb') as f:
                result = chardet.detect(f.read())
            detected_encoding = result['encoding']
            print(f"  检测到编码: {detected_encoding}")
            # 使用检测到的编码读取文件
            df = pd.read_csv(file, encoding=detected_encoding)
            print(f"  成功读取 (使用 {detected_encoding} 编码), 数据行数: {len(df)}")
        except UnicodeDecodeError as e:
            print(f"  [警告] 使用 {detected_encoding} 编码失败: {str(e)}")
            # 如果检测到的编码失败，尝试使用utf-8-sig编码
            try:
                df = pd.read_csv(file, encoding='utf-8-sig')
                print(f"  成功读取 (使用 utf-8-sig 编码), 数据行数: {len(df)}")
            except UnicodeDecodeError as e2:
                print(f"  [警告] 使用 utf-8-sig 编码失败: {str(e2)}")
                # 如果utf-8-sig也失败，尝试使用gbk编码
                df = pd.read_csv(file, encoding='gbk')
                print(f"  成功读取 (使用 gbk 编码), 数据行数: {len(df)}")
        except Exception as e:
            print(f"  [错误] 读取CSV文件失败: {type(e).__name__}: {str(e)}")
            raise
    else:
        df = pd.read_excel(file)
        print(f"  成功读取Excel文件, 数据行数: {len(df)}")
    return df


def main():
    """
    主函数
    """
    # 创建GUI
    root = tk.Tk()
    root.title("土壤属性数据报告生成器")
    root.geometry("700x500")
    
    def select_mapping_files():
        """选择制图数据文件"""
        files = filedialog.askopenfilenames(
            title="选择制图数据文件",
            filetypes=[("Excel文件", "*.xlsx *.xls"), ("CSV文件", "*.csv"), ("所有文件", "*.*")]
        )
        mapping_file_list.delete(0, tk.END)
        for file in files:
            mapping_file_list.insert(tk.END, file)
    
    def select_sample_files():
        """选择样点数据文件"""
        files = filedialog.askopenfilenames(
            title="选择样点数据文件",
            filetypes=[("Excel文件", "*.xlsx *.xls"), ("CSV文件", "*.csv"), ("所有文件", "*.*")]
        )
        sample_file_list.delete(0, tk.END)
        for file in files:
            sample_file_list.insert(tk.END, file)
    
    def generate_report():
        """生成报告"""
        mapping_files = mapping_file_list.get(0, tk.END)
        sample_files = sample_file_list.get(0, tk.END)
        
        if not mapping_files and not sample_files:
            messagebox.showerror("错误", "请至少选择制图数据或样点数据文件")
            return
        
        # 自动生成带时间戳的文件名
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        default_filename = f"数据报告_{timestamp}.xlsx"
        # 选择输出文件
        output_file = filedialog.asksaveasfilename(
            title="保存报告",
            defaultextension=".xlsx",
            filetypes=[("Excel文件", "*.xlsx")],
            initialfile=default_filename
        )
        if not output_file:
            return
        
        try:
            print("\n========== 开始生成报告 ==========")
            print(f"输出文件: {output_file}")
            
            # 读取制图数据
            df_mapping = None
            if mapping_files:
                print(f"\n读取制图数据文件列表:")
                for idx, file in enumerate(mapping_files, 1):
                    print(f"  {idx}. {file}")
                
                dfs_mapping = []
                for file in mapping_files:
                    print(f"\n正在读取: {file}")
                    df = read_data_file(file)
                    dfs_mapping.append(df)
                
                # 合并数据
                print(f"\n合并制图数据...")
                df_mapping = pd.concat(dfs_mapping, ignore_index=True)
                print(f"合并后总行数: {len(df_mapping)}")
                print(f"数据列: {list(df_mapping.columns)}")
            
            # 读取样点数据
            df_sample = None
            if sample_files:
                print(f"\n读取样点数据文件列表:")
                for idx, file in enumerate(sample_files, 1):
                    print(f"  {idx}. {file}")
                
                dfs_sample = []
                for file in sample_files:
                    print(f"\n正在读取: {file}")
                    df = read_data_file(file)
                    dfs_sample.append(df)
                
                # 合并数据
                print(f"\n合并样点数据...")
                df_sample = pd.concat(dfs_sample, ignore_index=True)
                print(f"合并后总行数: {len(df_sample)}")
                print(f"数据列: {list(df_sample.columns)}")
            
            # 创建Excel工作簿
            print(f"\n创建Excel工作簿...")
            wb = Workbook()
            
            # 为每个属性生成表格
            print(f"\n开始生成统计表...")
            processed_any = False
            total_attrs = len(SOIL_ATTR_CONFIG)
            
            for attr_idx, attr_key in enumerate(SOIL_ATTR_CONFIG.keys()):
                config = SOIL_ATTR_CONFIG[attr_key]
                attr_name = config['name']
                
                print(f"\n处理属性 ({attr_idx+1}/{total_attrs}): {attr_name} ({attr_key})")
                
                try:
                    # 应用用地类型过滤
                    land_filter = config.get('land_filter')
                    if land_filter:
                        filter_desc = {
                            'cultivated_garden': '耕园地',
                            'paddy_only': '水田',
                            'cultivated_only': '耕地'
                        }
                        print(f"  应用用地过滤: 仅统计{filter_desc.get(land_filter, land_filter)}")
                    
                    # 处理制图数据（生成1-3张表）
                    if df_mapping is not None and attr_key in df_mapping.columns:
                        df_filtered = apply_land_filter(df_mapping.copy(), attr_key)
                        
                        # 检查是否有有效数据
                        df_filtered[attr_key] = pd.to_numeric(df_filtered[attr_key], errors='coerce')
                        valid_count = len(df_filtered[df_filtered[attr_key] > 0])
                        if valid_count > 0:
                            print(f"  制图数据有效行数: {valid_count}")
                            
                            # 创建工作表
                            print(f"  生成分乡镇统计表...")
                            ws1 = wb.create_sheet()
                            generate_town_summary(ws1, df_filtered, attr_key)
                            print(f"    ✓ 完成")
                            
                            print(f"  生成土地利用类型统计表...")
                            ws2 = wb.create_sheet()
                            generate_land_use_summary(ws2, df_filtered, attr_key)
                            print(f"    ✓ 完成")
                            
                            print(f"  生成土壤类型统计表...")
                            ws3 = wb.create_sheet()
                            generate_soil_type_summary(ws3, df_filtered, attr_key)
                            print(f"    ✓ 完成")
                            
                            processed_any = True
                        else:
                            print(f"  [警告] 制图数据无有效数据")
                    else:
                        if df_mapping is not None:
                            print(f"  [警告] 制图数据中未找到属性列: {attr_key}")
                    
                    # 处理样点数据（生成第4、5、6、7张表）
                    if df_sample is not None and attr_key in df_sample.columns:
                        df_sample_filtered = df_sample.copy()
                        
                        # 检查是否有有效数据
                        df_sample_filtered[attr_key] = pd.to_numeric(df_sample_filtered[attr_key], errors='coerce')
                        valid_count = len(df_sample_filtered[df_sample_filtered[attr_key] > 0])
                        if valid_count > 0:
                            print(f"  样点数据有效行数: {valid_count}")
                            
                            print(f"  生成样点统计表（第4张表）...")
                            ws4 = wb.create_sheet()
                            generate_sample_point_summary(ws4, df_sample_filtered, attr_key)
                            print(f"    ✓ 完成")
                            
                            print(f"  生成分行政区样点统计表（第5张表）...")
                            ws5 = wb.create_sheet()
                            generate_town_sample_summary(ws5, df_sample_filtered, attr_key)
                            print(f"    ✓ 完成")
                            
                            print(f"  生成土地利用类型样点统计表（第6张表）...")
                            ws6 = wb.create_sheet()
                            generate_land_use_sample_summary(ws6, df_sample_filtered, attr_key)
                            print(f"    ✓ 完成")
                            
                            print(f"  生成土壤类型样点统计表（第7张表）...")
                            ws7 = wb.create_sheet()
                            generate_soil_type_sample_summary(ws7, df_sample_filtered, attr_key)
                            print(f"    ✓ 完成")
                            
                            processed_any = True
                        else:
                            print(f"  [警告] 样点数据无有效数据")
                    else:
                        if df_sample is not None:
                            print(f"  [警告] 样点数据中未找到属性列: {attr_key}")
                    
                except Exception as e:
                    # 容错机制：单个属性失败不影响其他属性（参考土壤属性多文件分析.py）
                    tb = traceback.extract_tb(e.__traceback__)
                    loc = ''
                    code_line = ''
                    if tb:
                        last = tb[-1]
                        loc = f"位置: {last.filename}:{last.lineno} {last.name}()"
                        code_line = linecache.getline(last.filename, last.lineno).strip()
                    err_type = type(e).__name__
                    print(f"  [错误] 处理失败:")
                    print(f"    错误类型: {err_type}")
                    print(f"    错误信息: {e}")
                    print(f"    {loc}")
                    if code_line:
                        print(f"    代码: {code_line}")
                    print(f"  → 继续处理其他属性...")
                    continue
            
            if not processed_any:
                raise RuntimeError("所有检测到的属性均无法处理（缺少数据或格式错误）")
            
            # 生成第8张表：全域属性统计汇怺表（只生成一次）
            if df_sample is not None:
                print(f"\n生成全域属性统计汇怺表（第8张表）...")
                try:
                    generate_overall_statistics_summary(wb, df_sample)
                    print(f"  ✓ 完成")
                except Exception as e:
                    print(f"  [错误] 生成失败: {str(e)}")
                
                print(f"\n生成全域属性百分位数统计表（第9张表）...")
                try:
                    generate_percentile_statistics_summary(wb, df_sample)
                    print(f"  ✓ 完成")
                except Exception as e:
                    print(f"  [错误] 生成失败: {str(e)}")
            
            # 删除默认工作表
            if 'Sheet' in wb.sheetnames:
                del wb['Sheet']
                print(f"\n删除默认工作表")
            
            # 保存文件
            print(f"\n保存Excel文件: {output_file}")
            wb.save(output_file)
            print(f"\n========== 报告生成成功 ==========\n")
            messagebox.showinfo("成功", f"报告已生成：{output_file}")
        except Exception as e:
            error_msg = f"生成报告时出错:\n\n错误类型: {type(e).__name__}\n错误信息: {str(e)}\n\n详细错误位置:\n{traceback.format_exc()}"
            print(f"\n========== 错误 ==========\n{error_msg}\n========================\n")
            messagebox.showerror("错误", f"生成报告时出错：{str(e)}\n\n详细信息已输出到终端")
    
    # GUI布局
    frame = ttk.Frame(root, padding="10")
    frame.pack(fill=tk.BOTH, expand=True)
    
    # 制图数据选择区域
    ttk.Label(frame, text="制图数据文件（用于生成表1-3）：").pack(anchor=tk.W)
    mapping_file_list = tk.Listbox(frame, height=5, selectmode=tk.MULTIPLE)
    mapping_file_list.pack(fill=tk.BOTH, expand=True, pady=5)
    
    button_frame1 = ttk.Frame(frame)
    button_frame1.pack(fill=tk.X, pady=5)
    ttk.Button(button_frame1, text="选择制图数据", command=select_mapping_files).pack(side=tk.LEFT, padx=5)
    
    # 样点数据选择区域
    ttk.Label(frame, text="样点数据文件（用于生成表4-9）：").pack(anchor=tk.W, pady=(10, 0))
    sample_file_list = tk.Listbox(frame, height=5, selectmode=tk.MULTIPLE)
    sample_file_list.pack(fill=tk.BOTH, expand=True, pady=5)
    
    button_frame2 = ttk.Frame(frame)
    button_frame2.pack(fill=tk.X, pady=5)
    ttk.Button(button_frame2, text="选择样点数据", command=select_sample_files).pack(side=tk.LEFT, padx=5)
    
    # 生成报告按钮
    button_frame3 = ttk.Frame(frame)
    button_frame3.pack(fill=tk.X, pady=10)
    ttk.Button(button_frame3, text="生成报告", command=generate_report).pack(side=tk.RIGHT, padx=5)
    
    root.mainloop()


if __name__ == "__main__":
    print("土壤属性数据报告生成器 v1.0")
    print("="*50)
    try:
        main()
    except Exception as e:
        print(f"\n程序异常退出:")
        print(f"错误类型: {type(e).__name__}")
        print(f"错误信息: {str(e)}")
        print(f"\n完整错误堆栈:\n{traceback.format_exc()}")
        sys.exit(1)