"""Excel输出生成模块

负责将统计结果写入Excel工作簿。
"""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.topics.data_report.classifiers import format_percentage, format_small_value
from app.topics.data_report.config import (
    SOIL_ATTR_CONFIG,
    get_grade_order,
    get_grade_ranges,
)
from app.topics.data_report.stats import AttributeStatsSummary

# 通用样式
THIN_BORDER = Border(
    top=Side(border_style="thin"),
    left=Side(border_style="thin"),
    right=Side(border_style="thin"),
    bottom=Side(border_style="thin"),
)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
HEADER_FONT = Font(bold=True)
TITLE_FONT = Font(bold=True, size=14)


def _apply_border_and_align(
    ws: Worksheet,
    min_row: int,
    max_row: int,
    min_col: int,
    max_col: int,
) -> None:
    """应用边框和居中对齐"""
    for row in ws.iter_rows(
        min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col
    ):
        for cell in row:
            cell.border = THIN_BORDER
            if cell.alignment.horizontal is None:
                cell.alignment = CENTER_ALIGN


def _set_column_widths(ws: Worksheet, widths: dict[str, int]) -> None:
    """设置列宽"""
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def write_town_summary(
    ws: Worksheet,
    summary: AttributeStatsSummary,
) -> None:
    """写入分乡镇统计表

    Args:
        ws: 工作表
        summary: 属性统计汇总
    """
    grade_order = get_grade_order(summary.attr_key)
    grade_ranges = get_grade_ranges(summary.attr_key)

    ws.title = f"{summary.attr_name}分乡镇统计"

    # 标题 - 先写数据
    ws["A1"] = f"{summary.attr_name}分乡镇统计"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = CENTER_ALIGN

    # 表头 - 先写数据
    ws["A2"] = "行政县"
    ws["A2"].font = HEADER_FONT
    ws["A2"].alignment = CENTER_ALIGN

    ws["C2"] = "分级"
    ws["C2"].font = HEADER_FONT
    ws["C2"].alignment = CENTER_ALIGN

    ws["H2"] = "指标平均等级"
    ws["H2"].font = HEADER_FONT
    ws["H2"].alignment = CENTER_ALIGN

    # 等级名称
    for i, grade in enumerate(grade_order[:5], start=3):
        col = get_column_letter(i)
        ws[f"{col}3"] = grade.replace("级", "")
        ws[f"{col}3"].font = HEADER_FONT
        ws[f"{col}3"].alignment = CENTER_ALIGN
        ws[f"{col}4"] = grade_ranges.get(grade, "")
        ws[f"{col}4"].font = HEADER_FONT
        ws[f"{col}4"].alignment = CENTER_ALIGN

    # 数据行 - 先写入所有数据，记录需要合并的行
    current_row = 5
    merge_rows: list[int] = []  # 记录需要合并A列和H列的行

    for town_stats in summary.town_stats:
        # 面积行
        ws.cell(current_row, 1, town_stats.town)
        ws.cell(current_row, 2, "面积")

        for i, grade in enumerate(grade_order[:5], start=3):
            gs = town_stats.grade_stats.get(grade)
            ws.cell(current_row, i, format_small_value(gs.area) if gs else 0)

        ws.cell(
            current_row,
            8,
            town_stats.avg_grade if town_stats.avg_grade is not None else "",
        )

        # 记录需要合并的行
        merge_rows.append(current_row)

        # 占比行
        ws.cell(current_row + 1, 2, "占比")
        for i, grade in enumerate(grade_order[:5], start=3):
            gs = town_stats.grade_stats.get(grade)
            ws.cell(current_row + 1, i, format_percentage(gs.percentage) if gs else 0)

        current_row += 2

    # 全域统计 - 写入数据
    global_row = current_row
    ws.cell(global_row, 1, "全域")
    ws.cell(global_row, 2, "面积")
    for i, grade in enumerate(grade_order[:5], start=3):
        gs = summary.grade_stats.get(grade)
        ws.cell(global_row, i, format_small_value(gs.area) if gs else 0)
    ws.cell(
        global_row,
        8,
        summary.global_avg_grade if summary.global_avg_grade is not None else "",
    )

    ws.cell(global_row + 1, 2, "占比")
    for i, grade in enumerate(grade_order[:5], start=3):
        gs = summary.grade_stats.get(grade)
        ws.cell(global_row + 1, i, format_percentage(gs.percentage) if gs else 0)

    # 现在进行所有合并操作
    # 标题合并
    ws.merge_cells("A1:H1")

    # 表头合并
    ws.merge_cells("A2:B4")
    ws.merge_cells("C2:G2")
    ws.merge_cells("H2:H4")

    # 数据行合并
    for row in merge_rows:
        ws.merge_cells(f"A{row}:A{row + 1}")
        ws.cell(row, 1).alignment = CENTER_ALIGN
        ws.merge_cells(f"H{row}:H{row + 1}")
        ws.cell(row, 8).alignment = CENTER_ALIGN

    # 全域统计合并
    ws.merge_cells(f"A{global_row}:A{global_row + 1}")
    ws.cell(global_row, 1).alignment = CENTER_ALIGN
    ws.merge_cells(f"H{global_row}:H{global_row + 1}")
    ws.cell(global_row, 8).alignment = CENTER_ALIGN

    current_row = global_row + 1

    # 样式
    _apply_border_and_align(ws, 2, current_row, 1, 8)
    _set_column_widths(
        ws, {"A": 14, "B": 12, "C": 12, "D": 12, "E": 12, "F": 12, "G": 12, "H": 12}
    )


def write_land_use_summary(
    ws: Worksheet,
    summary: AttributeStatsSummary,
) -> None:
    """写入土地利用类型统计表"""
    grade_order = get_grade_order(summary.attr_key)
    grade_ranges = get_grade_ranges(summary.attr_key)

    ws.title = f"{summary.attr_name}土地利用类型统计"

    # 标题 - 先写数据
    ws["A1"] = f"{summary.attr_name}土地利用类型统计"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = CENTER_ALIGN

    # 表头 - 先写数据
    ws["A2"] = "土地利用类型"
    ws["A2"].font = HEADER_FONT
    ws["A2"].alignment = CENTER_ALIGN

    ws["D2"] = "分级"
    ws["D2"].font = HEADER_FONT
    ws["D2"].alignment = CENTER_ALIGN

    ws["I2"] = "指标平均等级"
    ws["I2"].font = HEADER_FONT
    ws["I2"].alignment = CENTER_ALIGN

    # 等级名称
    for i, grade in enumerate(grade_order[:5], start=4):
        col = get_column_letter(i)
        ws[f"{col}3"] = grade.replace("级", "")
        ws[f"{col}3"].font = HEADER_FONT
        ws[f"{col}3"].alignment = CENTER_ALIGN
        ws[f"{col}4"] = grade_ranges.get(grade, "")
        ws[f"{col}4"].font = HEADER_FONT
        ws[f"{col}4"].alignment = CENTER_ALIGN

    # 数据行 - 先写入所有数据，记录合并信息
    current_row = 5
    current_primary = None
    primary_start_row = None

    # 记录需要合并的区域
    primary_merges: list[tuple[str, int, int]] = []  # (primary, start, end)
    ab_merges: list[int] = []  # 一级地类所在行，合并A:B
    b_merges: list[int] = []  # 二级地类所在行，合并B列
    i_merges: list[int] = []  # 需要合并I列的行

    for land_stats in summary.land_use_stats:
        if land_stats.secondary == "":
            # 一级地类
            if current_primary is not None and primary_start_row is not None:
                primary_merges.append((current_primary, primary_start_row, current_row - 1))

            current_primary = land_stats.primary
            primary_start_row = current_row

            # 写入一级地类数据
            ws.cell(current_row, 1, land_stats.primary)
            ab_merges.append(current_row)

            ws.cell(current_row, 3, "面积")
            for i, grade in enumerate(grade_order[:5], start=4):
                gs = land_stats.grade_stats.get(grade)
                ws.cell(current_row, i, format_small_value(gs.area) if gs else 0)
            ws.cell(
                current_row,
                9,
                land_stats.avg_grade if land_stats.avg_grade is not None else "",
            )
            i_merges.append(current_row)

            ws.cell(current_row + 1, 3, "占比")
            for i, grade in enumerate(grade_order[:5], start=4):
                gs = land_stats.grade_stats.get(grade)
                ws.cell(
                    current_row + 1, i, format_percentage(gs.percentage) if gs else 0
                )

            current_row += 2
        else:
            # 二级地类
            ws.cell(current_row, 2, land_stats.secondary)
            b_merges.append(current_row)

            ws.cell(current_row, 3, "面积")
            for i, grade in enumerate(grade_order[:5], start=4):
                gs = land_stats.grade_stats.get(grade)
                ws.cell(current_row, i, format_small_value(gs.area) if gs else 0)
            ws.cell(
                current_row,
                9,
                land_stats.avg_grade if land_stats.avg_grade is not None else "",
            )
            i_merges.append(current_row)

            ws.cell(current_row + 1, 3, "占比")
            for i, grade in enumerate(grade_order[:5], start=4):
                gs = land_stats.grade_stats.get(grade)
                ws.cell(
                    current_row + 1, i, format_percentage(gs.percentage) if gs else 0
                )

            current_row += 2

    # 记录最后一个一级地类
    if current_primary is not None and primary_start_row is not None:
        primary_merges.append((current_primary, primary_start_row, current_row - 1))

    # 全域统计 - 写入数据
    global_row = current_row
    ws.cell(global_row, 1, "全域")

    ws.cell(global_row, 3, "面积")
    for i, grade in enumerate(grade_order[:5], start=4):
        gs = summary.grade_stats.get(grade)
        ws.cell(global_row, i, format_small_value(gs.area) if gs else 0)
    ws.cell(
        global_row,
        9,
        summary.global_avg_grade if summary.global_avg_grade is not None else "",
    )

    ws.cell(global_row + 1, 3, "占比")
    for i, grade in enumerate(grade_order[:5], start=4):
        gs = summary.grade_stats.get(grade)
        ws.cell(global_row + 1, i, format_percentage(gs.percentage) if gs else 0)

    # 现在进行所有合并操作
    # 标题合并
    ws.merge_cells("A1:I1")

    # 表头合并
    ws.merge_cells("A2:C4")
    ws.merge_cells("D2:H2")
    ws.merge_cells("I2:I4")

    # 一级地类A列合并（跨多行）
    for primary, start, end in primary_merges:
        if end > start:
            ws.merge_cells(f"A{start}:A{end}")
            ws.cell(start, 1).alignment = CENTER_ALIGN

    # 一级地类A:B合并
    for row in ab_merges:
        ws.merge_cells(f"A{row}:B{row + 1}")
        ws.cell(row, 1).alignment = CENTER_ALIGN

    # 二级地类B列合并
    for row in b_merges:
        ws.merge_cells(f"B{row}:B{row + 1}")
        ws.cell(row, 2).alignment = CENTER_ALIGN

    # I列合并
    for row in i_merges:
        ws.merge_cells(f"I{row}:I{row + 1}")
        ws.cell(row, 9).alignment = CENTER_ALIGN

    # 全域统计合并
    ws.merge_cells(f"A{global_row}:B{global_row + 1}")
    ws.cell(global_row, 1).alignment = CENTER_ALIGN
    ws.merge_cells(f"I{global_row}:I{global_row + 1}")
    ws.cell(global_row, 9).alignment = CENTER_ALIGN

    current_row = global_row + 1

    # 样式
    _apply_border_and_align(ws, 1, current_row, 1, 9)
    _set_column_widths(
        ws,
        {
            "A": 12,
            "B": 12,
            "C": 12,
            "D": 12,
            "E": 12,
            "F": 12,
            "G": 12,
            "H": 12,
            "I": 12,
        },
    )


def write_sample_point_summary(
    ws: Worksheet,
    summary: AttributeStatsSummary,
) -> None:
    """写入样点统计表"""
    grade_order = get_grade_order(summary.attr_key)
    grade_ranges = get_grade_ranges(summary.attr_key)
    num_grades = len(grade_order)
    total_cols = 1 + num_grades
    last_col_letter = get_column_letter(total_cols)

    ws.title = f"{summary.attr_name}样点统计"

    # 标题
    ws.merge_cells(f"A1:{last_col_letter}1")
    ws["A1"] = f"{summary.attr_name}样点统计"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = CENTER_ALIGN

    # 表头
    ws["A2"] = "等级"
    ws["A2"].font = HEADER_FONT
    ws["A2"].alignment = CENTER_ALIGN

    for i, grade in enumerate(grade_order, start=2):
        col = get_column_letter(i)
        ws[f"{col}2"] = grade.replace("级", "")
        ws[f"{col}2"].font = HEADER_FONT
        ws[f"{col}2"].alignment = CENTER_ALIGN

    ws["A3"] = "级别范围"
    ws["A3"].font = HEADER_FONT
    ws["A3"].alignment = CENTER_ALIGN

    for i, grade in enumerate(grade_order, start=2):
        col = get_column_letter(i)
        ws[f"{col}3"] = grade_ranges.get(grade, "")
        ws[f"{col}3"].font = HEADER_FONT
        ws[f"{col}3"].alignment = CENTER_ALIGN

    # 样点数
    ws["A4"] = "样点数"
    for i, grade in enumerate(grade_order, start=2):
        col = get_column_letter(i)
        gs = summary.grade_stats.get(grade)
        ws[f"{col}4"] = gs.count if gs else 0

    # 样点数占比
    ws["A5"] = "样点数占比"
    for i, grade in enumerate(grade_order, start=2):
        col = get_column_letter(i)
        gs = summary.grade_stats.get(grade)
        pct = (
            (gs.count / summary.total_samples * 100)
            if gs and summary.total_samples > 0
            else 0
        )
        ws[f"{col}5"] = format_percentage(pct)

    # 全域统计
    ws["A6"] = "全域均值"
    ws.merge_cells("B6:C6")
    ws["B6"] = format_small_value(summary.global_mean)
    ws["B6"].alignment = CENTER_ALIGN

    ws["D6"] = "全域最小值"
    ws.merge_cells("E6:F6")
    ws["E6"] = format_small_value(summary.global_min)
    ws["E6"].alignment = CENTER_ALIGN

    ws["A7"] = "全域中位值"
    ws.merge_cells("B7:C7")
    ws["B7"] = format_small_value(summary.global_median)
    ws["B7"].alignment = CENTER_ALIGN

    ws["D7"] = "全域最大值"
    ws.merge_cells("E7:F7")
    ws["E7"] = format_small_value(summary.global_max)
    ws["E7"].alignment = CENTER_ALIGN

    # 样式
    _apply_border_and_align(ws, 2, 7, 1, total_cols)
    for i in range(1, total_cols + 1):
        ws.column_dimensions[get_column_letter(i)].width = 15


def write_town_sample_summary(
    ws: Worksheet,
    summary: AttributeStatsSummary,
) -> None:
    """写入分行政区样点统计表"""
    ws.title = f"{summary.attr_name}分行政区样点统计"

    # 标题
    ws.merge_cells("A1:F1")
    ws["A1"] = f"{summary.attr_name}分行政区样点统计"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = CENTER_ALIGN

    # 表头
    ws.merge_cells("A2:A3")
    ws["A2"] = "行政区"
    ws["A2"].font = HEADER_FONT
    ws["A2"].alignment = CENTER_ALIGN

    ws.merge_cells("B2:F2")
    ws["B2"] = "样点统计"
    ws["B2"].font = HEADER_FONT
    ws["B2"].alignment = CENTER_ALIGN

    headers = ["均值", "最小值", "最大值", "样点数", "样点占比"]
    for i, header in enumerate(headers, start=2):
        col = get_column_letter(i)
        ws[f"{col}3"] = header
        ws[f"{col}3"].font = HEADER_FONT
        ws[f"{col}3"].alignment = CENTER_ALIGN

    # 数据行
    current_row = 4
    for town_stats in summary.town_stats:
        ws.cell(current_row, 1, town_stats.town)
        ws.cell(current_row, 2, format_small_value(town_stats.sample_mean))
        ws.cell(current_row, 3, format_small_value(town_stats.sample_min))
        ws.cell(current_row, 4, format_small_value(town_stats.sample_max))
        ws.cell(current_row, 5, town_stats.sample_count)
        ws.cell(current_row, 6, format_percentage(town_stats.sample_percentage))
        current_row += 1

    # 全域统计
    ws.cell(current_row, 1, "全域")
    ws.cell(current_row, 2, format_small_value(summary.global_mean))
    ws.cell(current_row, 3, format_small_value(summary.global_min))
    ws.cell(current_row, 4, format_small_value(summary.global_max))
    ws.cell(current_row, 5, summary.total_samples)
    ws.cell(current_row, 6, "100")

    # 样式
    _apply_border_and_align(ws, 2, current_row, 1, 6)
    _set_column_widths(ws, {"A": 14, "B": 12, "C": 12, "D": 12, "E": 12, "F": 12})


def write_land_use_sample_summary(
    ws: Worksheet,
    summary: AttributeStatsSummary,
) -> None:
    """写入土地利用类型样点统计表"""
    ws.title = f"{summary.attr_name}土地利用类型样点统计"

    # 标题
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{summary.attr_name}土地利用类型样点统计"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = CENTER_ALIGN

    # 表头
    ws.merge_cells("A2:B3")
    ws["A2"] = "土地利用类型"
    ws["A2"].font = HEADER_FONT
    ws["A2"].alignment = CENTER_ALIGN

    ws.merge_cells("C2:G2")
    ws["C2"] = "样点统计"
    ws["C2"].font = HEADER_FONT
    ws["C2"].alignment = CENTER_ALIGN

    headers = ["均值", "最小值", "最大值", "样点数", "样点占比"]
    for i, header in enumerate(headers, start=3):
        col = get_column_letter(i)
        ws[f"{col}3"] = header
        ws[f"{col}3"].font = HEADER_FONT
        ws[f"{col}3"].alignment = CENTER_ALIGN

    # 数据行
    current_row = 4
    current_primary = None
    primary_start_row = None

    for land_stats in summary.land_use_stats:
        if land_stats.secondary == "":
            # 一级地类
            if current_primary is not None and primary_start_row is not None:
                if current_row - 1 > primary_start_row:
                    ws.merge_cells(f"A{primary_start_row}:A{current_row - 1}")
                    ws.cell(primary_start_row, 1).alignment = CENTER_ALIGN

            current_primary = land_stats.primary
            primary_start_row = current_row

            ws.merge_cells(f"A{current_row}:B{current_row}")
            ws.cell(current_row, 1, land_stats.primary)
            ws.cell(current_row, 1).alignment = CENTER_ALIGN
        else:
            ws.cell(current_row, 2, land_stats.secondary)

        ws.cell(current_row, 3, format_small_value(land_stats.sample_mean))
        ws.cell(current_row, 4, format_small_value(land_stats.sample_min))
        ws.cell(current_row, 5, format_small_value(land_stats.sample_max))
        ws.cell(current_row, 6, land_stats.sample_count)
        ws.cell(current_row, 7, format_percentage(land_stats.sample_percentage))
        current_row += 1

    # 处理最后一个一级地类
    if current_primary is not None and primary_start_row is not None:
        if current_row - 1 > primary_start_row:
            ws.merge_cells(f"A{primary_start_row}:A{current_row - 1}")
            ws.cell(primary_start_row, 1).alignment = CENTER_ALIGN

    # 全域统计
    ws.merge_cells(f"A{current_row}:B{current_row}")
    ws.cell(current_row, 1, "全域")
    ws.cell(current_row, 1).alignment = CENTER_ALIGN
    ws.cell(current_row, 3, format_small_value(summary.global_mean))
    ws.cell(current_row, 4, format_small_value(summary.global_min))
    ws.cell(current_row, 5, format_small_value(summary.global_max))
    ws.cell(current_row, 6, summary.total_samples)
    ws.cell(current_row, 7, "100")

    # 样式
    _apply_border_and_align(ws, 2, current_row, 1, 7)
    _set_column_widths(
        ws, {"A": 12, "B": 12, "C": 12, "D": 12, "E": 12, "F": 12, "G": 12}
    )


def write_soil_type_summary(
    ws: Worksheet,
    summary: AttributeStatsSummary,
) -> None:
    """写入土壤类型统计表"""
    grade_order = get_grade_order(summary.attr_key)
    grade_ranges = get_grade_ranges(summary.attr_key)

    ws.title = f"{summary.attr_name}土壤类型统计"

    # 标题
    ws.merge_cells("A1:I1")
    ws["A1"] = f"{summary.attr_name}土壤类型统计"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = CENTER_ALIGN

    # 表头
    ws.merge_cells("A2:A3")
    ws["A2"] = "土类"
    ws["A2"].font = HEADER_FONT
    ws["A2"].alignment = CENTER_ALIGN

    ws.merge_cells("B2:B3")
    ws["B2"] = "亚类"
    ws["B2"].font = HEADER_FONT
    ws["B2"].alignment = CENTER_ALIGN

    ws.merge_cells("C2:C3")
    ws["C2"] = "土属"
    ws["C2"].font = HEADER_FONT
    ws["C2"].alignment = CENTER_ALIGN

    for i, grade in enumerate(grade_order[:5], start=4):
        col = get_column_letter(i)
        ws[f"{col}2"] = grade
        ws[f"{col}2"].font = HEADER_FONT
        ws[f"{col}2"].alignment = CENTER_ALIGN
        ws[f"{col}3"] = grade_ranges.get(grade, "")
        ws[f"{col}3"].font = HEADER_FONT
        ws[f"{col}3"].alignment = CENTER_ALIGN

    ws.merge_cells("I2:I3")
    ws["I2"] = "平均等级"
    ws["I2"].font = HEADER_FONT
    ws["I2"].alignment = CENTER_ALIGN

    # 数据行 - 先写入所有数据，最后再合并单元格
    current_row = 4
    major_ranges: dict[str, tuple[int, int]] = {}
    sub_ranges: dict[tuple[str, str], tuple[int, int]] = {}
    genus_rows: list[int] = []  # 记录需要合并土属和平均等级的行

    current_major = None
    current_sub = None
    major_start = None
    sub_start = None

    for soil_stats in summary.soil_type_stats:
        # 跟踪土类变化
        if current_major != soil_stats.major:
            if current_major is not None and major_start is not None:
                major_ranges[current_major] = (major_start, current_row - 1)
            current_major = soil_stats.major
            major_start = current_row

        # 跟踪亚类变化
        key = (soil_stats.major, soil_stats.sub)
        if current_sub != key:
            if current_sub is not None and sub_start is not None:
                sub_ranges[current_sub] = (sub_start, current_row - 1)
            current_sub = key
            sub_start = current_row

        # 面积行
        ws.cell(current_row, 1, soil_stats.major)
        ws.cell(current_row, 2, soil_stats.sub)
        ws.cell(current_row, 3, soil_stats.genus)

        for i, grade in enumerate(grade_order[:5], start=4):
            gs = soil_stats.grade_stats.get(grade)
            ws.cell(current_row, i, format_small_value(gs.area) if gs else 0)

        ws.cell(
            current_row,
            9,
            soil_stats.avg_grade if soil_stats.avg_grade is not None else "",
        )

        # 记录需要合并的行
        genus_rows.append(current_row)

        # 占比行 - 只写入等级数据列
        for i, grade in enumerate(grade_order[:5], start=4):
            gs = soil_stats.grade_stats.get(grade)
            ws.cell(current_row + 1, i, format_percentage(gs.percentage) if gs else 0)

        current_row += 2

    # 记录最后的范围
    if current_major is not None and major_start is not None:
        major_ranges[current_major] = (major_start, current_row - 1)
    if current_sub is not None and sub_start is not None:
        sub_ranges[current_sub] = (sub_start, current_row - 1)

    # 全域统计行 - 先写入数据
    global_row = current_row
    ws.cell(global_row, 1, "全域")

    for i, grade in enumerate(grade_order[:5], start=4):
        gs = summary.grade_stats.get(grade)
        ws.cell(global_row, i, format_small_value(gs.area) if gs else 0)

    ws.cell(
        global_row,
        9,
        summary.global_avg_grade if summary.global_avg_grade is not None else "",
    )

    for i, grade in enumerate(grade_order[:5], start=4):
        gs = summary.grade_stats.get(grade)
        ws.cell(global_row + 1, i, format_percentage(gs.percentage) if gs else 0)

    # 现在进行所有合并操作
    # 合并土类单元格
    for major, (start, end) in major_ranges.items():
        if end > start:
            ws.merge_cells(f"A{start}:A{end}")
            ws.cell(start, 1).alignment = CENTER_ALIGN

    # 合并亚类单元格
    for (major, sub), (start, end) in sub_ranges.items():
        if end > start:
            ws.merge_cells(f"B{start}:B{end}")
            ws.cell(start, 2).alignment = CENTER_ALIGN

    # 合并土属和平均等级单元格
    for row in genus_rows:
        ws.merge_cells(f"C{row}:C{row + 1}")
        ws.cell(row, 3).alignment = CENTER_ALIGN
        ws.merge_cells(f"I{row}:I{row + 1}")
        ws.cell(row, 9).alignment = CENTER_ALIGN

    # 合并全域统计行
    ws.merge_cells(f"A{global_row}:C{global_row + 1}")
    ws.cell(global_row, 1).alignment = CENTER_ALIGN
    ws.merge_cells(f"I{global_row}:I{global_row + 1}")
    ws.cell(global_row, 9).alignment = CENTER_ALIGN

    current_row = global_row + 1

    # 样式
    _apply_border_and_align(ws, 1, current_row + 1, 1, 9)
    _set_column_widths(
        ws,
        {
            "A": 12,
            "B": 12,
            "C": 12,
            "D": 12,
            "E": 12,
            "F": 12,
            "G": 12,
            "H": 12,
            "I": 12,
        },
    )


def write_soil_type_sample_summary(
    ws: Worksheet,
    summary: AttributeStatsSummary,
) -> None:
    """写入土壤类型样点统计表"""
    ws.title = f"{summary.attr_name}土壤类型样点统计"

    # 标题
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{summary.attr_name}土壤类型样点统计"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = CENTER_ALIGN

    # 表头
    ws.merge_cells("A2:B3")
    ws["A2"] = "土壤类型"
    ws["A2"].font = HEADER_FONT
    ws["A2"].alignment = CENTER_ALIGN

    ws.merge_cells("C2:G2")
    ws["C2"] = "样点统计"
    ws["C2"].font = HEADER_FONT
    ws["C2"].alignment = CENTER_ALIGN

    headers = ["均值", "最小值", "最大值", "样点数", "样点占比"]
    for i, header in enumerate(headers, start=3):
        col = get_column_letter(i)
        ws[f"{col}3"] = header
        ws[f"{col}3"].font = HEADER_FONT
        ws[f"{col}3"].alignment = CENTER_ALIGN

    # 数据行
    current_row = 4
    sub_ranges: dict[str, tuple[int, int]] = {}
    current_sub = None
    sub_start = None

    for soil_stats in summary.soil_type_stats:
        # 跟踪亚类变化
        if current_sub != soil_stats.sub:
            if current_sub is not None and sub_start is not None:
                sub_ranges[current_sub] = (sub_start, current_row - 1)
            current_sub = soil_stats.sub
            sub_start = current_row

        ws.cell(current_row, 1, soil_stats.sub)
        ws.cell(current_row, 2, soil_stats.genus)
        ws.cell(current_row, 3, format_small_value(soil_stats.sample_mean))
        ws.cell(current_row, 4, format_small_value(soil_stats.sample_min))
        ws.cell(current_row, 5, format_small_value(soil_stats.sample_max))
        ws.cell(current_row, 6, soil_stats.sample_count)
        ws.cell(current_row, 7, format_percentage(soil_stats.sample_percentage))

        current_row += 1

    # 记录最后的范围
    if current_sub is not None and sub_start is not None:
        sub_ranges[current_sub] = (sub_start, current_row - 1)

    # 合并亚类单元格
    for sub, (start, end) in sub_ranges.items():
        if end > start:
            ws.merge_cells(f"A{start}:A{end}")
            ws.cell(start, 1).alignment = CENTER_ALIGN

    # 全域统计
    ws.merge_cells(f"A{current_row}:B{current_row}")
    ws.cell(current_row, 1, "全域")
    ws.cell(current_row, 1).alignment = CENTER_ALIGN
    ws.cell(current_row, 3, format_small_value(summary.global_mean))
    ws.cell(current_row, 4, format_small_value(summary.global_min))
    ws.cell(current_row, 5, format_small_value(summary.global_max))
    ws.cell(current_row, 6, summary.total_samples)
    ws.cell(current_row, 7, "100")

    # 样式
    _apply_border_and_align(ws, 2, current_row, 1, 7)
    _set_column_widths(
        ws, {"A": 12, "B": 12, "C": 12, "D": 12, "E": 12, "F": 12, "G": 12}
    )


def write_overall_statistics_summary(
    ws: Worksheet,
    summaries: list[AttributeStatsSummary],
) -> None:
    """写入全域属性统计汇总表"""
    ws.title = "全域属性统计汇总"

    # 标题
    ws.merge_cells("A1:I1")
    ws["A1"] = "全域土壤属性统计汇总"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = CENTER_ALIGN

    # 表头
    headers = ["土壤属性", "样本数", "最小值", "最大值", "极差", "中位数", "平均值", "标准差", "变异系数"]
    for i, header in enumerate(headers, start=1):
        ws.cell(2, i, header)
        ws.cell(2, i).font = HEADER_FONT
        ws.cell(2, i).alignment = CENTER_ALIGN

    # 数据行
    current_row = 3
    for summary in summaries:
        if summary.total_samples == 0:
            continue

        display_name = (
            f"{summary.attr_name}（{summary.unit}）"
            if summary.unit
            else summary.attr_name
        )

        ws.cell(current_row, 1, display_name)
        ws.cell(current_row, 2, summary.total_samples)
        ws.cell(current_row, 3, format_small_value(summary.global_min))
        ws.cell(current_row, 4, format_small_value(summary.global_max))
        ws.cell(
            current_row, 5, format_small_value(summary.global_max - summary.global_min)
        )
        ws.cell(current_row, 6, format_small_value(summary.global_median))
        ws.cell(current_row, 7, format_small_value(summary.global_mean))
        ws.cell(current_row, 8, format_small_value(summary.global_std))
        ws.cell(current_row, 9, format_small_value(summary.global_cv))

        current_row += 1

    # 样式
    _apply_border_and_align(ws, 2, current_row - 1, 1, 9)
    ws.column_dimensions["A"].width = 20
    for col in ["B", "C", "D", "E", "F", "G", "H", "I"]:
        ws.column_dimensions[col].width = 12


def write_percentile_statistics_summary(
    ws: Worksheet,
    summaries: list[AttributeStatsSummary],
) -> None:
    """写入全域属性百分位数统计表"""
    ws.title = "全域属性百分位数统计"

    # 标题
    ws.merge_cells("A1:I1")
    ws["A1"] = "全域土壤属性百分位数统计"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = CENTER_ALIGN

    # 表头
    headers = ["土壤属性", "2%", "5%", "10%", "20%", "80%", "90%", "95%", "98%"]
    for i, header in enumerate(headers, start=1):
        ws.cell(2, i, header)
        ws.cell(2, i).font = HEADER_FONT
        ws.cell(2, i).alignment = CENTER_ALIGN

    # 数据行
    current_row = 3
    for summary in summaries:
        if summary.total_samples == 0 or not summary.percentiles:
            continue

        display_name = (
            f"{summary.attr_name}（{summary.unit}）"
            if summary.unit
            else summary.attr_name
        )

        ws.cell(current_row, 1, display_name)

        percentile_keys = ["2%", "5%", "10%", "20%", "80%", "90%", "95%", "98%"]
        for i, key in enumerate(percentile_keys, start=2):
            value = summary.percentiles.get(key, 0)
            ws.cell(current_row, i, format_small_value(value))

        current_row += 1

    # 样式
    _apply_border_and_align(ws, 2, current_row - 1, 1, 9)
    ws.column_dimensions["A"].width = 20
    for col in ["B", "C", "D", "E", "F", "G", "H", "I"]:
        ws.column_dimensions[col].width = 12
