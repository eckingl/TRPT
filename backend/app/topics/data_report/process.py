"""数据报告处理主模块

协调数据读取、统计计算和Excel生成的完整流程。
"""

import io
import traceback
from collections.abc import Callable
from pathlib import Path

import pandas as pd
from openpyxl import Workbook

from app.topics.data_report.config import (
    SOIL_ATTR_CONFIG,
    detect_available_attributes,
)
from app.topics.data_report.land_use import apply_land_filter
from app.topics.data_report.stats import AttributeStatsSummary, compute_attribute_stats
from app.topics.data_report.writers import (
    write_land_use_sample_summary,
    write_land_use_summary,
    write_overall_statistics_summary,
    write_percentile_statistics_summary,
    write_sample_point_summary,
    write_soil_type_sample_summary,
    write_soil_type_summary,
    write_town_sample_summary,
    write_town_summary,
)


def read_data_file(file_path: str | Path) -> pd.DataFrame:
    """读取数据文件（CSV或Excel）

    支持自动编码检测。

    Args:
        file_path: 文件路径

    Returns:
        数据框

    Raises:
        ValueError: 文件读取失败
    """
    file_path = Path(file_path)

    if file_path.suffix.lower() == ".csv":
        # 尝试多种编码
        encodings = ["utf-8", "utf-8-sig", "gbk", "gb2312", "gb18030"]

        for encoding in encodings:
            try:
                return pd.read_csv(file_path, encoding=encoding)
            except UnicodeDecodeError:
                continue

        # 使用chardet检测编码
        try:
            import chardet

            with open(file_path, "rb") as f:
                result = chardet.detect(f.read())
            detected = result.get("encoding", "utf-8")
            return pd.read_csv(file_path, encoding=detected)
        except Exception as e:
            raise ValueError(f"无法读取CSV文件: {e}")
    else:
        # Excel文件
        return pd.read_excel(file_path)


def load_multiple_files(file_paths: list[str | Path]) -> pd.DataFrame:
    """读取并合并多个数据文件

    Args:
        file_paths: 文件路径列表

    Returns:
        合并后的数据框
    """
    if not file_paths:
        return pd.DataFrame()

    dfs = []
    for path in file_paths:
        try:
            df = read_data_file(path)
            dfs.append(df)
        except Exception as e:
            print(f"  [警告] 读取文件失败: {path}, 错误: {e}")

    if not dfs:
        return pd.DataFrame()

    return pd.concat(dfs, ignore_index=True)


def process_data_report(
    mapping_paths: list[str | Path] | None = None,
    sample_paths: list[str | Path] | None = None,
    progress_callback: Callable[[int, str], None] | None = None,
) -> tuple[bool, bytes | str]:
    """处理数据报告生成

    Args:
        mapping_paths: 制图数据文件路径列表
        sample_paths: 样点数据文件路径列表
        progress_callback: 进度回调函数 (进度百分比, 状态描述)

    Returns:
        (成功标志, Excel字节内容或错误信息)
    """
    # 打印接收到的参数
    print(f"[process_data_report] 开始处理:")
    print(f"  mapping_paths: {mapping_paths}")
    print(f"  sample_paths: {sample_paths}")

    try:
        if progress_callback:
            progress_callback(0, "正在读取数据文件...")

        # 读取制图数据
        df_mapping = None
        if mapping_paths:
            if progress_callback:
                progress_callback(5, "正在读取制图数据...")
            df_mapping = load_multiple_files(mapping_paths)
            if df_mapping.empty:
                df_mapping = None

        # 读取样点数据
        df_sample = None
        if sample_paths:
            if progress_callback:
                progress_callback(10, "正在读取样点数据...")
            df_sample = load_multiple_files(sample_paths)
            if df_sample.empty:
                df_sample = None

        if df_mapping is None and df_sample is None:
            return False, "未能读取任何数据文件"

        if progress_callback:
            progress_callback(15, "正在检测可用属性...")

        # 检测可用属性
        available_attrs: list[tuple[str, str]] = []
        if df_mapping is not None:
            available_attrs = detect_available_attributes(list(df_mapping.columns))
        if not available_attrs and df_sample is not None:
            available_attrs = detect_available_attributes(list(df_sample.columns))

        if not available_attrs:
            return False, "未检测到可用的土壤属性列"

        if progress_callback:
            progress_callback(20, f"检测到 {len(available_attrs)} 个可用属性")

        # 创建Excel工作簿
        wb = Workbook()
        if wb.active:
            wb.remove(wb.active)

        # 计算所有属性的统计
        all_summaries: list[AttributeStatsSummary] = []
        total_attrs = len(available_attrs)

        for idx, (orig_col, attr_key) in enumerate(available_attrs):
            config = SOIL_ATTR_CONFIG.get(attr_key, {})
            attr_name = config.get("name", attr_key)

            if progress_callback:
                progress = 20 + int((idx / total_attrs) * 50)
                progress_callback(progress, f"正在处理: {attr_name}")

            try:
                # 准备数据
                df_m = None
                df_s = None

                if df_mapping is not None and orig_col in df_mapping.columns:
                    df_m = apply_land_filter(df_mapping.copy(), attr_key)

                if df_sample is not None and orig_col in df_sample.columns:
                    df_s = df_sample.copy()

                # 重命名列为标准键
                if df_m is not None and orig_col != attr_key:
                    df_m = df_m.rename(columns={orig_col: attr_key})
                if df_s is not None and orig_col != attr_key:
                    df_s = df_s.rename(columns={orig_col: attr_key})

                # 计算统计
                summary = compute_attribute_stats(df_m, df_s, attr_key)
                all_summaries.append(summary)

                # 生成表格
                if summary.total_area > 0 or summary.total_samples > 0:
                    # 表1: 分乡镇统计（制图数据）
                    if summary.town_stats and summary.total_area > 0:
                        ws1 = wb.create_sheet()
                        write_town_summary(ws1, summary)

                    # 表2: 土地利用类型统计（制图数据）
                    if summary.land_use_stats and summary.total_area > 0:
                        ws2 = wb.create_sheet()
                        write_land_use_summary(ws2, summary)

                    # 表3: 土壤类型统计（制图数据）
                    if summary.soil_type_stats and summary.total_area > 0:
                        ws3 = wb.create_sheet()
                        write_soil_type_summary(ws3, summary)

                    # 表4: 样点统计
                    if summary.total_samples > 0:
                        ws4 = wb.create_sheet()
                        write_sample_point_summary(ws4, summary)

                    # 表5: 分行政区样点统计（只需要样点数据）
                    if summary.total_samples > 0:
                        ws5 = wb.create_sheet()
                        write_town_sample_summary(ws5, summary)

                    # 表6: 土地利用类型样点统计（只需要样点数据）
                    if summary.total_samples > 0:
                        ws6 = wb.create_sheet()
                        write_land_use_sample_summary(ws6, summary)

                    # 表7: 土壤类型样点统计（只需要样点数据）
                    if summary.total_samples > 0:
                        ws7 = wb.create_sheet()
                        write_soil_type_sample_summary(ws7, summary)

            except Exception as e:
                print(f"  [警告] 处理属性 {attr_name} 失败: {e}")
                continue

        if progress_callback:
            progress_callback(75, "正在生成汇总表...")

        # 表8: 全域属性统计汇总（只生成一次）
        sample_summaries = [s for s in all_summaries if s.total_samples > 0]
        if sample_summaries:
            ws8 = wb.create_sheet()
            write_overall_statistics_summary(ws8, sample_summaries)

            # 表9: 全域属性百分位数统计
            ws9 = wb.create_sheet()
            write_percentile_statistics_summary(ws9, sample_summaries)

        if not wb.sheetnames:
            return False, "未能生成任何统计表"

        if progress_callback:
            progress_callback(90, "正在保存文件...")

        # 保存到字节流
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        if progress_callback:
            progress_callback(100, "完成！")

        return True, output.getvalue()

    except Exception as e:
        error_detail = traceback.format_exc()
        error_msg = f"处理数据时发生错误：{type(e).__name__}: {e}\n\n{error_detail}"
        return False, error_msg
