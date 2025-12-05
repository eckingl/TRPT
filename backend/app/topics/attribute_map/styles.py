"""Excel样式和格式化工具"""

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

# 预创建样式对象（避免重复创建）
THIN_SIDE = Side(border_style="thin")
BORDER = Border(top=THIN_SIDE, left=THIN_SIDE, right=THIN_SIDE, bottom=THIN_SIDE)
CENTER = Alignment(horizontal="center", vertical="center")
CENTER_HORIZONTAL = Alignment(horizontal="center")
BOLD_FONT = Font(bold=True)
TITLE_FONT = Font(bold=True, size=14)
SUBTITLE_FONT = Font(bold=True, size=12)


def format_value(value: float, decimals: int = 3) -> float | str:
    """格式化数值"""
    if pd.isna(value):
        return value
    if value == 0:
        return 0
    abs_val = abs(value)
    if abs_val >= 0.001:
        return round(value, decimals)
    return f"{value:.3g}"


def format_percentage(value: float) -> float | str:
    """格式化百分比"""
    if pd.isna(value):
        return value
    if value > 100:
        return 100
    return format_value(value)


def format_range(min_val: float, max_val: float) -> str:
    """格式化范围值"""
    if pd.isna(min_val) or pd.isna(max_val):
        return ""
    return f"{min_val:.3f}～{max_val:.3f}"


def apply_excel_styles(ws, max_row: int, max_col: int) -> None:
    """应用Excel样式"""
    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.border = BORDER
            cell.alignment = CENTER

    for col in range(1, max_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 12


def apply_border_and_center(ws, max_row: int, max_col: int) -> None:
    """应用边框和居中样式（不设置列宽）"""
    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.border = BORDER
            cell.alignment = CENTER_HORIZONTAL


def set_column_widths(ws, col_letters: list[str], width: int = 12) -> None:
    """设置列宽"""
    for col in col_letters:
        ws.column_dimensions[col].width = width
