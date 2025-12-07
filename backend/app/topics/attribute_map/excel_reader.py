"""Excel结果文件读取器

从属性图数据处理生成的Excel文件中读取统计数据，用于报告生成。
"""

from dataclasses import dataclass, field
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from app.topics.attribute_map.config import SOIL_ATTR_CONFIG


@dataclass
class ExcelAttributeStats:
    """从Excel读取的属性统计数据"""

    attr_key: str
    attr_name: str
    unit: str

    # 总体统计
    sample_total: int = 0
    sample_mean: float = 0.0
    sample_median: float = 0.0
    sample_min: float = 0.0
    sample_max: float = 0.0
    area_total: float = 0.0
    area_mean: float = 0.0
    area_median: float = 0.0
    area_min: float = 0.0
    area_max: float = 0.0

    # 等级统计
    grade_sample_counts: dict[str, int] = field(default_factory=dict)
    grade_area_sums: dict[str, float] = field(default_factory=dict)

    # 土地利用类型统计（直接从Excel读取）
    land_use_stats: pd.DataFrame = field(default_factory=pd.DataFrame)

    # 乡镇统计
    town_stats: pd.DataFrame = field(default_factory=pd.DataFrame)

    # 土壤类型统计
    soil_type_stats: pd.DataFrame = field(default_factory=pd.DataFrame)


def read_excel_stats(
    excel_path: str | Path, attr_name: str
) -> ExcelAttributeStats | None:
    """从Excel文件读取指定属性的统计数据

    Args:
        excel_path: Excel文件路径
        attr_name: 属性名称（如"有机质"、"全氮"等）

    Returns:
        ExcelAttributeStats对象，如果读取失败返回None
    """
    excel_path = Path(excel_path)
    if not excel_path.exists():
        return None

    # 查找属性键名
    attr_key = None
    for key, config in SOIL_ATTR_CONFIG.items():
        if config["name"] == attr_name:
            attr_key = key
            break

    if not attr_key:
        return None

    config = SOIL_ATTR_CONFIG[attr_key]
    unit = config["unit"]

    try:
        wb = load_workbook(excel_path, data_only=True)
    except Exception:
        return None

    stats = ExcelAttributeStats(
        attr_key=attr_key,
        attr_name=attr_name,
        unit=unit,
    )

    # 读取总体情况sheet（名称格式：{attr_name}总体情况）
    overall_sheet_name = f"{attr_name}总体情况"
    if overall_sheet_name in wb.sheetnames:
        _read_overall_sheet(wb[overall_sheet_name], stats, config)

    # 读取土地利用类型sheet（名称格式：{attr_name}不同土地利用类型）
    land_use_sheet_name = f"{attr_name}不同土地利用类型"
    if land_use_sheet_name in wb.sheetnames:
        stats.land_use_stats = _read_land_use_sheet(wb[land_use_sheet_name])

    # 读取乡镇统计sheet（名称格式：{attr_name}乡镇统计）
    town_sheet_name = f"{attr_name}乡镇统计"
    if town_sheet_name in wb.sheetnames:
        stats.town_stats = _read_town_sheet(wb[town_sheet_name])

    # 读取土壤类型sheet（名称格式：{attr_name}分土壤类型）
    soil_type_sheet_name = f"{attr_name}分土壤类型"
    if soil_type_sheet_name in wb.sheetnames:
        stats.soil_type_stats = _read_soil_type_sheet(wb[soil_type_sheet_name])

    wb.close()
    return stats


def _read_overall_sheet(ws, stats: ExcelAttributeStats, config: dict) -> None:
    """读取总体情况sheet"""
    levels = config.get("levels", [])
    grade_order = [level[1] for level in levels]

    # 数据从第4行开始
    for i, grade in enumerate(grade_order):
        row = 4 + i
        try:
            # 样点数量
            count_val = ws.cell(row, 3).value
            stats.grade_sample_counts[grade] = int(count_val) if count_val else 0

            # 面积（第5列）
            area_val = ws.cell(row, 5).value
            if area_val:
                if isinstance(area_val, str):
                    area_val = area_val.replace(",", "")
                stats.grade_area_sums[grade] = float(area_val)
            else:
                stats.grade_area_sums[grade] = 0.0
        except (ValueError, TypeError):
            stats.grade_sample_counts[grade] = 0
            stats.grade_area_sums[grade] = 0.0

    # 读取合计行（等级数量 + 4）
    summary_row = 4 + len(grade_order)
    try:
        stats.sample_total = int(ws.cell(summary_row, 3).value or 0)
        area_total_val = ws.cell(summary_row, 5).value
        if area_total_val:
            if isinstance(area_total_val, str):
                area_total_val = area_total_val.replace(",", "")
            stats.area_total = float(area_total_val)
    except (ValueError, TypeError):
        pass

    # 读取均值行
    mean_row = summary_row + 1
    try:
        mean_sample_val = ws.cell(mean_row, 3).value
        if mean_sample_val:
            # 处理合并单元格的情况
            if isinstance(mean_sample_val, str):
                stats.sample_mean = float(mean_sample_val)
            else:
                stats.sample_mean = float(mean_sample_val)

        mean_area_val = ws.cell(mean_row, 5).value
        if mean_area_val:
            if isinstance(mean_area_val, str):
                stats.area_mean = float(mean_area_val)
            else:
                stats.area_mean = float(mean_area_val)
    except (ValueError, TypeError):
        pass

    # 读取中位值行
    median_row = mean_row + 1
    try:
        median_sample_val = ws.cell(median_row, 3).value
        if median_sample_val:
            stats.sample_median = (
                float(median_sample_val)
                if isinstance(median_sample_val, (int, float))
                else float(median_sample_val)
            )

        median_area_val = ws.cell(median_row, 5).value
        if median_area_val:
            stats.area_median = (
                float(median_area_val)
                if isinstance(median_area_val, (int, float))
                else float(median_area_val)
            )
    except (ValueError, TypeError):
        pass

    # 读取范围行
    range_row = median_row + 1
    try:
        range_sample_val = ws.cell(range_row, 3).value
        if (
            range_sample_val
            and isinstance(range_sample_val, str)
            and "～" in range_sample_val
        ):
            parts = range_sample_val.split("～")
            stats.sample_min = float(parts[0])
            stats.sample_max = float(parts[1])

        range_area_val = ws.cell(range_row, 5).value
        if (
            range_area_val
            and isinstance(range_area_val, str)
            and "～" in range_area_val
        ):
            parts = range_area_val.split("～")
            stats.area_min = float(parts[0])
            stats.area_max = float(parts[1])
    except (ValueError, TypeError):
        pass


def _read_land_use_sheet(ws) -> pd.DataFrame:
    """读取土地利用类型sheet"""
    data = []

    # 从第4行开始读取数据（跳过标题行）
    current_primary = ""
    for row_idx in range(4, ws.max_row + 1):
        primary_val = ws.cell(row_idx, 1).value
        secondary_val = ws.cell(row_idx, 2).value

        # 跳过空行和全区行
        if secondary_val is None or secondary_val == "":
            continue
        if str(secondary_val).strip() == "全区":
            continue

        # 更新一级分类
        if primary_val and str(primary_val).strip():
            current_primary = str(primary_val).strip()

        secondary = str(secondary_val).strip()

        # 读取各列数据
        row_data = {
            "一级": current_primary,
            "二级": secondary,
            "样点均值": _parse_float(ws.cell(row_idx, 3).value),
            "样点中位数": _parse_float(ws.cell(row_idx, 4).value),
            "样点范围": ws.cell(row_idx, 5).value or "",
            "样点数量": _parse_int(ws.cell(row_idx, 6).value),
            "制图均值": _parse_float(ws.cell(row_idx, 7).value),
            "制图面积": _parse_float(ws.cell(row_idx, 8).value),
            "制图范围": ws.cell(row_idx, 9).value or "",
        }
        data.append(row_data)

    return pd.DataFrame(data) if data else pd.DataFrame()


def _read_town_sheet(ws) -> pd.DataFrame:
    """读取乡镇统计sheet"""
    data = []

    # 确定表头行（第2行）
    headers = []
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(2, col_idx).value
        if val:
            headers.append(str(val).strip())
        else:
            headers.append(f"col_{col_idx}")

    # 从第3行开始读取数据
    for row_idx in range(3, ws.max_row + 1):
        town_val = ws.cell(row_idx, 1).value
        if not town_val or str(town_val).strip() == "全区":
            continue

        row_data = {"乡镇": str(town_val).strip()}

        # 读取样点数
        row_data["样点数"] = _parse_int(ws.cell(row_idx, 2).value)

        # 读取面积
        row_data["面积"] = _parse_float(ws.cell(row_idx, 3).value)

        # 读取等级占比（从第4列开始，倒数第二列是均值）
        for col_idx in range(4, ws.max_column):
            header = (
                headers[col_idx - 1] if col_idx - 1 < len(headers) else f"col_{col_idx}"
            )
            val = ws.cell(row_idx, col_idx).value
            if val is not None:
                # 处理百分比值
                if isinstance(val, str) and "%" in val:
                    row_data[f"{header}_pct"] = _parse_float(val.replace("%", ""))
                else:
                    row_data[f"{header}_pct"] = _parse_float(val)

        # 读取均值（最后一列）
        mean_val = ws.cell(row_idx, ws.max_column).value
        row_data["均值"] = (
            _parse_float(mean_val) if mean_val and mean_val != "-" else None
        )

        data.append(row_data)

    return pd.DataFrame(data) if data else pd.DataFrame()


def _read_soil_type_sheet(ws) -> pd.DataFrame:
    """读取土壤类型sheet"""
    data = []

    # 从第4行开始读取数据
    current_yl = ""
    for row_idx in range(4, ws.max_row + 1):
        yl_val = ws.cell(row_idx, 1).value
        ts_val = ws.cell(row_idx, 2).value

        # 跳过空行和全区行
        if ts_val is None or str(ts_val).strip() == "":
            continue
        if str(ts_val).strip() == "全区":
            continue

        # 更新亚类
        if yl_val and str(yl_val).strip():
            current_yl = str(yl_val).strip()

        ts = str(ts_val).strip()

        row_data = {
            "YL": current_yl,
            "TS": ts,
            "sample_mean": _parse_float(ws.cell(row_idx, 3).value),
            "sample_median": _parse_float(ws.cell(row_idx, 4).value),
            "sample_range": ws.cell(row_idx, 5).value or "",
            "sample_count": _parse_int(ws.cell(row_idx, 6).value),
            "area_mean": _parse_float(ws.cell(row_idx, 7).value),
            "area_sum": _parse_float(ws.cell(row_idx, 8).value),
            "area_range": ws.cell(row_idx, 9).value or "",
        }

        # 解析范围值获取min/max
        if row_data["sample_range"] and "～" in str(row_data["sample_range"]):
            parts = str(row_data["sample_range"]).split("～")
            try:
                row_data["sample_min"] = float(parts[0])
                row_data["sample_max"] = float(parts[1])
            except ValueError:
                pass

        if row_data["area_range"] and "～" in str(row_data["area_range"]):
            parts = str(row_data["area_range"]).split("～")
            try:
                row_data["area_min"] = float(parts[0])
                row_data["area_max"] = float(parts[1])
            except ValueError:
                pass

        data.append(row_data)

    return pd.DataFrame(data) if data else pd.DataFrame()


def _parse_float(val) -> float:
    """安全解析浮点数"""
    if val is None or val == "" or val == "-":
        return 0.0
    try:
        if isinstance(val, str):
            val = val.replace(",", "").replace("%", "")
        return float(val)
    except (ValueError, TypeError):
        return 0.0


def _parse_int(val) -> int:
    """安全解析整数"""
    if val is None or val == "" or val == "-":
        return 0
    try:
        if isinstance(val, str):
            val = val.replace(",", "")
        return int(float(val))
    except (ValueError, TypeError):
        return 0


def read_excel_sheet_as_table(
    excel_path: str | Path, sheet_name: str
) -> list[list[str]]:
    """读取 Excel sheet 的原始表格数据

    Args:
        excel_path: Excel文件路径
        sheet_name: Sheet名称

    Returns:
        二维列表，每行是一个列表
    """
    excel_path = Path(excel_path)
    if not excel_path.exists():
        return []

    try:
        wb = load_workbook(excel_path, data_only=True)
        if sheet_name not in wb.sheetnames:
            wb.close()
            return []

        ws = wb[sheet_name]
        data = []

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
            row_data = []
            for cell in row:
                val = cell.value
                if val is None:
                    row_data.append("")
                elif isinstance(val, float):
                    # 格式化浮点数
                    if val == int(val):
                        row_data.append(str(int(val)))
                    else:
                        row_data.append(f"{val:.3f}")
                else:
                    row_data.append(str(val))
            data.append(row_data)

        wb.close()
        return data
    except Exception:
        return []


def get_available_attributes_from_excel(excel_path: str | Path) -> list[str]:
    """从Excel文件获取可用的属性列表

    Args:
        excel_path: Excel文件路径

    Returns:
        属性名称列表
    """
    excel_path = Path(excel_path)
    if not excel_path.exists():
        return []

    try:
        wb = load_workbook(excel_path, data_only=True)
        sheet_names = wb.sheetnames
        wb.close()

        # 从sheet名称中提取属性名（sheet名称格式：{attr_name}总体情况）
        attributes = []
        suffix = "总体情况"
        for name in sheet_names:
            if name.endswith(suffix):
                attr_name = name[: -len(suffix)]
                # 验证是否是有效属性
                for config in SOIL_ATTR_CONFIG.values():
                    if config["name"] == attr_name:
                        attributes.append(attr_name)
                        break

        return attributes
    except Exception:
        return []
