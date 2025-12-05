"""Excel写入函数"""

import pandas as pd
from openpyxl.styles import Alignment

from app.topics.attribute_map.config import get_level_value_ranges
from app.topics.attribute_map.stats import AttributeStats
from app.topics.attribute_map.styles import (
    BOLD_FONT,
    BORDER,
    CENTER,
    CENTER_HORIZONTAL,
    SUBTITLE_FONT,
    TITLE_FONT,
    apply_excel_styles,
    format_percentage,
    format_range,
    format_value,
    set_column_widths,
)

# 土地利用类型配置
LAND_USE_CONFIG = {
    "耕地": ["水田", "水浇地", "旱地"],
    "园地": ["果园", "茶园", "其他园地"],
    "林地": ["林地"],
    "草地": ["草地"],
    "其他": ["其他"],
}


def write_overall_summary(ws, stats: AttributeStats, grade_order: list[str]) -> None:
    """写入总体情况统计表"""
    ws.title = f"{stats.attr_name}总体情况"
    unit = stats.unit
    range_desc = get_level_value_ranges(stats.attr_key)

    total_samples = stats.sample_total
    total_area = stats.area_total

    # 计算占比
    sample_pct = {
        g: round(stats.grade_sample_counts[g] / total_samples * 100, 2)
        if total_samples > 0
        else 0
        for g in grade_order
    }
    area_pct = {
        g: round(stats.grade_area_sums[g] / total_area * 100, 2)
        if total_area > 0
        else 0
        for g in grade_order
    }

    # 写入表头
    ws.merge_cells("A1:F1")
    ws["A1"] = f"土壤{stats.attr_name}分级分布统计"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = CENTER

    ws.merge_cells("A2:B2")
    ws["A2"] = "土壤三普分级"
    ws["A2"].alignment = CENTER
    ws.merge_cells("C2:D2")
    ws["C2"] = "样点统计"
    ws["C2"].alignment = CENTER
    ws.merge_cells("E2:F2")
    ws["E2"] = "制图统计"
    ws["E2"].alignment = CENTER

    ws["A3"] = "分级"
    ws["B3"] = f"值域/({unit})" if unit else "值域"
    ws["C3"] = "数量/个"
    ws["D3"] = "占比/%"
    ws["E3"] = "面积/亩"
    ws["F3"] = "占比/%"

    # 写入数据
    for i, grade in enumerate(grade_order):
        row = 4 + i
        rng = range_desc[i] if i < len(range_desc) else ""
        ws.cell(row, 1, grade)
        ws.cell(row, 2, rng)
        ws.cell(row, 3, stats.grade_sample_counts.get(grade, 0))
        ws.cell(row, 4, format_percentage(sample_pct.get(grade, 0)))
        ws.cell(row, 5, format_value(stats.grade_area_sums.get(grade, 0)))
        ws.cell(row, 6, format_percentage(area_pct.get(grade, 0)))

    # 合计行
    summary_row = 4 + len(grade_order)
    ws.merge_cells(f"A{summary_row}:B{summary_row}")
    ws[f"A{summary_row}"] = "全区"
    ws[f"A{summary_row}"].alignment = CENTER
    ws[f"C{summary_row}"] = total_samples
    ws[f"D{summary_row}"] = format_percentage(sum(sample_pct.values()))
    ws[f"E{summary_row}"] = format_value(total_area)
    ws[f"F{summary_row}"] = format_percentage(sum(area_pct.values()))

    # 均值行
    stat_row = summary_row + 1
    ws.merge_cells(f"A{stat_row}:B{stat_row}")
    ws[f"A{stat_row}"] = f"全区均值/({unit})" if unit else "全区均值"
    ws[f"A{stat_row}"].alignment = CENTER
    ws.merge_cells(f"C{stat_row}:D{stat_row}")
    ws[f"C{stat_row}"] = f"{stats.sample_mean:.3f}"
    ws[f"C{stat_row}"].alignment = CENTER
    ws.merge_cells(f"E{stat_row}:F{stat_row}")
    ws[f"E{stat_row}"] = f"{stats.area_mean:.3f}"
    ws[f"E{stat_row}"].alignment = CENTER

    # 中位值行
    stat_row += 1
    ws.merge_cells(f"A{stat_row}:B{stat_row}")
    ws[f"A{stat_row}"] = f"全区中位值/({unit})" if unit else "全区中位值"
    ws[f"A{stat_row}"].alignment = CENTER
    ws.merge_cells(f"C{stat_row}:D{stat_row}")
    ws[f"C{stat_row}"] = f"{stats.sample_median:.3f}"
    ws[f"C{stat_row}"].alignment = CENTER
    ws.merge_cells(f"E{stat_row}:F{stat_row}")
    ws[f"E{stat_row}"] = f"{stats.area_median:.3f}"
    ws[f"E{stat_row}"].alignment = CENTER

    # 范围行
    stat_row += 1
    ws.merge_cells(f"A{stat_row}:B{stat_row}")
    ws[f"A{stat_row}"] = f"全区范围/({unit})" if unit else "全区范围"
    ws[f"A{stat_row}"].alignment = CENTER
    ws.merge_cells(f"C{stat_row}:D{stat_row}")
    ws[f"C{stat_row}"] = format_range(stats.sample_min, stats.sample_max)
    ws[f"C{stat_row}"].alignment = CENTER
    ws.merge_cells(f"E{stat_row}:F{stat_row}")
    ws[f"E{stat_row}"] = format_range(stats.area_min, stats.area_max)
    ws[f"E{stat_row}"].alignment = CENTER

    apply_excel_styles(ws, stat_row, 6)


def write_land_use_summary(ws, stats: AttributeStats) -> None:
    """写入土地利用类型统计表（包含一级和二级分类）"""
    ws.title = f"{stats.attr_name}不同土地利用类型"
    attr_key = stats.attr_key
    unit = stats.unit
    df_sample = stats.df_sample_clean
    df_area = stats.df_area_clean

    # 标题
    ws.merge_cells("A1:I1")
    ws["A1"] = f"不同土地利用类型{stats.attr_name}统计"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = CENTER

    # 表头 - 第2行
    ws.merge_cells("A2:A3")
    ws["A2"] = "一级"
    ws["A2"].font = BOLD_FONT
    ws["A2"].alignment = CENTER

    ws.merge_cells("B2:B3")
    ws["B2"] = "二级"
    ws["B2"].font = BOLD_FONT
    ws["B2"].alignment = CENTER

    ws.merge_cells("C2:F2")
    ws["C2"] = "样点统计"
    ws["C2"].alignment = CENTER

    ws.merge_cells("G2:I2")
    ws["G2"] = "制图统计"
    ws["G2"].alignment = CENTER

    # 表头 - 第3行
    if attr_key == "ph":
        ws["C3"] = "均值"
        ws["D3"] = "中位数"
        ws["E3"] = "范围"
        ws["G3"] = "均值"
        ws["I3"] = "范围"
    else:
        ws["C3"] = f"均值/({unit})"
        ws["D3"] = f"中位数/({unit})"
        ws["E3"] = f"范围/({unit})"
        ws["G3"] = f"均值/({unit})"
        ws["I3"] = f"范围/({unit})"
    ws["F3"] = "数量/个"
    ws["H3"] = "面积/亩"

    for col in ["C", "D", "E", "F", "G", "H", "I"]:
        ws[f"{col}2"].font = BOLD_FONT
        ws[f"{col}3"].font = BOLD_FONT

    current_row = 4
    start_row_map = {}

    for primary, secondaries in LAND_USE_CONFIG.items():
        # 筛选该一级类型的数据
        if "一级" in df_sample.columns:
            sample_primary = df_sample[df_sample["一级"] == primary]
        else:
            sample_primary = pd.DataFrame()

        if "一级" in df_area.columns:
            area_primary = df_area[df_area["一级"] == primary]
        else:
            area_primary = pd.DataFrame()

        total_count_p = 0
        total_area_p = 0.0
        vals_all_sample = []
        vals_all_area = []

        start_row_map[primary] = current_row

        for idx, sec in enumerate(secondaries):
            # 林地和草地只有一级，不按二级筛选
            if primary in ["林地", "草地", "其他"]:
                sample_sec = sample_primary
                area_sec = area_primary
            else:
                if "二级" in sample_primary.columns:
                    sample_sec = sample_primary[sample_primary["二级"] == sec]
                else:
                    sample_sec = pd.DataFrame()
                if "二级" in area_primary.columns:
                    area_sec = area_primary[area_primary["二级"] == sec]
                else:
                    area_sec = pd.DataFrame()

            count = len(sample_sec)
            vals_s = (
                sample_sec[attr_key].dropna()
                if not sample_sec.empty and attr_key in sample_sec.columns
                else pd.Series([], dtype="float64")
            )

            mean_sample = f"{vals_s.mean():.3f}" if len(vals_s) > 0 else ""
            median_sample = f"{vals_s.median():.3f}" if len(vals_s) > 0 else ""
            range_sample = (
                f"{vals_s.min():.3f}～{vals_s.max():.3f}" if len(vals_s) > 0 else ""
            )

            if not area_sec.empty and "面积" in area_sec.columns:
                area_val = area_sec["面积"].sum()
                vals_a = (
                    area_sec[attr_key].dropna()
                    if attr_key in area_sec.columns
                    else pd.Series([], dtype="float64")
                )
                mean_area = f"{vals_a.mean():.3f}" if len(vals_a) > 0 else ""
                range_area = (
                    f"{vals_a.min():.3f}～{vals_a.max():.3f}" if len(vals_a) > 0 else ""
                )
                vals_all_area.extend(vals_a.tolist())
            else:
                area_val = 0.0
                mean_area = ""
                range_area = ""

            # 写入行
            if idx == 0:
                ws.cell(row=current_row, column=1, value=primary)
            else:
                ws.cell(row=current_row, column=1, value="")

            ws.cell(row=current_row, column=2, value=sec)
            ws.cell(row=current_row, column=3, value=mean_sample)
            ws.cell(row=current_row, column=4, value=median_sample)
            ws.cell(row=current_row, column=5, value=range_sample)
            ws.cell(row=current_row, column=6, value=count)
            ws.cell(row=current_row, column=7, value=mean_area)
            ws.cell(row=current_row, column=8, value=format_value(area_val))
            ws.cell(row=current_row, column=9, value=range_area)

            total_count_p += count
            total_area_p += area_val
            vals_all_sample.extend(vals_s.tolist())
            current_row += 1

        # 耕地和园地需要合计行
        if primary in ["耕地", "园地"]:
            vals_series_sample = pd.Series(vals_all_sample)
            mean_all_sample = (
                f"{vals_series_sample.mean():.3f}"
                if len(vals_series_sample) > 0
                else ""
            )
            median_all_sample = (
                f"{vals_series_sample.median():.3f}"
                if len(vals_series_sample) > 0
                else ""
            )
            range_all_sample = (
                f"{vals_series_sample.min():.3f}～{vals_series_sample.max():.3f}"
                if len(vals_series_sample) > 0
                else ""
            )

            vals_series_area = pd.Series(vals_all_area)
            mean_all_area = (
                f"{vals_series_area.mean():.3f}" if len(vals_series_area) > 0 else ""
            )
            range_all_area = (
                f"{vals_series_area.min():.3f}～{vals_series_area.max():.3f}"
                if len(vals_series_area) > 0
                else ""
            )

            ws.cell(row=current_row, column=1, value="")
            ws.cell(row=current_row, column=2, value="合计")
            ws.cell(row=current_row, column=3, value=mean_all_sample)
            ws.cell(row=current_row, column=4, value=median_all_sample)
            ws.cell(row=current_row, column=5, value=range_all_sample)
            ws.cell(row=current_row, column=6, value=total_count_p)
            ws.cell(row=current_row, column=7, value=mean_all_area)
            ws.cell(row=current_row, column=8, value=format_value(total_area_p))
            ws.cell(row=current_row, column=9, value=range_all_area)
            current_row += 1

            # 合并一级列单元格
            end_row = current_row - 1
            if end_row > start_row_map[primary]:
                ws.merge_cells(f"A{start_row_map[primary]}:A{end_row}")
                ws.cell(row=start_row_map[primary], column=1).alignment = Alignment(
                    vertical="center", horizontal="center"
                )

    # 全区统计
    if attr_key in df_sample.columns:
        all_vals_sample = df_sample[attr_key].dropna()
    else:
        all_vals_sample = pd.Series([], dtype="float64")

    global_total_count = len(all_vals_sample)
    global_mean_sample = (
        f"{all_vals_sample.mean():.3f}" if len(all_vals_sample) > 0 else ""
    )
    global_median_sample = (
        f"{all_vals_sample.median():.3f}" if len(all_vals_sample) > 0 else ""
    )
    global_range_sample = (
        f"{all_vals_sample.min():.3f}～{all_vals_sample.max():.3f}"
        if len(all_vals_sample) > 0
        else ""
    )

    if attr_key in df_area.columns:
        all_vals_area = df_area[attr_key].dropna()
    else:
        all_vals_area = pd.Series([], dtype="float64")

    global_total_area = df_area["面积"].sum() if "面积" in df_area.columns else 0
    global_mean_area = f"{all_vals_area.mean():.3f}" if len(all_vals_area) > 0 else ""
    global_range_area = (
        f"{all_vals_area.min():.3f}～{all_vals_area.max():.3f}"
        if len(all_vals_area) > 0
        else ""
    )

    # 合并全区行
    ws.merge_cells(f"A{current_row}:B{current_row}")
    ws.cell(row=current_row, column=1, value="全区")
    ws.cell(row=current_row, column=1).alignment = CENTER
    ws.cell(row=current_row, column=3, value=global_mean_sample)
    ws.cell(row=current_row, column=4, value=global_median_sample)
    ws.cell(row=current_row, column=5, value=global_range_sample)
    ws.cell(row=current_row, column=6, value=global_total_count)
    ws.cell(row=current_row, column=7, value=global_mean_area)
    ws.cell(row=current_row, column=8, value=format_value(global_total_area))
    ws.cell(row=current_row, column=9, value=global_range_area)

    # 应用样式
    for row in ws.iter_rows(min_row=1, max_row=current_row, min_col=1, max_col=9):
        for cell in row:
            cell.border = BORDER
            cell.alignment = CENTER_HORIZONTAL

    set_column_widths(ws, ["A", "B", "C", "D", "E", "F", "G", "H", "I"])


def write_town_summary(ws, stats: AttributeStats, grade_order: list[str]) -> None:
    """写入乡镇统计表"""
    ws.title = f"{stats.attr_name}乡镇统计"
    df = stats.town_stats

    # 写入表头
    ws.merge_cells(
        start_row=1, start_column=1, end_row=1, end_column=len(grade_order) + 4
    )
    ws["A1"] = f"土壤{stats.attr_name}分级分布统计（分乡镇）"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = CENTER

    ws["A2"] = "乡镇"
    ws["B2"] = "样点/个"
    ws["C2"] = "面积/亩"
    for i, grade in enumerate(grade_order):
        ws.cell(row=2, column=4 + i, value=grade)
    ws.cell(row=2, column=4 + len(grade_order), value="均值")

    for col in range(1, 5 + len(grade_order)):
        ws.cell(row=2, column=col).font = BOLD_FONT
        ws.cell(row=2, column=col).alignment = CENTER

    current_row = 3
    if len(df) == 0:
        ws.cell(row=current_row, column=1, value="无数据")
        current_row += 1
    else:
        for _, row_data in df.iterrows():
            ws.cell(row=current_row, column=1, value=row_data["乡镇"])
            ws.cell(row=current_row, column=2, value=int(row_data["样点数"]))
            ws.cell(row=current_row, column=3, value=format_value(row_data["面积"]))

            for i, grade in enumerate(grade_order):
                pct = row_data.get(f"{grade}_pct", 0)
                ws.cell(row=current_row, column=4 + i, value=format_percentage(pct))

            mean_val = row_data["均值"]
            ws.cell(
                row=current_row,
                column=4 + len(grade_order),
                value=f"{mean_val:.3f}" if pd.notna(mean_val) else "-",
            )
            current_row += 1

    # 全区合计
    ws.cell(row=current_row, column=1, value="全区")
    ws.cell(row=current_row, column=2, value=stats.sample_total)
    ws.cell(row=current_row, column=3, value=format_value(stats.area_total))

    for i, grade in enumerate(grade_order):
        pct = (
            stats.grade_area_sums[grade] / stats.area_total * 100
            if stats.area_total > 0
            else 0
        )
        ws.cell(row=current_row, column=4 + i, value=format_percentage(pct))

    ws.cell(
        row=current_row,
        column=4 + len(grade_order),
        value=f"{stats.sample_mean:.3f}" if stats.sample_total > 0 else "-",
    )

    apply_excel_styles(ws, current_row, 4 + len(grade_order))


def write_soil_type_summary(ws, stats: AttributeStats) -> None:
    """写入分土壤类型统计表"""
    ws.title = f"{stats.attr_name}分土壤类型"
    unit = stats.unit
    df = stats.soil_type_stats

    # 写入标题
    ws.merge_cells("A1:I1")
    ws["A1"] = f"不同土壤类型{stats.attr_name}统计"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = CENTER

    ws.merge_cells("A2:B2")
    ws["A2"] = "土壤类型"
    ws["A2"].font = SUBTITLE_FONT
    ws["A2"].alignment = CENTER

    ws.merge_cells("C2:F2")
    ws["C2"] = "样点统计"
    ws["C2"].alignment = CENTER
    ws.merge_cells("G2:J2")
    ws["G2"] = "制图统计"
    ws["G2"].alignment = CENTER

    # 列标题
    if stats.attr_key == "ph":
        headers = [
            "亚类",
            "土属",
            "均值",
            "中位值",
            "范围",
            "数量/个",
            "均值",
            "面积/亩",
            "范围",
        ]
    else:
        headers = [
            "亚类",
            "土属",
            f"均值/({unit})",
            f"中位值/({unit})",
            f"范围/({unit})",
            "数量/个",
            f"均值/({unit})",
            "面积/亩",
            f"范围/({unit})",
        ]

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = BOLD_FONT
        cell.alignment = CENTER

    if len(df) == 0:
        apply_excel_styles(ws, 3, 9)
        return

    # 土壤类型排序顺序
    soil_type_order_map = get_soil_type_order_map()
    sub_order = list(soil_type_order_map.keys())

    def get_soil_order(soil_name: str) -> int:
        try:
            return sub_order.index(soil_name)
        except ValueError:
            return len(sub_order)

    # 按亚类和土属分组整理数据
    df["YL"] = df["YL"].astype(str).str.strip()
    df["TS"] = df["TS"].astype(str).str.strip()

    # 按亚类分组
    # 注意：不能使用 dict(grouped)，因为 DataFrameGroupBy 不支持直接转换
    grouped = df.groupby("YL", observed=True)
    sub_to_data = {name: group for name, group in grouped}  # noqa: C416
    sorted_subs = sorted(sub_to_data.keys(), key=get_soil_order)

    current_row = 4
    for sub in sorted_subs:
        sub_df = sub_to_data[sub]

        # 对土属排序
        if sub in soil_type_order_map:
            predefined = soil_type_order_map[sub]
            ts_list = sub_df["TS"].tolist()
            sorted_ts = [ts for ts in predefined if ts in ts_list]
            sorted_ts.extend(sorted([ts for ts in ts_list if ts not in predefined]))
        else:
            sorted_ts = sorted(sub_df["TS"].tolist())

        start_row_for_sub = current_row
        is_multi = len(sorted_ts) > 1

        for idx, ts in enumerate(sorted_ts):
            row_data = (
                sub_df[sub_df["TS"] == ts].iloc[0]
                if ts in sub_df["TS"].values
                else None
            )

            if row_data is None:
                continue

            # 样点统计
            count = (
                int(row_data.get("sample_count", 0))
                if pd.notna(row_data.get("sample_count"))
                else 0
            )
            mean_s = (
                f"{row_data['sample_mean']:.3f}"
                if pd.notna(row_data.get("sample_mean"))
                else ""
            )
            median_s = (
                f"{row_data['sample_median']:.3f}"
                if pd.notna(row_data.get("sample_median"))
                else ""
            )
            range_s = format_range(
                row_data.get("sample_min"), row_data.get("sample_max")
            )

            # 面积统计
            area_sum = (
                row_data.get("area_sum", 0) if pd.notna(row_data.get("area_sum")) else 0
            )
            mean_a = (
                f"{row_data['area_mean']:.3f}"
                if pd.notna(row_data.get("area_mean"))
                else ""
            )
            range_a = format_range(row_data.get("area_min"), row_data.get("area_max"))

            # 写入行
            if idx == 0:
                ws.cell(row=current_row, column=1, value=sub)
            else:
                ws.cell(row=current_row, column=1, value="")

            ws.cell(row=current_row, column=2, value=ts)
            ws.cell(row=current_row, column=3, value=mean_s)
            ws.cell(row=current_row, column=4, value=median_s)
            ws.cell(row=current_row, column=5, value=range_s)
            ws.cell(row=current_row, column=6, value=count)
            ws.cell(row=current_row, column=7, value=mean_a)
            ws.cell(row=current_row, column=8, value=format_value(area_sum))
            ws.cell(row=current_row, column=9, value=range_a)

            current_row += 1

        # 多土属时添加合计行
        if is_multi and len(sorted_ts) > 1:
            count_all = (
                int(sub_df["sample_count"].sum())
                if "sample_count" in sub_df.columns
                else 0
            )

            # 样点统计加权平均
            if count_all > 0:
                weights = sub_df["sample_count"].fillna(0)
                if weights.sum() > 0:
                    mean_s_all = (
                        sub_df["sample_mean"].fillna(0) * weights
                    ).sum() / weights.sum()
                    mean_s_all = f"{mean_s_all:.3f}"
                else:
                    mean_s_all = ""
                median_s_all = (
                    f"{sub_df['sample_median'].mean():.3f}"
                    if sub_df["sample_median"].notna().any()
                    else ""
                )
                s_min = sub_df["sample_min"].min()
                s_max = sub_df["sample_max"].max()
                range_s_all = format_range(s_min, s_max)
            else:
                mean_s_all = ""
                median_s_all = ""
                range_s_all = ""

            # 面积统计
            total_area_all = (
                sub_df["area_sum"].sum() if "area_sum" in sub_df.columns else 0
            )
            if total_area_all > 0:
                area_weights = sub_df["area_sum"].fillna(0)
                if area_weights.sum() > 0:
                    mean_a_all = (
                        sub_df["area_mean"].fillna(0) * area_weights
                    ).sum() / area_weights.sum()
                    mean_a_all = f"{mean_a_all:.3f}"
                else:
                    mean_a_all = ""
                a_min = sub_df["area_min"].min()
                a_max = sub_df["area_max"].max()
                range_a_all = format_range(a_min, a_max)
            else:
                mean_a_all = ""
                range_a_all = ""

            ws.cell(row=current_row, column=1, value="")
            ws.cell(row=current_row, column=2, value="合计")
            ws.cell(row=current_row, column=3, value=mean_s_all)
            ws.cell(row=current_row, column=4, value=median_s_all)
            ws.cell(row=current_row, column=5, value=range_s_all)
            ws.cell(row=current_row, column=6, value=count_all)
            ws.cell(row=current_row, column=7, value=mean_a_all)
            ws.cell(row=current_row, column=8, value=format_value(total_area_all))
            ws.cell(row=current_row, column=9, value=range_a_all)
            current_row += 1

            # 合并亚类单元格
            end_row_for_sub = current_row - 1
            if end_row_for_sub > start_row_for_sub:
                ws.merge_cells(f"A{start_row_for_sub}:A{end_row_for_sub}")
                ws.cell(row=start_row_for_sub, column=1).alignment = CENTER

    # 全区行
    ws.merge_cells(f"A{current_row}:B{current_row}")
    ws.cell(row=current_row, column=1, value="全区")
    ws.cell(row=current_row, column=1).alignment = CENTER
    ws.cell(
        row=current_row,
        column=3,
        value=f"{stats.sample_mean:.3f}" if stats.sample_total > 0 else "",
    )
    ws.cell(
        row=current_row,
        column=4,
        value=f"{stats.sample_median:.3f}" if stats.sample_total > 0 else "",
    )
    ws.cell(
        row=current_row,
        column=5,
        value=format_range(stats.sample_min, stats.sample_max)
        if stats.sample_total > 0
        else "",
    )
    ws.cell(row=current_row, column=6, value=stats.sample_total)
    ws.cell(
        row=current_row,
        column=7,
        value=f"{stats.area_mean:.3f}" if stats.area_total > 0 else "",
    )
    ws.cell(row=current_row, column=8, value=format_value(stats.area_total))
    ws.cell(
        row=current_row,
        column=9,
        value=format_range(stats.area_min, stats.area_max)
        if stats.area_total > 0
        else "",
    )

    # 应用样式
    for row in ws.iter_rows(min_row=3, max_row=current_row, min_col=1, max_col=9):
        for cell in row:
            cell.border = BORDER
            cell.alignment = CENTER_HORIZONTAL

    set_column_widths(ws, ["A", "B", "C", "D", "E", "F", "G", "H", "I"])


def get_soil_type_order_map() -> dict[str, list[str]]:
    """获取土壤类型排序映射"""
    return {
        "棕红壤": ["红泥质棕红壤"],
        "红壤性土": ["砂泥质红壤性土", "麻砂质红壤性土"],
        "典型黄棕壤": [
            "暗泥质黄棕壤",
            "麻砂质黄棕壤",
            "红砂质黄棕壤",
            "黄土质黄棕壤",
            "砂泥质黄棕壤",
        ],
        "黄棕壤性土": ["砂泥质黄棕壤性土"],
        "典型棕壤": ["麻砂质典型棕壤"],
        "白浆化棕壤": ["麻砂质白浆化棕壤", "泥砂质白浆化棕壤"],
        "潮棕壤": ["泥砂质潮棕壤"],
        "淋溶褐土": ["黄土质淋溶褐土", "灰泥质淋溶褐土", "暗泥质淋溶褐土"],
        "潮褐土": ["泥砂质潮褐土"],
        "红黏土": ["红黏土"],
        "黑色石灰土": ["黑色石灰土"],
        "棕色石灰土": ["棕色石灰土"],
        "暗火山灰土": ["暗火山灰土"],
        "酸性紫色土": ["壤质酸性紫色土", "黏质酸性紫色土"],
        "中性紫色土": ["砂质中性紫色土", "壤质中性紫色土", "黏质中性紫色土"],
        "石灰性紫色土": ["壤质石灰性紫色土"],
        "酸性粗骨土": ["麻砂质酸性粗骨土", "硅质酸性粗骨土"],
        "中性粗骨土": ["麻砂质中性粗骨土"],
        "钙质粗骨土": ["灰泥质钙质粗骨土"],
        "典型潮土": ["砂质潮土", "壤质潮土", "黏质潮土"],
        "灰潮土": ["灰潮土", "石灰性灰潮土"],
        "盐化潮土（含碱化潮土）": ["氯化物盐化潮土", "硫酸盐盐化潮土", "苏打盐化潮土"],
        "典型砂姜黑土": ["黑腐砂姜黑土（黑姜土）", "覆泥砂姜黑土（覆泥黑姜土）"],
        "盐化砂姜黑土": ["氯化物盐化砂姜黑土"],
        "腐泥沼泽土": ["腐泥沼泽土"],
        "草甸沼泽土": ["草甸沼泽土", "石灰性草甸沼泽土"],
        "典型滨海盐土": ["氯化物滨海盐土"],
        "滨海沼泽盐土": ["氯化物沼泽滨海盐土"],
        "滨海潮滩盐土": ["氯化物潮滩滨海盐土"],
        "淹育水稻土": ["浅马肝泥田"],
        "渗育水稻土": [
            "渗灰泥田",
            "渗潮泥砂田",
            "渗潮泥田",
            "渗湖泥田",
            "渗涂泥田",
            "渗淡涂泥田",
            "渗麻砂泥田",
            "渗潮白土田",
            "渗马肝泥田",
        ],
        "潴育水稻土": ["潮泥田", "湖泥田", "马肝泥田"],
        "潜育水稻土": ["青湖泥田", "青马肝泥田"],
        "脱潜水稻土": ["黄斑黏田", "黄斑泥田"],
        "漂洗水稻土": ["漂潮白土田", "漂马肝泥田"],
        "盐渍水稻土": [
            "氯化物潮泥田",
            "氯化物涂泥田",
            "氯化物湖泥田",
            "硫酸盐潮泥田",
            "硫酸盐涂泥田",
            "苏打潮泥田",
            "苏打涂泥田",
            "苏打湖泥田",
        ],
        "填充土": ["工矿填充土", "城镇填充土"],
        "扰动土": ["运移扰动土"],
    }
