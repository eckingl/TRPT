"""属性图上图处理模块

只处理制图统计数据，输出分级面积统计
"""

import io
from collections.abc import Callable
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

from app.core.data import (
    get_land_use_class,
    get_pinyin_sort_key,
    load_multiple_csv,
    normalize_dlmc_column,
)
from app.topics.attribute_map.config import (
    ATTR_LAND_USE_FILTERS,
    FARMLAND_GARDEN_ATTRS,
    ROMAN_NUMERALS,
    SOIL_ATTR_CONFIG,
    SOIL_TEXTURE_MAPPING,
    classify_series,
    detect_available_attributes,
    get_grade_order,
    get_level_value_ranges,
)


def format_value(value: float, decimals: int = 3) -> float | str:
    """格式化数值"""
    if pd.isna(value):
        return value
    if value == 0:
        return 0
    abs_val = abs(value)
    if abs_val >= 0.001:
        return round(value, decimals)
    return f"{value:.3g}"


def format_percentage(value: float) -> float | str:
    """格式化百分比"""
    if pd.isna(value):
        return value
    if value > 100:
        return 100
    return format_value(value)


def apply_excel_styles(ws, max_row: int, max_col: int) -> None:
    """应用Excel样式"""
    thin = Side(border_style="thin")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.border = border
            cell.alignment = center

    for col in range(1, max_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 12


def filter_by_land_use(df: pd.DataFrame, attr_key: str) -> pd.DataFrame:
    """根据属性过滤土地利用类型

    Args:
        df: 数据框
        attr_key: 属性键名

    Returns:
        过滤后的数据框
    """
    if "DLMC" not in df.columns:
        return df

    # 特定属性过滤
    if attr_key in ATTR_LAND_USE_FILTERS:
        allowed = ATTR_LAND_USE_FILTERS[attr_key]
        return df[df["DLMC"].isin(allowed)]

    # 耕园地属性过滤
    if attr_key in FARMLAND_GARDEN_ATTRS:
        farmland_garden = ["水田", "水浇地", "旱地", "果园", "茶园"]
        mask = df["DLMC"].isin(farmland_garden) | df["DLMC"].str.contains(
            "园地", case=False, na=False
        )
        return df[mask]

    return df


def generate_land_use_area_by_level(ws, df_area: pd.DataFrame, attr_key: str) -> None:
    """生成土地利用类型 × 属性分级面积表

    Args:
        ws: Excel工作表
        df_area: 制图数据
        attr_key: 属性键名
    """
    config = SOIL_ATTR_CONFIG[attr_key]
    attr_name = config["name"]
    ws.title = f"{attr_name}面积(土地利用)"

    grade_order = get_grade_order(attr_key)
    range_desc = get_level_value_ranges(attr_key)
    level_to_range = dict(zip(grade_order, range_desc))

    # 预处理数据
    df = df_area.copy()
    df[attr_key] = pd.to_numeric(df[attr_key], errors="coerce")

    if "面积" not in df.columns:
        raise ValueError("制图数据缺少'面积'列")

    df["面积"] = pd.to_numeric(df["面积"], errors="coerce")
    df = df[df["面积"].notna() & (df["面积"] > 0) & df[attr_key].notna()].copy()

    df["等级"] = classify_series(df[attr_key], attr_key)

    # 添加土地利用分类
    if "DLMC" in df.columns:
        df["土地利用"] = df["DLMC"].apply(lambda x: get_land_use_class(x)[0])
    else:
        df["土地利用"] = "其他"

    df = df.dropna(subset=["等级", "土地利用"])

    land_types = ["耕地", "园地", "林地", "草地", "其他"]

    if df.empty:
        grouped = pd.DataFrame(
            0.0, index=pd.Index(grade_order), columns=pd.Index(land_types)
        )
    else:
        grouped = (
            df.groupby(["等级", "土地利用"], observed=True)["面积"]
            .sum()
            .unstack(fill_value=0.0)
        )
        grouped = grouped.reindex(index=grade_order, fill_value=0.0)
        grouped = grouped.reindex(columns=land_types, fill_value=0.0)

    total_by_level = grouped.sum(axis=1)
    grand_total = float(total_by_level.sum())

    # 写入表头
    ws.merge_cells("A1:I1")
    ws["A1"] = f"{attr_name}分级面积统计（按土地利用类型）"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    headers = ["级别", "分级标准"] + land_types + ["总面积/亩", "占比/%"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 写入数据
    for i, level in enumerate(grade_order):
        row_num = 3 + i
        roman = ROMAN_NUMERALS[i] if i < len(ROMAN_NUMERALS) else str(i + 1)
        ws.cell(row=row_num, column=1, value=roman)
        ws.cell(row=row_num, column=2, value=level_to_range.get(level, ""))

        for j, lt in enumerate(land_types):
            val = (
                float(grouped.at[level, lt])
                if (level in grouped.index and lt in grouped.columns)
                else 0.0
            )
            ws.cell(row=row_num, column=3 + j, value=format_value(val))

        total_area_level = float(total_by_level.get(level, 0.0))
        ws.cell(row=row_num, column=8, value=format_value(total_area_level))
        pct = round(total_area_level / grand_total * 100, 2) if grand_total > 0 else 0.0
        ws.cell(row=row_num, column=9, value=format_percentage(pct))

    # 合计行
    summary_row = 3 + len(grade_order)
    ws.merge_cells(f"A{summary_row}:B{summary_row}")
    ws.cell(row=summary_row, column=1, value="合计")
    ws.cell(row=summary_row, column=1).alignment = Alignment(
        horizontal="center", vertical="center"
    )

    for j, lt in enumerate(land_types):
        col_sum = float(grouped[lt].sum()) if lt in grouped.columns else 0.0
        ws.cell(row=summary_row, column=3 + j, value=format_value(col_sum))

    ws.cell(row=summary_row, column=8, value=format_value(grand_total))
    ws.cell(row=summary_row, column=9, value=100.0)

    apply_excel_styles(ws, summary_row, 9)


def generate_town_area_by_level(ws, df_area: pd.DataFrame, attr_key: str) -> None:
    """生成乡镇 × 属性分级面积表

    Args:
        ws: Excel工作表
        df_area: 制图数据
        attr_key: 属性键名
    """
    config = SOIL_ATTR_CONFIG[attr_key]
    attr_name = config["name"]
    ws.title = f"{attr_name}面积(乡镇)"

    grade_order = get_grade_order(attr_key)
    range_desc = get_level_value_ranges(attr_key)
    level_to_range = dict(zip(grade_order, range_desc))

    # 应用土地利用过滤（乡镇统计只统计耕地）
    df = df_area.copy()
    if "DLMC" in df.columns:
        df = df[df["DLMC"].isin(["水田", "水浇地", "旱地"])]

        # 有效硅只统计水田
        if attr_key == "ASI":
            df = df[df["DLMC"].str.contains("水田", case=False, na=False)]

    df[attr_key] = pd.to_numeric(df[attr_key], errors="coerce")

    if "面积" not in df.columns:
        raise ValueError("制图数据缺少'面积'列")

    df["面积"] = pd.to_numeric(df["面积"], errors="coerce")
    df = df[df["面积"].notna() & (df["面积"] > 0) & df[attr_key].notna()].copy()

    df["等级"] = classify_series(df[attr_key], attr_key)
    df = df.dropna(subset=["等级", "行政区名称"]).copy()
    df["行政区名称"] = df["行政区名称"].astype(str).str.strip()

    if df.empty:
        towns = ["无乡镇数据"]
        grouped = pd.DataFrame(
            0.0, index=pd.Index(grade_order), columns=pd.Index(towns)
        )
    else:
        towns = sorted(df["行政区名称"].unique(), key=get_pinyin_sort_key)
        grouped = (
            df.groupby(["等级", "行政区名称"], observed=True)["面积"]
            .sum()
            .unstack(fill_value=0.0)
        )
        grouped = grouped.reindex(index=grade_order, fill_value=0.0)
        grouped = grouped.reindex(columns=pd.Index(towns), fill_value=0.0)

    total_by_level = grouped.sum(axis=1)
    grand_total = float(total_by_level.sum())

    # 写入表头
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(towns) + 4)
    ws["A1"] = f"{attr_name}分级面积统计（分乡镇）"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws.cell(row=2, column=1, value="级别")
    ws.cell(row=2, column=2, value="分级标准")
    for i, town in enumerate(towns):
        ws.cell(row=2, column=3 + i, value=town)
    ws.cell(row=2, column=3 + len(towns), value="总面积/亩")
    ws.cell(row=2, column=3 + len(towns) + 1, value="占比/%")

    for col in range(1, 3 + len(towns) + 2):
        ws.cell(row=2, column=col).font = Font(bold=True)
        ws.cell(row=2, column=col).alignment = Alignment(
            horizontal="center", vertical="center"
        )

    # 写入数据
    for i, level in enumerate(grade_order):
        row_num = 3 + i
        roman = ROMAN_NUMERALS[i] if i < len(ROMAN_NUMERALS) else str(i + 1)
        ws.cell(row=row_num, column=1, value=roman)
        ws.cell(row=row_num, column=2, value=level_to_range.get(level, ""))

        for j, town in enumerate(towns):
            val = (
                float(grouped.at[level, town])
                if (level in grouped.index and town in grouped.columns)
                else 0.0
            )
            ws.cell(row=row_num, column=3 + j, value=format_value(val))

        total_area_level = float(total_by_level.get(level, 0.0))
        ws.cell(
            row=row_num, column=3 + len(towns), value=format_value(total_area_level)
        )
        pct = round(total_area_level / grand_total * 100, 2) if grand_total > 0 else 0.0
        ws.cell(row=row_num, column=3 + len(towns) + 1, value=format_percentage(pct))

    # 合计行
    summary_row = 3 + len(grade_order)
    ws.merge_cells(f"A{summary_row}:B{summary_row}")
    ws.cell(row=summary_row, column=1, value="合计")
    ws.cell(row=summary_row, column=1).alignment = Alignment(
        horizontal="center", vertical="center"
    )

    for j, town in enumerate(towns):
        col_sum = float(grouped[town].sum()) if town in grouped.columns else 0.0
        ws.cell(row=summary_row, column=3 + j, value=format_value(col_sum))

    ws.cell(row=summary_row, column=3 + len(towns), value=format_value(grand_total))
    ws.cell(row=summary_row, column=3 + len(towns) + 1, value=100.0)

    max_col = 3 + len(towns) + 1
    apply_excel_styles(ws, summary_row, max_col)


def generate_soil_texture_by_land_use(ws, df_area: pd.DataFrame) -> None:
    """生成土壤质地按土地利用类型统计表"""
    ws.title = "土壤质地面积(土地利用)"

    df = df_area.copy()

    if "TRZD" not in df.columns:
        ws.merge_cells("A1:D1")
        ws["A1"] = "土壤质地面积统计（缺少TRZD列）"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        return

    if "面积" not in df.columns:
        raise ValueError("制图数据缺少'面积'列")

    df["面积"] = pd.to_numeric(df["面积"], errors="coerce")
    df = df[df["面积"].notna() & (df["面积"] != 0)].copy()

    # 映射到大类
    df["TRZD"] = df["TRZD"].astype(str).str.strip()
    texture_class_map = {k: v[0] for k, v in SOIL_TEXTURE_MAPPING.items()}
    df["质地大类"] = df["TRZD"].map(texture_class_map)

    # 土地利用分类
    if "DLMC" in df.columns:
        df["土地利用"] = df["DLMC"].apply(lambda x: get_land_use_class(x)[0])
    else:
        df["土地利用"] = "其他"

    df = df.dropna(subset=["质地大类", "土地利用"])

    texture_order = [
        "砂土类",
        "砂壤类",
        "轻壤类",
        "中壤类",
        "黏壤类",
        "轻黏类",
        "黏土类",
    ]
    land_types = ["耕地", "园地", "林地", "草地", "其他"]

    if df.empty:
        grouped = pd.DataFrame(
            0.0, index=pd.Index(texture_order), columns=pd.Index(land_types)
        )
    else:
        grouped = (
            df.groupby(["质地大类", "土地利用"], observed=True)["面积"]
            .sum()
            .unstack(fill_value=0.0)
        )
        grouped = grouped.reindex(index=texture_order, fill_value=0.0)
        grouped = grouped.reindex(columns=land_types, fill_value=0.0)

    total_by_texture = grouped.sum(axis=1)
    grand_total = float(total_by_texture.sum())

    # 写入表头
    ws.merge_cells("A1:I1")
    ws["A1"] = "土壤质地面积统计（按土地利用类型）"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    headers = ["级别", "质地类别"] + land_types + ["总面积/亩", "占比/%"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = Font(bold=True)

    # 写入数据
    for i, texture_class in enumerate(texture_order):
        row_num = 3 + i
        roman = ROMAN_NUMERALS[i] if i < len(ROMAN_NUMERALS) else str(i + 1)
        ws.cell(row=row_num, column=1, value=roman)
        ws.cell(row=row_num, column=2, value=texture_class)

        for j, lt in enumerate(land_types):
            val = (
                float(grouped.at[texture_class, lt])
                if (texture_class in grouped.index and lt in grouped.columns)
                else 0.0
            )
            ws.cell(row=row_num, column=3 + j, value=format_value(val))

        total_area = float(total_by_texture.get(texture_class, 0.0))
        ws.cell(row=row_num, column=8, value=format_value(total_area))
        pct = round(total_area / grand_total * 100, 2) if grand_total > 0 else 0.0
        ws.cell(row=row_num, column=9, value=format_percentage(pct))

    # 合计行
    summary_row = 3 + len(texture_order)
    ws.merge_cells(f"A{summary_row}:B{summary_row}")
    ws.cell(row=summary_row, column=1, value="合计")

    for j, lt in enumerate(land_types):
        col_sum = float(grouped[lt].sum()) if lt in grouped.columns else 0.0
        ws.cell(row=summary_row, column=3 + j, value=format_value(col_sum))

    ws.cell(row=summary_row, column=8, value=format_value(grand_total))
    ws.cell(row=summary_row, column=9, value=100.0)

    apply_excel_styles(ws, summary_row, 9)


def generate_soil_texture_by_town(ws, df_area: pd.DataFrame) -> None:
    """生成土壤质地按乡镇统计表"""
    ws.title = "土壤质地面积(乡镇)"

    # 只统计耕地
    df = df_area.copy()
    if "DLMC" in df.columns:
        df = df[df["DLMC"].isin(["水田", "水浇地", "旱地"])]

    if "TRZD" not in df.columns:
        ws.merge_cells("A1:D1")
        ws["A1"] = "土壤质地面积统计（缺少TRZD列）"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        return

    if "面积" not in df.columns:
        raise ValueError("制图数据缺少'面积'列")

    df["面积"] = pd.to_numeric(df["面积"], errors="coerce")
    df = df[df["面积"].notna() & (df["面积"] != 0)].copy()

    df["TRZD"] = df["TRZD"].astype(str).str.strip()
    texture_class_map = {k: v[0] for k, v in SOIL_TEXTURE_MAPPING.items()}
    df["质地大类"] = df["TRZD"].map(texture_class_map)

    df = df.dropna(subset=["质地大类", "行政区名称"]).copy()
    df["行政区名称"] = df["行政区名称"].astype(str).str.strip()

    texture_order = [
        "砂土类",
        "砂壤类",
        "轻壤类",
        "中壤类",
        "黏壤类",
        "轻黏类",
        "黏土类",
    ]

    if df.empty:
        towns = ["无乡镇数据"]
        grouped = pd.DataFrame(
            0.0, index=pd.Index(texture_order), columns=pd.Index(towns)
        )
    else:
        towns = sorted(df["行政区名称"].unique(), key=get_pinyin_sort_key)
        grouped = (
            df.groupby(["质地大类", "行政区名称"], observed=True)["面积"]
            .sum()
            .unstack(fill_value=0.0)
        )
        grouped = grouped.reindex(index=texture_order, fill_value=0.0)
        grouped = grouped.reindex(columns=pd.Index(towns), fill_value=0.0)

    total_by_texture = grouped.sum(axis=1)
    grand_total = float(total_by_texture.sum())

    # 写入表头
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(towns) + 4)
    ws["A1"] = "土壤质地面积统计（分乡镇）"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws.cell(row=2, column=1, value="级别")
    ws.cell(row=2, column=2, value="质地类别")
    for i, town in enumerate(towns):
        ws.cell(row=2, column=3 + i, value=town)
    ws.cell(row=2, column=3 + len(towns), value="总面积/亩")
    ws.cell(row=2, column=3 + len(towns) + 1, value="占比/%")

    for col in range(1, 3 + len(towns) + 2):
        ws.cell(row=2, column=col).font = Font(bold=True)

    # 写入数据
    for i, texture_class in enumerate(texture_order):
        row_num = 3 + i
        roman = ROMAN_NUMERALS[i] if i < len(ROMAN_NUMERALS) else str(i + 1)
        ws.cell(row=row_num, column=1, value=roman)
        ws.cell(row=row_num, column=2, value=texture_class)

        for j, town in enumerate(towns):
            val = (
                float(grouped.at[texture_class, town])
                if (texture_class in grouped.index and town in grouped.columns)
                else 0.0
            )
            ws.cell(row=row_num, column=3 + j, value=format_value(val))

        total_area = float(total_by_texture.get(texture_class, 0.0))
        ws.cell(row=row_num, column=3 + len(towns), value=format_value(total_area))
        pct = round(total_area / grand_total * 100, 2) if grand_total > 0 else 0.0
        ws.cell(row=row_num, column=3 + len(towns) + 1, value=format_percentage(pct))

    # 合计行
    summary_row = 3 + len(texture_order)
    ws.merge_cells(f"A{summary_row}:B{summary_row}")
    ws.cell(row=summary_row, column=1, value="合计")

    for j, town in enumerate(towns):
        col_sum = float(grouped[town].sum()) if town in grouped.columns else 0.0
        ws.cell(row=summary_row, column=3 + j, value=format_value(col_sum))

    ws.cell(row=summary_row, column=3 + len(towns), value=format_value(grand_total))
    ws.cell(row=summary_row, column=3 + len(towns) + 1, value=100.0)

    max_col = 3 + len(towns) + 1
    apply_excel_styles(ws, summary_row, max_col)


def process_mapping_data(
    area_paths: list[str | Path],
    progress_callback: Callable[[int, str], None] | None = None,
) -> tuple[bool, bytes | str]:
    """处理属性图上图数据

    Args:
        area_paths: 制图统计CSV文件路径列表
        progress_callback: 进度回调函数

    Returns:
        (成功标志, Excel文件字节或错误信息)
    """
    try:
        if progress_callback:
            progress_callback(0, "正在读取制图数据...")

        df_area = load_multiple_csv(area_paths)

        if progress_callback:
            progress_callback(15, "正在标准化列名...")

        try:
            df_area = normalize_dlmc_column(df_area)
        except ValueError:
            pass

        if progress_callback:
            progress_callback(20, "正在检测可用属性...")

        available_attrs = detect_available_attributes(list(df_area.columns))

        if not available_attrs:
            return False, "未在制图统计数据中找到任何支持的土壤属性列"

        if progress_callback:
            progress_callback(25, f"检测到 {len(available_attrs)} 个可用属性")

        wb = Workbook()
        if wb.active:
            wb.remove(wb.active)

        processed_any = False
        total_attrs = len(available_attrs)

        for idx, (orig_col, attr_key) in enumerate(available_attrs):
            try:
                if progress_callback:
                    progress = 25 + int((idx / max(total_attrs, 1)) * 60)
                    attr_name = SOIL_ATTR_CONFIG.get(attr_key, {}).get("name", attr_key)
                    progress_callback(progress, f"正在处理: {attr_name}")

                # 准备数据
                df_proc = df_area.copy()
                if orig_col in df_proc.columns and orig_col != attr_key:
                    df_proc = df_proc.rename(columns={orig_col: attr_key})

                if attr_key not in df_proc.columns:
                    continue

                # 应用土地利用过滤
                df_filtered = filter_by_land_use(df_proc, attr_key)

                # 检查是否有有效数据
                vals = pd.to_numeric(df_filtered[attr_key], errors="coerce")
                vals = vals[vals > 0].dropna()
                if len(vals) == 0:
                    continue

                attr_name = SOIL_ATTR_CONFIG[attr_key]["name"]

                # 生成两个表
                ws1 = wb.create_sheet(title=f"{attr_name}面积(土地利用)")
                generate_land_use_area_by_level(ws1, df_filtered, attr_key)

                ws2 = wb.create_sheet(title=f"{attr_name}面积(乡镇)")
                generate_town_area_by_level(ws2, df_filtered, attr_key)

                processed_any = True

            except Exception:
                continue

        # 生成土壤质地统计表
        if "TRZD" in df_area.columns:
            try:
                if progress_callback:
                    progress_callback(90, "正在生成土壤质地统计表...")

                ws_texture_land = wb.create_sheet(title="土壤质地面积(土地利用)")
                generate_soil_texture_by_land_use(ws_texture_land, df_area)

                ws_texture_town = wb.create_sheet(title="土壤质地面积(乡镇)")
                generate_soil_texture_by_town(ws_texture_town, df_area)

                processed_any = True
            except Exception:
                pass

        if not processed_any:
            return False, "所有检测到的属性均无法处理（缺少数据或格式错误）"

        if progress_callback:
            progress_callback(95, "正在生成文件...")

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        if progress_callback:
            progress_callback(100, "完成！")

        return True, output.getvalue()

    except Exception as e:
        return False, str(e)
