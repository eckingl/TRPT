"""属性图数据处理模块

包含：机械组成统计、土壤属性统计、土壤质地统计
"""

import io
from collections.abc import Callable
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

from app.core.data import (
    get_land_use_class,
    get_pinyin_sort_key,
    load_multiple_csv,
    normalize_dlmc_column,
    normalize_soil_type_columns,
)
from app.topics.attribute_map.config import (
    SOIL_ATTR_CONFIG,
    SOIL_TEXTURE_MAPPING,
    TEXTURE_COLS,
    classify_series,
    classify_value,
    detect_available_attributes,
    get_grade_order,
    get_level_value_ranges,
)


def format_value(value: float, decimals: int = 3) -> float | str:
    """格式化数值

    Args:
        value: 数值
        decimals: 小数位数

    Returns:
        格式化后的值
    """
    if pd.isna(value):
        return value
    if value == 0:
        return 0
    abs_val = abs(value)
    if abs_val >= 0.001:
        return round(value, decimals)
    # 极小值使用科学计数法
    return f"{value:.3g}"


def format_percentage(value: float) -> float | str:
    """格式化百分比"""
    if pd.isna(value):
        return value
    if value > 100:
        return 100
    return format_value(value)


def apply_excel_styles(ws, max_row: int, max_col: int) -> None:
    """应用Excel样式

    Args:
        ws: 工作表
        max_row: 最大行数
        max_col: 最大列数
    """
    thin = Side(border_style="thin")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.border = border
            cell.alignment = center

    for col in range(1, max_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 12


# =============================================================================
# 土壤属性统计
# =============================================================================


def generate_overall_summary(
    ws, df_sample: pd.DataFrame, df_area: pd.DataFrame, attr_key: str
) -> None:
    """生成总体情况统计表

    Args:
        ws: Excel工作表
        df_sample: 样点数据
        df_area: 制图数据
        attr_key: 属性键名
    """
    config = SOIL_ATTR_CONFIG[attr_key]
    attr_name = config["name"]
    unit = config["unit"]
    ws.title = f"{attr_name}总体情况"

    grade_order = get_grade_order(attr_key)
    range_desc = get_level_value_ranges(attr_key)

    # 处理样点数据
    df_sample_clean = df_sample.copy()
    df_sample_clean[attr_key] = pd.to_numeric(
        df_sample_clean[attr_key], errors="coerce"
    )
    df_sample_clean = df_sample_clean[
        (df_sample_clean[attr_key] > 0) & df_sample_clean[attr_key].notna()
    ].copy()
    df_sample_clean["等级"] = df_sample_clean[attr_key].apply(
        lambda x: classify_value(x, attr_key)
    )

    # 处理制图数据
    df_area_clean = df_area.copy()
    df_area_clean[attr_key] = pd.to_numeric(df_area_clean[attr_key], errors="coerce")
    df_area_clean = df_area_clean[
        (df_area_clean[attr_key] > 0) & df_area_clean[attr_key].notna()
    ].copy()
    df_area_clean["等级"] = df_area_clean[attr_key].apply(
        lambda x: classify_value(x, attr_key)
    )

    # 样点统计
    total_samples = len(df_sample_clean)
    sample_counts = dict.fromkeys(grade_order, 0)
    for grade in df_sample_clean["等级"].dropna():
        if grade in sample_counts:
            sample_counts[grade] += 1

    sample_pct = {
        g: round(sample_counts[g] / total_samples * 100, 2) if total_samples > 0 else 0
        for g in grade_order
    }

    # 制图统计
    area_sum = dict.fromkeys(grade_order, 0.0)
    for _, row in df_area_clean.iterrows():
        grade = row["等级"]
        if pd.notna(grade) and grade in area_sum:
            area_val = row.get("面积", 0)
            if pd.notna(area_val):
                area_sum[grade] += float(area_val)

    total_area = sum(area_sum.values())
    area_pct = {
        g: round(area_sum[g] / total_area * 100, 2) if total_area > 0 else 0
        for g in grade_order
    }

    # 统计数值
    sample_mean = df_sample_clean[attr_key].mean() if total_samples > 0 else 0
    sample_median = df_sample_clean[attr_key].median() if total_samples > 0 else 0
    sample_min = df_sample_clean[attr_key].min() if total_samples > 0 else 0
    sample_max = df_sample_clean[attr_key].max() if total_samples > 0 else 0

    area_mean = df_area_clean[attr_key].mean() if len(df_area_clean) > 0 else 0
    area_median = df_area_clean[attr_key].median() if len(df_area_clean) > 0 else 0
    area_min = df_area_clean[attr_key].min() if len(df_area_clean) > 0 else 0
    area_max = df_area_clean[attr_key].max() if len(df_area_clean) > 0 else 0

    # 写入表头
    ws.merge_cells("A1:F1")
    ws["A1"] = f"土壤{attr_name}分级分布统计"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:B2")
    ws["A2"] = "土壤三普分级"
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("C2:D2")
    ws["C2"] = "样点统计"
    ws["C2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("E2:F2")
    ws["E2"] = "制图统计"
    ws["E2"].alignment = Alignment(horizontal="center", vertical="center")

    ws["A3"] = "分级"
    ws["B3"] = f"值域/({unit})" if unit else "值域"
    ws["C3"] = "数量/个"
    ws["D3"] = "占比/%"
    ws["E3"] = "面积/亩"
    ws["F3"] = "占比/%"

    # 写入数据
    for i, grade in enumerate(grade_order):
        row = 4 + i
        rng = range_desc[i] if i < len(range_desc) else ""
        ws.cell(row, 1, grade)
        ws.cell(row, 2, rng)
        ws.cell(row, 3, sample_counts.get(grade, 0))
        ws.cell(row, 4, format_percentage(sample_pct.get(grade, 0)))
        ws.cell(row, 5, format_value(area_sum.get(grade, 0)))
        ws.cell(row, 6, format_percentage(area_pct.get(grade, 0)))

    # 合计行
    summary_row = 4 + len(grade_order)
    ws.merge_cells(f"A{summary_row}:B{summary_row}")
    ws[f"A{summary_row}"] = "全区"
    ws[f"A{summary_row}"].alignment = Alignment(horizontal="center", vertical="center")
    ws[f"C{summary_row}"] = sum(sample_counts.values())
    ws[f"D{summary_row}"] = format_percentage(sum(sample_pct.values()))
    ws[f"E{summary_row}"] = format_value(sum(area_sum.values()))
    ws[f"F{summary_row}"] = format_percentage(sum(area_pct.values()))

    # 均值行
    stat_row = summary_row + 1
    ws.merge_cells(f"A{stat_row}:B{stat_row}")
    ws[f"A{stat_row}"] = f"全区均值/({unit})" if unit else "全区均值"
    ws[f"A{stat_row}"].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(f"C{stat_row}:D{stat_row}")
    ws[f"C{stat_row}"] = f"{sample_mean:.3f}"
    ws[f"C{stat_row}"].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(f"E{stat_row}:F{stat_row}")
    ws[f"E{stat_row}"] = f"{area_mean:.3f}"
    ws[f"E{stat_row}"].alignment = Alignment(horizontal="center", vertical="center")

    # 中位值行
    stat_row += 1
    ws.merge_cells(f"A{stat_row}:B{stat_row}")
    ws[f"A{stat_row}"] = f"全区中位值/({unit})" if unit else "全区中位值"
    ws[f"A{stat_row}"].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(f"C{stat_row}:D{stat_row}")
    ws[f"C{stat_row}"] = f"{sample_median:.3f}"
    ws[f"C{stat_row}"].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(f"E{stat_row}:F{stat_row}")
    ws[f"E{stat_row}"] = f"{area_median:.3f}"
    ws[f"E{stat_row}"].alignment = Alignment(horizontal="center", vertical="center")

    # 范围行
    stat_row += 1
    ws.merge_cells(f"A{stat_row}:B{stat_row}")
    ws[f"A{stat_row}"] = f"全区范围/({unit})" if unit else "全区范围"
    ws[f"A{stat_row}"].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(f"C{stat_row}:D{stat_row}")
    ws[f"C{stat_row}"] = f"{sample_min:.3f}～{sample_max:.3f}"
    ws[f"C{stat_row}"].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(f"E{stat_row}:F{stat_row}")
    ws[f"E{stat_row}"] = f"{area_min:.3f}～{area_max:.3f}"
    ws[f"E{stat_row}"].alignment = Alignment(horizontal="center", vertical="center")

    apply_excel_styles(ws, stat_row, 6)


def generate_land_use_summary(
    ws, df_sample: pd.DataFrame, df_area: pd.DataFrame, attr_key: str
) -> None:
    """生成土地利用类型统计表"""
    config = SOIL_ATTR_CONFIG[attr_key]
    attr_name = config["name"]
    unit = config["unit"]
    ws.title = f"{attr_name}土地利用"

    grade_order = get_grade_order(attr_key)
    range_desc = get_level_value_ranges(attr_key)

    # 预处理数据
    df_sample = df_sample.copy()
    df_area = df_area.copy()
    df_sample[attr_key] = pd.to_numeric(df_sample[attr_key], errors="coerce")
    df_area[attr_key] = pd.to_numeric(df_area[attr_key], errors="coerce")
    df_sample = df_sample[(df_sample[attr_key] > 0) & df_sample[attr_key].notna()]
    df_area = df_area[(df_area[attr_key] > 0) & df_area[attr_key].notna()]
    df_sample["等级"] = classify_series(df_sample[attr_key], attr_key)
    df_area["等级"] = classify_series(df_area[attr_key], attr_key)

    # 添加土地利用分类
    if "DLMC" in df_sample.columns:
        df_sample[["一级", "二级"]] = pd.DataFrame(
            df_sample["DLMC"].apply(get_land_use_class).tolist(), index=df_sample.index
        )
    if "DLMC" in df_area.columns:
        df_area[["一级", "二级"]] = pd.DataFrame(
            df_area["DLMC"].apply(get_land_use_class).tolist(), index=df_area.index
        )

    land_types = ["耕地", "园地", "林地", "草地", "其他"]

    # 写入表头
    ws.merge_cells(
        start_row=1, start_column=1, end_row=1, end_column=len(grade_order) + 4
    )
    ws["A1"] = f"土壤{attr_name}分级分布统计（按土地利用类型）"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    # 第二行表头
    ws["A2"] = "土地利用类型"
    ws["B2"] = "样点/个"
    ws["C2"] = "面积/亩"
    for i, grade in enumerate(grade_order):
        ws.cell(row=2, column=4 + i, value=grade)
    ws.cell(row=2, column=4 + len(grade_order), value="均值")

    for col in range(1, 5 + len(grade_order)):
        ws.cell(row=2, column=col).font = Font(bold=True)
        ws.cell(row=2, column=col).alignment = Alignment(
            horizontal="center", vertical="center"
        )

    current_row = 3
    for land_type in land_types:
        if "一级" in df_sample.columns:
            sample_subset = df_sample[df_sample["一级"] == land_type]
        else:
            sample_subset = pd.DataFrame()

        if "一级" in df_area.columns:
            area_subset = df_area[df_area["一级"] == land_type]
        else:
            area_subset = pd.DataFrame()

        sample_count = len(sample_subset)
        area_total = area_subset["面积"].sum() if "面积" in area_subset.columns else 0

        ws.cell(row=current_row, column=1, value=land_type)
        ws.cell(row=current_row, column=2, value=sample_count)
        ws.cell(row=current_row, column=3, value=format_value(area_total))

        # 各级别占比
        for i, grade in enumerate(grade_order):
            if len(area_subset) > 0 and area_total > 0:
                grade_area = area_subset[area_subset["等级"] == grade]["面积"].sum()
                pct = round(grade_area / area_total * 100, 2)
            else:
                pct = 0
            ws.cell(row=current_row, column=4 + i, value=format_percentage(pct))

        # 均值
        if len(sample_subset) > 0:
            mean_val = sample_subset[attr_key].mean()
            ws.cell(
                row=current_row, column=4 + len(grade_order), value=f"{mean_val:.3f}"
            )
        else:
            ws.cell(row=current_row, column=4 + len(grade_order), value="-")

        current_row += 1

    # 全区合计
    ws.cell(row=current_row, column=1, value="全区")
    ws.cell(row=current_row, column=2, value=len(df_sample))
    total_area = df_area["面积"].sum() if "面积" in df_area.columns else 0
    ws.cell(row=current_row, column=3, value=format_value(total_area))

    for i, grade in enumerate(grade_order):
        if len(df_area) > 0 and total_area > 0:
            grade_area = df_area[df_area["等级"] == grade]["面积"].sum()
            pct = round(grade_area / total_area * 100, 2)
        else:
            pct = 0
        ws.cell(row=current_row, column=4 + i, value=format_percentage(pct))

    if len(df_sample) > 0:
        mean_val = df_sample[attr_key].mean()
        ws.cell(row=current_row, column=4 + len(grade_order), value=f"{mean_val:.3f}")
    else:
        ws.cell(row=current_row, column=4 + len(grade_order), value="-")

    apply_excel_styles(ws, current_row, 4 + len(grade_order))


def generate_soil_type_summary(
    ws, df_sample: pd.DataFrame, df_area: pd.DataFrame, attr_key: str
) -> None:
    """生成分土壤类型统计表

    Args:
        ws: Excel工作表
        df_sample: 样点数据
        df_area: 制图数据
        attr_key: 属性键名
    """
    config = SOIL_ATTR_CONFIG[attr_key]
    attr_name = config["name"]
    unit = config["unit"]
    ws.title = f"{attr_name}分土壤类型"

    # 完全独立处理样点数据（只统计大于0的值）
    df_sample_clean = df_sample.copy()
    df_sample_clean[attr_key] = pd.to_numeric(
        df_sample_clean[attr_key], errors="coerce"
    )
    df_sample_clean = df_sample_clean[
        (df_sample_clean[attr_key] > 0) & df_sample_clean[attr_key].notna()
    ].copy()

    # 完全独立处理制图数据（只统计大于0的值）
    df_area_clean = df_area.copy()
    df_area_clean[attr_key] = pd.to_numeric(df_area_clean[attr_key], errors="coerce")
    df_area_clean = df_area_clean[
        (df_area_clean[attr_key] > 0) & df_area_clean[attr_key].notna()
    ].copy()

    # 添加标题行（第1行）
    ws.merge_cells("A1:I1")
    ws["A1"] = f"不同土壤类型{attr_name}统计"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    # 第2行：土壤类型标题
    ws.merge_cells("A2:B2")
    ws["A2"] = "土壤类型"
    ws["A2"].font = Font(bold=True, size=12)
    ws["A2"].alignment = Alignment(horizontal="center")

    ws.merge_cells("C2:F2")
    ws["C2"] = "样点统计"
    ws["C2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("G2:J2")
    ws["G2"] = "制图统计"
    ws["G2"].alignment = Alignment(horizontal="center", vertical="center")

    # 第3行：列标题，pH 不显示单位
    if attr_key == "ph":
        headers = [
            "亚类",
            "土属",
            "均值",
            "中位值",
            "范围",
            "数量/个",
            "均值",
            "面积/亩",
            "范围",
        ]
    else:
        headers = [
            "亚类",
            "土属",
            f"均值/({unit})",
            f"中位值/({unit})",
            f"范围/({unit})",
            "数量/个",
            f"均值/({unit})",
            "面积/亩",
            f"范围/({unit})",
        ]

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # 确保土壤类型字段存在且非空
    if "YL" not in df_sample_clean.columns or "TS" not in df_sample_clean.columns:
        # 没有土壤类型数据，只生成基本表头
        apply_excel_styles(ws, 3, 9)
        return

    df_sample_clean = df_sample_clean.dropna(subset=["YL", "TS"]).copy()
    df_area_clean = df_area_clean.dropna(subset=["YL", "TS"]).copy()

    for df in [df_sample_clean, df_area_clean]:
        df["YL"] = df["YL"].astype(str).str.strip()
        df["TS"] = df["TS"].astype(str).str.strip()

    if "面积" in df_area_clean.columns:
        df_area_clean["面积"] = pd.to_numeric(df_area_clean["面积"], errors="coerce")

    all_pairs = (
        pd.concat([df_sample_clean[["YL", "TS"]], df_area_clean[["YL", "TS"]]])
        .drop_duplicates()
        .reset_index(drop=True)
    )

    # 定义土壤类型精确排序顺序（亚类→土属列表）
    soil_type_order_map = {
        "棕红壤": ["红泥质棕红壤"],
        "红壤性土": ["砂泥质红壤性土", "麻砂质红壤性土"],
        "典型黄棕壤": [
            "暗泥质黄棕壤",
            "麻砂质黄棕壤",
            "红砂质黄棕壤",
            "黄土质黄棕壤",
            "砂泥质黄棕壤",
        ],
        "黄棕壤性土": ["砂泥质黄棕壤性土"],
        "典型棕壤": ["麻砂质典型棕壤"],
        "白浆化棕壤": ["麻砂质白浆化棕壤", "泥砂质白浆化棕壤"],
        "潮棕壤": ["泥砂质潮棕壤"],
        "淋溶褐土": ["黄土质淋溶褐土", "灰泥质淋溶褐土", "暗泥质淋溶褐土"],
        "潮褐土": ["泥砂质潮褐土"],
        "红黏土": ["红黏土"],
        "黑色石灰土": ["黑色石灰土"],
        "棕色石灰土": ["棕色石灰土"],
        "暗火山灰土": ["暗火山灰土"],
        "酸性紫色土": ["壤质酸性紫色土", "黏质酸性紫色土"],
        "中性紫色土": ["砂质中性紫色土", "壤质中性紫色土", "黏质中性紫色土"],
        "石灰性紫色土": ["壤质石灰性紫色土"],
        "酸性粗骨土": ["麻砂质酸性粗骨土", "硅质酸性粗骨土"],
        "中性粗骨土": ["麻砂质中性粗骨土"],
        "钙质粗骨土": ["灰泥质钙质粗骨土"],
        "典型潮土": ["砂质潮土", "壤质潮土", "黏质潮土"],
        "灰潮土": ["灰潮土", "石灰性灰潮土"],
        "盐化潮土（含碱化潮土）": ["氯化物盐化潮土", "硫酸盐盐化潮土", "苏打盐化潮土"],
        "典型砂姜黑土": ["黑腐砂姜黑土（黑姜土）", "覆泥砂姜黑土（覆泥黑姜土）"],
        "盐化砂姜黑土": ["氯化物盐化砂姜黑土"],
        "腐泥沼泽土": ["腐泥沼泽土"],
        "草甸沼泽土": ["草甸沼泽土", "石灰性草甸沼泽土"],
        "典型滨海盐土": ["氯化物滨海盐土"],
        "滨海沼泽盐土": ["氯化物沼泽滨海盐土"],
        "滨海潮滩盐土": ["氯化物潮滩滨海盐土"],
        "淹育水稻土": ["浅马肝泥田"],
        "渗育水稻土": [
            "渗灰泥田",
            "渗潮泥砂田",
            "渗潮泥田",
            "渗湖泥田",
            "渗涂泥田",
            "渗淡涂泥田",
            "渗麻砂泥田",
            "渗潮白土田",
            "渗马肝泥田",
        ],
        "潴育水稻土": ["潮泥田", "湖泥田", "马肝泥田"],
        "潜育水稻土": ["青湖泥田", "青马肝泥田"],
        "脱潜水稻土": ["黄斑黏田", "黄斑泥田"],
        "漂洗水稻土": ["漂潮白土田", "漂马肝泥田"],
        "盐渍水稻土": [
            "氯化物潮泥田",
            "氯化物涂泥田",
            "氯化物湖泥田",
            "硫酸盐潮泥田",
            "硫酸盐涂泥田",
            "苏打潮泥田",
            "苏打涂泥田",
            "苏打湖泥田",
        ],
        "填充土": ["工矿填充土", "城镇填充土"],
        "扰动土": ["运移扰动土"],
    }

    # 亚类顺序
    sub_order = list(soil_type_order_map.keys())

    # 创建排序映射
    sub_to_gens: dict[str, set | list] = {}
    for _, row in all_pairs.iterrows():
        sub, gen = row["YL"], row["TS"]
        if sub not in sub_to_gens:
            sub_to_gens[sub] = set()
        sub_to_gens[sub].add(gen)

    # 按照预定义顺序排序亚类，未在列表中的排到最后
    def get_soil_order(soil_name: str) -> int:
        try:
            return sub_order.index(soil_name)
        except ValueError:
            return len(sub_order)

    sorted_subs = sorted(sub_to_gens.keys(), key=get_soil_order)

    # 对每个亚类的土属按预定义顺序排序
    for sub in sorted_subs:
        gens_set = sub_to_gens[sub]
        if sub in soil_type_order_map:
            predefined = soil_type_order_map[sub]
            sorted_gens = [g for g in predefined if g in gens_set]
            sorted_gens.extend(sorted([g for g in gens_set if g not in predefined]))
            sub_to_gens[sub] = sorted_gens
        else:
            sub_to_gens[sub] = sorted(gens_set)

    current_row = 4

    for sub in sorted_subs:
        gens = sub_to_gens[sub]
        is_multi = len(gens) > 1

        start_row_for_sub = current_row

        all_vals_sample = pd.Series([], dtype="float64")
        all_vals_area = pd.Series([], dtype="float64")
        all_areas = pd.Series([], dtype="float64")

        for idx, gen in enumerate(gens):
            mask_s = (df_sample_clean["YL"] == sub) & (df_sample_clean["TS"] == gen)
            vals = df_sample_clean.loc[mask_s, attr_key]
            count = len(vals)
            mean_s = f"{vals.mean():.3f}" if count > 0 else ""
            median_s = f"{vals.median():.3f}" if count > 0 else ""
            range_s = f"{vals.min():.3f}～{vals.max():.3f}" if count > 0 else ""

            if count > 0:
                all_vals_sample = pd.concat(
                    [all_vals_sample, vals.astype("float64")], ignore_index=True
                )

            mask_a = (df_area_clean["YL"] == sub) & (df_area_clean["TS"] == gen)
            area_df = df_area_clean[mask_a]

            area_sum = 0.0
            mean_a = ""
            range_a = ""
            if not area_df.empty and "面积" in area_df.columns:
                area_col = area_df["面积"]
                val_col = area_df[attr_key]
                valid_mask = area_col.notna() & val_col.notna()
                area_valid = area_col[valid_mask]
                val_valid = val_col[valid_mask]

                if len(val_valid) > 0:
                    area_sum = area_valid.sum()
                    mean_a = f"{val_valid.mean():.3f}"
                    range_a = f"{val_valid.min():.3f}～{val_valid.max():.3f}"

                    all_vals_area = pd.concat(
                        [all_vals_area, val_valid.astype("float64")], ignore_index=True
                    )
                    all_areas = pd.concat(
                        [all_areas, area_valid.astype("float64")], ignore_index=True
                    )

            # 第一个土属显示亚类，其他不显示（留空）
            if idx == 0:
                ws.cell(row=current_row, column=1, value=sub)
            else:
                ws.cell(row=current_row, column=1, value="")

            ws.cell(row=current_row, column=2, value=gen)
            ws.cell(row=current_row, column=3, value=mean_s)
            ws.cell(row=current_row, column=4, value=median_s)
            ws.cell(row=current_row, column=5, value=range_s)
            ws.cell(row=current_row, column=6, value=count)
            ws.cell(row=current_row, column=7, value=mean_a)
            ws.cell(row=current_row, column=8, value=format_value(area_sum))
            ws.cell(row=current_row, column=9, value=range_a)

            current_row += 1

        if is_multi:
            count_all = len(all_vals_sample)
            mean_s_all = (
                f"{all_vals_sample.mean():.3f}" if len(all_vals_sample) > 0 else ""
            )
            median_s_all = (
                f"{all_vals_sample.median():.3f}" if len(all_vals_sample) > 0 else ""
            )
            range_s_all = (
                f"{all_vals_sample.min():.3f}～{all_vals_sample.max():.3f}"
                if len(all_vals_sample) > 0
                else ""
            )

            total_area_all = all_areas.sum()
            mean_a_all = f"{all_vals_area.mean():.3f}" if len(all_vals_area) > 0 else ""
            range_a_all = (
                f"{all_vals_area.min():.3f}～{all_vals_area.max():.3f}"
                if len(all_vals_area) > 0
                else ""
            )

            ws.cell(row=current_row, column=1, value="")
            ws.cell(row=current_row, column=2, value="合计")
            ws.cell(row=current_row, column=3, value=mean_s_all)
            ws.cell(row=current_row, column=4, value=median_s_all)
            ws.cell(row=current_row, column=5, value=range_s_all)
            ws.cell(row=current_row, column=6, value=count_all)
            ws.cell(row=current_row, column=7, value=mean_a_all)
            ws.cell(row=current_row, column=8, value=format_value(total_area_all))
            ws.cell(row=current_row, column=9, value=range_a_all)
            current_row += 1

            # 合并亚类单元格
            end_row_for_sub = current_row - 1
            if end_row_for_sub > start_row_for_sub:
                ws.merge_cells(f"A{start_row_for_sub}:A{end_row_for_sub}")
                ws.cell(row=start_row_for_sub, column=1).alignment = Alignment(
                    horizontal="center", vertical="center"
                )

    # "全区"行
    global_vals_sample = df_sample_clean[attr_key]
    count_global = len(global_vals_sample)
    mean_global = f"{global_vals_sample.mean():.3f}" if count_global > 0 else ""
    median_global = f"{global_vals_sample.median():.3f}" if count_global > 0 else ""
    range_global = (
        f"{global_vals_sample.min():.3f}～{global_vals_sample.max():.3f}"
        if count_global > 0
        else ""
    )

    area_global = df_area_clean["面积"].sum() if "面积" in df_area_clean.columns else 0
    vals_area_global = df_area_clean[attr_key]
    mean_area_global = (
        f"{vals_area_global.mean():.3f}" if len(vals_area_global) > 0 else ""
    )
    range_area_global = (
        f"{vals_area_global.min():.3f}～{vals_area_global.max():.3f}"
        if len(vals_area_global) > 0
        else ""
    )

    ws.merge_cells(f"A{current_row}:B{current_row}")
    ws.cell(row=current_row, column=1, value="全区")
    ws.cell(row=current_row, column=1).alignment = Alignment(
        horizontal="center", vertical="center"
    )
    ws.cell(row=current_row, column=3, value=mean_global)
    ws.cell(row=current_row, column=4, value=median_global)
    ws.cell(row=current_row, column=5, value=range_global)
    ws.cell(row=current_row, column=6, value=count_global)
    ws.cell(row=current_row, column=7, value=mean_area_global)
    ws.cell(row=current_row, column=8, value=format_value(area_global))
    ws.cell(row=current_row, column=9, value=range_area_global)

    # 应用样式
    thin = Side(border_style="thin")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    merged_cells = []
    for merge_range in ws.merged_cells.ranges:
        if merge_range.min_col == 1 and merge_range.max_col == 1:
            merged_cells.append((merge_range.min_row, merge_range.max_row))

    for row in ws.iter_rows(min_row=3, max_row=current_row, min_col=1, max_col=9):
        for cell in row:
            cell.border = border
            is_merged_start = any(
                cell.row == start_row and cell.column == 1
                for start_row, _ in merged_cells
            )
            if is_merged_start:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="center")

    for col in ["A", "B", "C", "D", "E", "F", "G", "H", "I"]:
        ws.column_dimensions[col].width = 12


def generate_town_summary(
    ws, df_sample: pd.DataFrame, df_area: pd.DataFrame, attr_key: str
) -> None:
    """生成乡镇统计表"""
    config = SOIL_ATTR_CONFIG[attr_key]
    attr_name = config["name"]
    ws.title = f"{attr_name}乡镇统计"

    grade_order = get_grade_order(attr_key)

    # 预处理数据
    df_sample = df_sample.copy()
    df_area = df_area.copy()
    df_sample[attr_key] = pd.to_numeric(df_sample[attr_key], errors="coerce")
    df_area[attr_key] = pd.to_numeric(df_area[attr_key], errors="coerce")
    df_sample = df_sample[(df_sample[attr_key] > 0) & df_sample[attr_key].notna()]
    df_area = df_area[(df_area[attr_key] > 0) & df_area[attr_key].notna()]
    df_sample["等级"] = classify_series(df_sample[attr_key], attr_key)
    df_area["等级"] = classify_series(df_area[attr_key], attr_key)

    # 获取乡镇列表（按拼音排序）
    towns = set()
    if "行政区名称" in df_sample.columns:
        towns.update(df_sample["行政区名称"].dropna().unique())
    if "行政区名称" in df_area.columns:
        towns.update(df_area["行政区名称"].dropna().unique())
    towns = sorted(towns, key=get_pinyin_sort_key)

    if not towns:
        towns = ["无数据"]

    # 写入表头
    ws.merge_cells(
        start_row=1, start_column=1, end_row=1, end_column=len(grade_order) + 4
    )
    ws["A1"] = f"土壤{attr_name}分级分布统计（分乡镇）"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws["A2"] = "乡镇"
    ws["B2"] = "样点/个"
    ws["C2"] = "面积/亩"
    for i, grade in enumerate(grade_order):
        ws.cell(row=2, column=4 + i, value=grade)
    ws.cell(row=2, column=4 + len(grade_order), value="均值")

    for col in range(1, 5 + len(grade_order)):
        ws.cell(row=2, column=col).font = Font(bold=True)
        ws.cell(row=2, column=col).alignment = Alignment(
            horizontal="center", vertical="center"
        )

    current_row = 3
    for town in towns:
        if "行政区名称" in df_sample.columns:
            sample_subset = df_sample[df_sample["行政区名称"] == town]
        else:
            sample_subset = pd.DataFrame()

        if "行政区名称" in df_area.columns:
            area_subset = df_area[df_area["行政区名称"] == town]
        else:
            area_subset = pd.DataFrame()

        sample_count = len(sample_subset)
        area_total = area_subset["面积"].sum() if "面积" in area_subset.columns else 0

        ws.cell(row=current_row, column=1, value=town)
        ws.cell(row=current_row, column=2, value=sample_count)
        ws.cell(row=current_row, column=3, value=format_value(area_total))

        for i, grade in enumerate(grade_order):
            if len(area_subset) > 0 and area_total > 0:
                grade_area = area_subset[area_subset["等级"] == grade]["面积"].sum()
                pct = round(grade_area / area_total * 100, 2)
            else:
                pct = 0
            ws.cell(row=current_row, column=4 + i, value=format_percentage(pct))

        if len(sample_subset) > 0:
            mean_val = sample_subset[attr_key].mean()
            ws.cell(
                row=current_row, column=4 + len(grade_order), value=f"{mean_val:.3f}"
            )
        else:
            ws.cell(row=current_row, column=4 + len(grade_order), value="-")

        current_row += 1

    # 全区合计
    ws.cell(row=current_row, column=1, value="全区")
    ws.cell(row=current_row, column=2, value=len(df_sample))
    total_area = df_area["面积"].sum() if "面积" in df_area.columns else 0
    ws.cell(row=current_row, column=3, value=format_value(total_area))

    for i, grade in enumerate(grade_order):
        if len(df_area) > 0 and total_area > 0:
            grade_area = df_area[df_area["等级"] == grade]["面积"].sum()
            pct = round(grade_area / total_area * 100, 2)
        else:
            pct = 0
        ws.cell(row=current_row, column=4 + i, value=format_percentage(pct))

    if len(df_sample) > 0:
        mean_val = df_sample[attr_key].mean()
        ws.cell(row=current_row, column=4 + len(grade_order), value=f"{mean_val:.3f}")
    else:
        ws.cell(row=current_row, column=4 + len(grade_order), value="-")

    apply_excel_styles(ws, current_row, 4 + len(grade_order))


# =============================================================================
# 土壤质地统计
# =============================================================================


def map_trzd_to_info(trzd_val) -> tuple[str | None, str | None, int | None]:
    """将TRZD值映射为(类别, 名称, 分级)"""
    if pd.isna(trzd_val):
        return None, None, None
    trzd_str = str(trzd_val).strip()
    if trzd_str in ["0", "/", ""]:
        return None, None, None
    if trzd_str in SOIL_TEXTURE_MAPPING:
        return SOIL_TEXTURE_MAPPING[trzd_str]
    return "其他", f"未知({trzd_str})", 99


def build_texture_row(
    sample_df: pd.DataFrame,
    area_df: pd.DataFrame,
    total_sample_count: int,
    total_area_sum: float,
) -> list:
    """构建质地行数据"""
    texture_counts = dict.fromkeys(TEXTURE_COLS, 0)
    texture_areas = dict.fromkeys(TEXTURE_COLS, 0.0)

    if not sample_df.empty and "质地名称" in sample_df.columns:
        counts_series = sample_df["质地名称"].value_counts()
        for name in TEXTURE_COLS:
            texture_counts[name] = counts_series.get(name, 0)

    if (
        not area_df.empty
        and "质地名称" in area_df.columns
        and "面积" in area_df.columns
    ):
        area_grouped = area_df.groupby("质地名称")["面积"].sum()
        for name in TEXTURE_COLS:
            texture_areas[name] = area_grouped.get(name, 0.0)

    sample_pcts = {}
    area_pcts = {}
    for name in TEXTURE_COLS:
        sample_pcts[name] = (
            round(texture_counts[name] / total_sample_count * 100, 3)
            if total_sample_count > 0
            else 0.0
        )
        area_pcts[name] = (
            round(texture_areas[name] / total_area_sum * 100, 3)
            if total_area_sum > 0
            else 0.0
        )

    sample_pct_list = [sample_pcts.get(name, 0.0) for name in TEXTURE_COLS]
    area_pct_list = [area_pcts.get(name, 0.0) for name in TEXTURE_COLS]

    return sample_pct_list + area_pct_list


def generate_texture_overall(
    ws, df_sample: pd.DataFrame, df_area: pd.DataFrame
) -> None:
    """生成土壤质地总体情况表"""
    ws.title = "土壤质地总体情况"

    # 映射TRZD
    def safe_map(x):
        if pd.isna(x) or str(x).strip() in {"0", "/", ""}:
            return (None, None, None)
        s = str(x).strip()
        return SOIL_TEXTURE_MAPPING.get(s, ("其他", f"未知({s})", 99))

    if "TRZD" in df_sample.columns:
        mapped = df_sample["TRZD"].apply(safe_map)
        df_sample[["质地类别", "质地名称", "分级"]] = pd.DataFrame(
            mapped.tolist(), index=df_sample.index
        )
        df_sample = df_sample.dropna(subset=["质地类别"]).copy()

    if "TRZD" in df_area.columns:
        mapped = df_area["TRZD"].apply(safe_map)
        df_area[["质地类别", "质地名称", "分级"]] = pd.DataFrame(
            mapped.tolist(), index=df_area.index
        )
        df_area = df_area.dropna(subset=["质地类别"]).copy()
        df_area["面积"] = pd.to_numeric(df_area["面积"], errors="coerce")
        df_area = df_area.dropna(subset=["面积"]).copy()

    total_sample_all = len(df_sample)
    total_area_all = df_area["面积"].sum() if "面积" in df_area.columns else 0

    # 数字转罗马数字映射
    grade_to_roman = {1: "Ⅰ", 2: "Ⅱ", 3: "Ⅲ", 4: "Ⅳ", 5: "Ⅴ", 6: "Ⅵ", 7: "Ⅶ"}

    # 写入表头
    ws.merge_cells("A1:G1")
    ws["A1"] = "土壤质地分级分布统计"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    headers = ["分级", "质地类别", "质地名称", "频数/个", "频率/%", "面积/亩", "比例/%"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=2, column=i, value=h)
        ws.cell(row=2, column=i).font = Font(bold=True)

    current_row = 3
    valid_grades = sorted(set(info[2] for info in SOIL_TEXTURE_MAPPING.values()))

    for grade in valid_grades:
        names_in_grade = [
            n for n, (_, _, g) in SOIL_TEXTURE_MAPPING.items() if g == grade
        ]
        cat_name = SOIL_TEXTURE_MAPPING[names_in_grade[0]][0]
        roman_grade = grade_to_roman.get(grade, str(grade))

        for name in names_in_grade:
            count = (
                len(df_sample[df_sample["质地名称"] == name])
                if "质地名称" in df_sample.columns
                else 0
            )
            freq = (
                round(count / total_sample_all * 100, 3)
                if total_sample_all > 0
                else 0.0
            )
            area_sum = (
                df_area.loc[df_area["质地名称"] == name, "面积"].sum()
                if "质地名称" in df_area.columns
                else 0
            )
            area_pct = (
                round(area_sum / total_area_all * 100, 3) if total_area_all > 0 else 0.0
            )

            ws.cell(row=current_row, column=1, value=roman_grade)
            ws.cell(row=current_row, column=2, value=cat_name)
            ws.cell(row=current_row, column=3, value=name)
            ws.cell(row=current_row, column=4, value=count)
            ws.cell(row=current_row, column=5, value=format_percentage(freq))
            ws.cell(row=current_row, column=6, value=format_value(area_sum))
            ws.cell(row=current_row, column=7, value=format_percentage(area_pct))
            current_row += 1

        # 合计行（仅对多个名称的分级）
        if len(names_in_grade) > 1:
            subset_s = (
                df_sample[df_sample["质地名称"].isin(names_in_grade)]
                if "质地名称" in df_sample.columns
                else pd.DataFrame()
            )
            subset_a = (
                df_area[df_area["质地名称"].isin(names_in_grade)]
                if "质地名称" in df_area.columns
                else pd.DataFrame()
            )
            count = len(subset_s)
            freq = (
                round(count / total_sample_all * 100, 3)
                if total_sample_all > 0
                else 0.0
            )
            area_sum = subset_a["面积"].sum() if "面积" in subset_a.columns else 0
            area_pct = (
                round(area_sum / total_area_all * 100, 3) if total_area_all > 0 else 0.0
            )

            ws.cell(row=current_row, column=1, value=roman_grade)
            ws.cell(row=current_row, column=2, value=cat_name)
            ws.cell(row=current_row, column=3, value="合计")
            ws.cell(row=current_row, column=4, value=count)
            ws.cell(row=current_row, column=5, value=format_percentage(freq))
            ws.cell(row=current_row, column=6, value=format_value(area_sum))
            ws.cell(row=current_row, column=7, value=format_percentage(area_pct))
            current_row += 1

    # 全市合计
    ws.cell(row=current_row, column=1, value="全市")
    ws.cell(row=current_row, column=2, value="全市")
    ws.cell(row=current_row, column=3, value="全市")
    ws.cell(row=current_row, column=4, value=total_sample_all)
    ws.cell(row=current_row, column=5, value=100.0)
    ws.cell(row=current_row, column=6, value=format_value(total_area_all))
    ws.cell(row=current_row, column=7, value=100.0)

    apply_excel_styles(ws, current_row, 7)


# =============================================================================
# 主处理函数
# =============================================================================


def process_attribute_data(
    sample_paths: list[str | Path],
    area_paths: list[str | Path],
    progress_callback: Callable[[int, str], None] | None = None,
) -> tuple[bool, bytes | str]:
    """处理属性图数据

    Args:
        sample_paths: 样点统计CSV文件路径列表
        area_paths: 制图统计CSV文件路径列表
        progress_callback: 进度回调函数

    Returns:
        (成功标志, Excel文件字节或错误信息)
    """
    try:
        if progress_callback:
            progress_callback(0, "正在读取样点数据...")

        # 读取数据
        df_sample = load_multiple_csv(sample_paths)

        if progress_callback:
            progress_callback(10, "正在读取制图数据...")

        df_area = load_multiple_csv(area_paths)

        if progress_callback:
            progress_callback(20, "正在标准化列名...")

        # 标准化列名
        try:
            df_sample = normalize_dlmc_column(df_sample)
        except ValueError:
            pass

        try:
            df_area = normalize_dlmc_column(df_area)
        except ValueError:
            pass

        try:
            df_sample = normalize_soil_type_columns(df_sample)
        except ValueError:
            pass

        try:
            df_area = normalize_soil_type_columns(df_area)
        except ValueError:
            pass

        if progress_callback:
            progress_callback(25, "正在检测可用属性...")

        # 检测可用属性
        available_attrs = detect_available_attributes(list(df_sample.columns))
        if not available_attrs:
            available_attrs = detect_available_attributes(list(df_area.columns))

        if progress_callback:
            progress_callback(30, f"检测到 {len(available_attrs)} 个可用属性")

        wb = Workbook()
        if wb.active:
            wb.remove(wb.active)

        processed_any = False
        total_attrs = len(available_attrs)

        for idx, (orig_col, attr_key) in enumerate(available_attrs):
            try:
                if progress_callback:
                    progress = 30 + int((idx / max(total_attrs, 1)) * 50)
                    attr_name = SOIL_ATTR_CONFIG.get(attr_key, {}).get("name", attr_key)
                    progress_callback(progress, f"正在处理: {attr_name}")

                # 确保列名标准化
                df_sample_proc = df_sample.copy()
                df_area_proc = df_area.copy()

                if orig_col in df_sample_proc.columns and orig_col != attr_key:
                    df_sample_proc = df_sample_proc.rename(columns={orig_col: attr_key})
                if orig_col in df_area_proc.columns and orig_col != attr_key:
                    df_area_proc = df_area_proc.rename(columns={orig_col: attr_key})

                if (
                    attr_key not in df_sample_proc.columns
                    and attr_key not in df_area_proc.columns
                ):
                    continue

                # 生成统计表
                attr_name = SOIL_ATTR_CONFIG[attr_key]["name"]

                ws1 = wb.create_sheet(title=f"{attr_name}总体")
                generate_overall_summary(ws1, df_sample_proc, df_area_proc, attr_key)

                ws2 = wb.create_sheet(title=f"{attr_name}土地利用")
                generate_land_use_summary(ws2, df_sample_proc, df_area_proc, attr_key)

                ws3 = wb.create_sheet(title=f"{attr_name}乡镇")
                generate_town_summary(ws3, df_sample_proc, df_area_proc, attr_key)

                ws4 = wb.create_sheet(title=f"{attr_name}土壤类型")
                generate_soil_type_summary(ws4, df_sample_proc, df_area_proc, attr_key)

                processed_any = True

            except Exception:
                continue

        # 处理土壤质地
        if "TRZD" in df_sample.columns or "TRZD" in df_area.columns:
            try:
                if progress_callback:
                    progress_callback(85, "正在处理土壤质地...")

                ws_texture = wb.create_sheet(title="土壤质地总体")
                generate_texture_overall(ws_texture, df_sample.copy(), df_area.copy())
                processed_any = True
            except Exception:
                pass

        if not processed_any:
            return False, "未能处理任何属性数据"

        if progress_callback:
            progress_callback(95, "正在生成文件...")

        # 保存到字节流
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        if progress_callback:
            progress_callback(100, "完成！")

        return True, output.getvalue()

    except Exception as e:
        return False, str(e)
